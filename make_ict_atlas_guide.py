#!/usr/bin/env python3
"""
ICT ATLAS EA Input Settings Guide — Excel Generator
Creates /home/user/Rattana/ICT_ATLAS_EA_Input_Guide.xlsx from scratch
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/home/user/Rattana/ICT_ATLAS_EA_Input_Guide.xlsx"

# ── Colours ──────────────────────────────────────────────────────────────────
FILL_TITLE      = PatternFill("solid", fgColor="B8860B")   # gold
FILL_SUBTITLE   = PatternFill("solid", fgColor="2D2D2D")   # dark gray
FILL_HEADER     = PatternFill("solid", fgColor="1E3A5F")   # dark blue
FILL_WARNING    = PatternFill("solid", fgColor="FFF3CD")   # yellow
FILL_SECTION    = PatternFill("solid", fgColor="2E4E7E")   # medium blue
FILL_EVEN       = PatternFill("solid", fgColor="F2F2F2")   # light gray
FILL_ODD        = PatternFill("solid", fgColor="FFFFFF")   # white

# ── Fonts ────────────────────────────────────────────────────────────────────
FONT_TITLE      = Font(bold=True,   color="FFFFFF", size=14)
FONT_SUBTITLE   = Font(italic=True, color="FFFFFF", size=9)
FONT_HEADER     = Font(bold=True,   color="FFFFFF", size=10)
FONT_WARNING    = Font(bold=True,   color="000000", size=9)
FONT_SECTION    = Font(bold=True,   color="FFFFFF", size=10)
FONT_DATA       = Font(color="000000", size=9)

# ── Alignment helpers ─────────────────────────────────────────────────────────
ALIGN_CENTER    = Alignment(horizontal="center", vertical="top",  wrap_text=True)
ALIGN_LEFT      = Alignment(horizontal="left",   vertical="top",  wrap_text=True)

# ── Border ───────────────────────────────────────────────────────────────────
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = {1: 22, 2: 55, 3: 12, 4: 12, 5: 12, 6: 12, 7: 50}

HEADERS = ["Input Name", "Description", "$50 acct", "$100 acct",
           "$1,000 acct", "$10,000 acct", "Notes / Tips"]

# ── All section data ──────────────────────────────────────────────────────────
SECTIONS = [
    ("  SYMBOL PRESET", [
        ("SymPreset", "Symbol preset — sets pip scale and symbol-specific logic",
         "SYM_XAUUSD", "SYM_XAUUSD", "SYM_XAUUSD", "SYM_XAUUSD",
         "Leave as XAUUSD. Other symbols not yet calibrated."),
    ]),
    ("  [01] BIAS ENGINE", [
        ("UseBiasEngine", "Require HTF bias alignment before any entry",
         "true", "true", "true", "true",
         "Always keep true — core ICT logic requires directional bias"),
        ("RequireWeeklyBias", "Weekly bias must match trade direction",
         "true", "true", "true", "true",
         "Weekly candle structure determines overall bias"),
        ("RequireDailyBias", "Daily bias must match trade direction",
         "true", "true", "true", "true",
         "Daily structure confirms intraday direction"),
        ("RequireH4Bias", "H4 bias must match trade direction",
         "false", "false", "false", "false",
         "Optional extra confirmation — adds strictness"),
        ("BiasSwingLookback", "Bars each side for swing point detection",
         "5", "5", "5", "5",
         "Higher = stronger swings required; 5 is balanced"),
    ]),
    ("  [02] LIQUIDITY ENGINE", [
        ("UseLiquidityEngine", "Require liquidity sweep before entry",
         "true", "true", "true", "true",
         "Core ICT concept — price must sweep a level first"),
        ("UseEQHEQL", "Detect Equal Highs / Equal Lows as sweep targets",
         "true", "true", "true", "true",
         "EQH/EQL are magnet levels for smart money"),
        ("EQHTolerance", "Pip tolerance for Equal High/Low detection",
         "5.0", "5.0", "5.0", "5.0",
         "In pips; 5 = levels within 5 pips are treated as equal"),
        ("LiqLookbackBars", "Bars back to scan for liquidity levels",
         "50", "50", "50", "50",
         "More bars = more levels identified"),
        ("SweepWickMinPips", "Minimum wick beyond level to qualify as sweep",
         "2", "2", "2", "2",
         "Prevents false sweeps; 2 pips minimum"),
    ]),
    ("  [03] MARKET STRUCTURE ENGINE", [
        ("UseMSSFilter", "MSS (Market Structure Shift) required after sweep",
         "true", "true", "true", "true",
         "MSS confirms directional flip — do not disable"),
        ("UseBOSFilter", "Also require Break of Structure confirmation",
         "false", "false", "false", "false",
         "BOS is broader confirmation; MSS is preferred"),
        ("MSSwingLookback", "Bars each side for MSS swing point detection",
         "3", "3", "3", "3",
         "3 is balanced — detects local swings on M15"),
        ("MSSLookbackBars", "Bars to look back for MSS level",
         "30", "30", "30", "30",
         "Scan last 30 bars for prior swing to break"),
    ]),
    ("  [04] DISPLACEMENT ENGINE", [
        ("UseDispFilter", "Require a strong displacement candle after MSS",
         "true", "true", "true", "true",
         "Displacement = aggressive push showing intent"),
        ("DispMinBodyPct", "Min body as % of candle range (0–1)",
         "0.60", "0.60", "0.60", "0.60",
         "0.60 = candle body must be 60% of its full range"),
        ("DispMinATRMulti", "Min candle range relative to ATR",
         "1.3", "1.3", "1.3", "1.3",
         "1.3 = candle must be 1.3× the average candle size"),
        ("DispLookbackBars", "Bars back to find displacement candle",
         "5", "5", "5", "5",
         "Scan last 5 bars; increase to 8 for more flexibility"),
    ]),
    ("  [05] FVG ENGINE", [
        ("UseFVGFilter", "Require price to retrace into a Fair Value Gap",
         "true", "true", "true", "true",
         "FVG = entry precision zone; always keep on"),
        ("UseFVGEntry", "Enter when price retraces into FVG",
         "true", "true", "true", "true",
         "Enables actual FVG entry logic"),
        ("UseIFVG", "Also detect Inverse FVGs (filled + rejected)",
         "true", "true", "true", "true",
         "IFVG = prior FVG became support/resistance"),
        ("UseCEentry", "Use CE (midpoint of FVG) as refined entry",
         "true", "true", "true", "true",
         "CE = 50% of FVG; improves RR"),
        ("FVGMaxAgeBars", "FVG expires after this many bars",
         "50", "50", "50", "50",
         "Old FVGs are less reliable; 50 bars ≈ 12.5 hours"),
        ("MaxFVGsTracked", "Max concurrent FVGs tracked in memory",
         "20", "20", "20", "20",
         "20 is sufficient for M15 XAUUSD"),
        ("UseFVGM5", "Also scan M5 chart FVGs for entry",
         "false", "false", "false", "false",
         "M5 FVGs add noise on M15 strategy; keep false"),
    ]),
    ("  [06] PD ARRAY ENGINE", [
        ("UsePDAEngine", "Enable PD Array selection for entry refinement",
         "true", "true", "true", "true",
         "PDA = ranked entry zone selection (ICT concept)"),
        ("PDA_UseFVG", "Include FVGs as PD Array entries",
         "true", "true", "true", "true",
         "FVG = most important PDA type"),
        ("PDA_UseIFVG", "Include Inverse FVGs as PD Array entries",
         "true", "true", "true", "true",
         "IFVG adds confluence"),
        ("PDA_UseOB", "Include Order Blocks as PD Array entries",
         "true", "true", "true", "true",
         "OB = candle before major move"),
        ("PDA_UseBreaker", "Include Breaker Blocks as PD Array entries",
         "true", "true", "true", "true",
         "Breaker = failed OB that flipped polarity"),
        ("PDA_UseMitigation", "Include Mitigation Blocks",
         "false", "false", "false", "false",
         "Experimental — keep false for stability"),
        ("PDA_UseLiqVoid", "Include Liquidity Voids",
         "false", "false", "false", "false",
         "Experimental — keep false for stability"),
        ("PDA_UseNDOG", "Include New Day Opening Gaps",
         "false", "false", "false", "false",
         "Gap-based; less reliable on M15"),
        ("PDA_UseNWOG", "Include New Week Opening Gaps",
         "false", "false", "false", "false",
         "Gap-based; less reliable on M15"),
        ("OBLookbackBars", "Bars back to find Order Blocks",
         "10", "10", "10", "10",
         "Scan 10 bars back for OB formation"),
        ("MaxPDAsTracked", "Max PD array objects tracked in memory",
         "15", "15", "15", "15",
         "15 is sufficient for M15 trading"),
    ]),
    ("  [07] PREMIUM / DISCOUNT ENGINE", [
        ("UsePremDiscFilter", "Require entry to be in premium/discount zone",
         "true", "true", "true", "true",
         "Buys in discount, sells in premium — core ICT rule"),
        ("DiscountZone", "Max Fibonacci % for discount zone (buys ≤ this)",
         "0.50", "0.50", "0.50", "0.50",
         "0.50 = lower 50% of dealing range is discount"),
        ("PremiumZone", "Min Fibonacci % for premium zone (sells ≥ 1-this)",
         "0.50", "0.50", "0.50", "0.50",
         "0.50 = upper 50% of dealing range is premium"),
        ("DealingRangeLookback", "H4 bars for dealing range calculation (60 ≈ 15 days)",
         "60", "60", "60", "60",
         "Higher = broader dealing range reference"),
    ]),
    ("  [08] SESSION & KILLZONE ENGINE", [
        ("UseSessionFilter", "Only trade during active killzone sessions",
         "true", "true", "true", "true",
         "ICT killzones = London open, NY open"),
        ("BrokerGMTOffset", "Your broker's GMT offset in hours",
         "0", "0", "0", "0",
         "Set if AutoGMTOffset doesn't detect correctly"),
        ("AutoGMTOffset", "Auto-detect broker GMT offset",
         "true", "true", "true", "true",
         "Leave true; EA detects offset automatically"),
        ("SessionAsian", "Trade Asian killzone session",
         "false", "false", "false", "false",
         "Asian = low volatility; not ideal for ICT setups"),
        ("SessionLondon", "Trade London killzone session",
         "true", "true", "true", "true",
         "London open = best ICT setups"),
        ("SessionNewYork", "Trade New York AM killzone session",
         "true", "true", "true", "true",
         "NY AM = highest volume and best moves"),
        ("SessionNYPM", "Trade NY PM session",
         "false", "false", "false", "false",
         "NY PM = low volume; avoid"),
        ("AsianStartHour", "Asian session start hour (GMT)",
         "0", "0", "0", "0",
         "Default: 00:00 GMT"),
        ("AsianEndHour", "Asian session end hour (GMT)",
         "7", "7", "7", "7",
         "Default: 07:00 GMT"),
        ("LondonStartHour", "London killzone start hour (GMT)",
         "7", "7", "7", "7",
         "London open: 07:00 GMT"),
        ("LondonEndHour", "London killzone end hour (GMT)",
         "10", "10", "10", "10",
         "London killzone ends 10:00 GMT"),
        ("NYStartHour", "New York AM killzone start hour (GMT)",
         "13", "13", "13", "13",
         "NY open: 13:00 GMT"),
        ("NYEndHour", "New York AM killzone end hour (GMT)",
         "16", "16", "16", "16",
         "NY AM ends 16:00 GMT"),
        ("NYPMStartHour", "NY PM session start hour (GMT)",
         "18", "18", "18", "18",
         "NY PM: 18:00 GMT"),
        ("NYPMEndHour", "NY PM session end hour (GMT)",
         "20", "20", "20", "20",
         "NY PM ends 20:00 GMT"),
    ]),
    ("  [09] POWER OF 3 ENGINE", [
        ("UsePO3Filter", "Require Power of 3 (Accumulation-Manipulation-Distribution) pattern",
         "false", "false", "false", "false",
         "Extra confluence; optional filter"),
        ("PO3ManipMinPips", "Minimum manipulation sweep size for PO3",
         "10.0", "10.0", "10.0", "10.0",
         "Sweep must be at least 10 pips for valid PO3"),
    ]),
    ("  [10] SMT DIVERGENCE ENGINE", [
        ("UseSMTFilter", "Enable SMT (Smart Money Technique) divergence filter",
         "false", "false", "false", "false",
         "Compares XAUUSD vs correlated asset for divergence"),
        ("SMTSymbol", "Correlated symbol for SMT comparison",
         "XAGUSD", "XAGUSD", "XAGUSD", "XAGUSD",
         "XAGUSD (Silver) diverges with Gold for SMT"),
        ("SMTLookbackBars", "Number of bars to compare for divergence",
         "5", "5", "5", "5",
         "Compare last 5 bars between symbols"),
    ]),
    ("  [11] ADR FILTER", [
        ("UseADRFilter", "Block entries when daily range is nearly complete",
         "true", "true", "true", "true",
         "Prevents buying tops and selling bottoms late in day"),
        ("ADRPeriod", "Days to calculate average daily range",
         "14", "14", "14", "14",
         "14-day ADR is the standard ICT reference"),
        ("ADRMaxPct", "Block trade if price has moved this % of ADR already",
         "0.80", "0.80", "0.80", "0.80",
         "0.80 = block if 80% of daily range already complete"),
    ]),
    ("  [12] NEWS FILTER", [
        ("UseNewsFilter", "Block trading around high-impact news events",
         "true", "true", "true", "true",
         "Avoids slippage and erratic moves during news"),
        ("NewsBlockBefore", "Minutes to block before news event",
         "30", "30", "30", "30",
         "Block 30 min before news to avoid manipulation"),
        ("NewsBlockAfter", "Minutes to block after news event",
         "30", "30", "30", "30",
         "Block 30 min after news to avoid fake moves"),
        ("NewsTime1", 'News event time 1 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         'e.g. "08:30" for 8:30 AM GMT news'),
        ("NewsTime2", 'News event time 2 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Add NFP, CPI, FOMC times here"),
        ("NewsTime3", 'News event time 3 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Add up to 8 news events per day"),
        ("NewsTime4", 'News event time 4 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Leave blank if fewer than 4 events"),
        ("NewsTime5", 'News event time 5 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Leave blank if fewer than 5 events"),
        ("NewsTime6", 'News event time 6 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Leave blank if fewer than 6 events"),
        ("NewsTime7", 'News event time 7 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Leave blank if fewer than 7 events"),
        ("NewsTime8", 'News event time 8 (format: HH:MM GMT)',
         '""', '""', '""', '""',
         "Leave blank if fewer than 8 events"),
    ]),
    ("  [13] MARKET CONDITION FILTER", [
        ("UseConditionFilter", "Filter by market regime (trend/range/choppy)",
         "true", "true", "true", "true",
         "Prevents trading in unfavorable market conditions"),
        ("TradeInTrend", "Allow trades in trending markets (ADX high)",
         "true", "true", "true", "true",
         "Trending = best ICT environment"),
        ("TradeInRanging", "Allow trades in ranging markets",
         "false", "false", "false", "false",
         "Ranging markets produce more false signals"),
        ("TradeInChoppy", "Allow trades in choppy/indecisive markets",
         "false", "false", "false", "false",
         "Choppy = worst environment for this strategy"),
        ("CondADXPeriod", "ADX indicator period for market regime detection",
         "14", "14", "14", "14",
         "Standard 14-period ADX"),
        ("CondADXTrend", "ADX value above which market is considered trending",
         "25.0", "25.0", "25.0", "25.0",
         "ADX > 25 = trending; standard threshold"),
        ("CondADXChoppy", "ADX value below which market is considered choppy",
         "18.0", "18.0", "18.0", "18.0",
         "ADX < 18 = choppy; block all entries"),
    ]),
    ("  [15-16] SPREAD & SLIPPAGE FILTERS", [
        ("MaxSpreadPips", "Maximum allowed spread in pips (0 = no check)",
         "50", "50", "30", "20",
         "XAUUSD normal spread ≈ 2-5 pips; 50 blocks extreme spikes"),
        ("MaxSlippagePips", "Maximum allowed slippage on order fill",
         "5", "5", "5", "3",
         "5 pips slippage tolerance on market orders"),
    ]),
    ("  [17] CORRELATION FILTER", [
        ("UseCorrelFilter", "Enable correlated symbol filter (e.g. DXY vs XAUUSD)",
         "false", "false", "false", "false",
         "Optional; DXY inversely correlated with Gold"),
        ("CorrelSymbol", "Symbol to use for correlation check",
         "DXY", "DXY", "DXY", "DXY",
         "DXY = US Dollar Index; inverse Gold correlation"),
    ]),
    ("  [18] CONFLUENCE SCORING SYSTEM", [
        ("UseScoringSystem", "Enable the confluence scoring gate",
         "true", "true", "true", "true",
         "Always on — it is the quality gate for every entry"),
        ("MinScore", "Minimum score (out of 125) required to enter trade",
         "70", "70", "80", "85",
         "Lower on small accounts for frequency; higher for quality"),
        ("ScoreWeeklyBias", "Points awarded for weekly bias alignment",
         "15", "15", "15", "15",
         "Max 15 pts; total = 125 pts across all components"),
        ("ScoreDailyBias", "Points awarded for daily bias alignment",
         "15", "15", "15", "15",
         "Max 15 pts for daily structure alignment"),
        ("ScoreLiqSweep", "Points awarded for liquidity sweep",
         "20", "20", "20", "20",
         "Max 20 pts; highest weight = most important"),
        ("ScoreMSS", "Points awarded for Market Structure Shift",
         "20", "20", "20", "20",
         "Max 20 pts; confirms direction change"),
        ("ScoreDisplacement", "Points awarded for displacement candle",
         "15", "15", "15", "15",
         "Max 15 pts; validates market intent"),
        ("ScoreFVG", "Points awarded for FVG entry zone",
         "10", "10", "10", "10",
         "Max 10 pts; entry precision"),
        ("ScoreKillzone", "Points awarded for active killzone",
         "10", "10", "10", "10",
         "Max 10 pts; session timing"),
        ("ScoreSMT", "Points awarded for SMT divergence",
         "5", "5", "5", "5",
         "Max 5 pts; optional filter"),
        ("ScoreADR", "Points awarded for ADR filter pass",
         "5", "5", "5", "5",
         "Max 5 pts; range position"),
        ("ScorePO3", "Points awarded for Power of 3 pattern",
         "5", "5", "5", "5",
         "Max 5 pts; AMD pattern"),
        ("ScorePremDisc", "Points awarded for premium/discount zone",
         "5", "5", "5", "5",
         "Max 5 pts; P/D zone alignment"),
    ]),
    ("  [19] TRADE QUALITY GRADES", [
        ("AllowedGrades", "Minimum grade allowed for entry",
         "A+ and A", "A+ and A", "A+ and A", "A+ and A",
         "Grades_A_UP = only A and A+ trades"),
        ("GradeAPlus", "Score threshold for A+ grade",
         "100", "100", "100", "100",
         "A+ = near-perfect confluence (score ≥ 100)"),
        ("GradeA", "Score threshold for A grade",
         "80", "80", "80", "80",
         "A = strong confluence (score ≥ 80)"),
        ("GradeB", "Score threshold for B grade",
         "60", "60", "60", "60",
         "B = acceptable confluence (score ≥ 60); not taken by default"),
    ]),
    ("  [25] RISK MANAGEMENT  ← Most important section", [
        ("RiskMode", "RISK_PCT = % of balance per trade; RISK_LOT = fixed lot",
         "RISK_LOT", "RISK_LOT", "RISK_PCT", "RISK_PCT",
         "Small accounts use RISK_LOT to avoid overleveraging"),
        ("RiskPercent", "Percentage of account balance to risk per trade",
         "N/A", "N/A", "1.0%", "0.5%",
         "Reduce to 0.5% when account grows over $5,000"),
        ("FixedLotSize", "Fixed lot size when using RISK_LOT mode",
         "0.01", "0.01", "0.01", "0.01",
         "Minimum lot; use 0.01 for safety on small accounts"),
        ("MaxLotCap", "Hard ceiling on lot size regardless of risk calc",
         "0.01", "0.02", "0.10", "0.50",
         "Prevents runaway lot sizes from compounding"),
        ("MaxTradesPerDay", "Maximum number of trades per day",
         "5", "5", "10", "10",
         "Limits daily exposure; increase with aggressive mode"),
        ("MaxConsecLosses", "Pause after this many consecutive losses",
         "3", "3", "5", "5",
         "Automatic circuit breaker on losing streaks"),
    ]),
    ("  [22] ADVANCED TRADE MANAGEMENT  (SL / TP)", [
        ("TP1_RR", "First take profit reward:risk ratio",
         "1.0", "1.0", "1.0", "1.0",
         "TP1 at 1R — take partial profit quickly"),
        ("TP2_RR", "Second take profit reward:risk ratio",
         "2.0", "2.0", "2.0", "2.0",
         "TP2 at 2R — mid-term target"),
        ("TP3_RR", "Third take profit / runner reward:risk ratio",
         "3.0", "3.0", "3.0", "3.0",
         "TP3 at 3R — runner position for extended move"),
        ("TP1_ClosePct", "Percentage of position to close at TP1",
         "40.0", "40.0", "40.0", "40.0",
         "Close 40% at TP1; locks in profit early"),
        ("TP2_ClosePct", "Percentage of position to close at TP2",
         "40.0", "40.0", "40.0", "40.0",
         "Close 40% at TP2; remaining 20% runs to TP3"),
        ("UseBreakeven", "Move SL to breakeven after TP1 is hit",
         "true", "true", "true", "true",
         "Protects capital once in profit"),
        ("BreakevenBufferPips", "Pips above entry to place breakeven SL",
         "5", "5", "5", "5",
         "5 pip buffer covers spread and commissions"),
        ("UseTrailingStop", "Enable trailing stop on the runner position",
         "false", "false", "false", "false",
         "Optional; can be activated for trending markets"),
        ("TrailStartPips", "Pips of profit before trailing stop activates",
         "50", "50", "50", "50",
         "Start trailing after 50 pips in profit"),
        ("TrailStepPips", "How many pips the trail moves per step",
         "15", "15", "15", "15",
         "15 pip step; tighter = more responsive"),
        ("SLBufferPips", "Extra pips added beyond structure for SL",
         "10", "10", "10", "10",
         "Buffer for spread and wicks; 10 pips standard"),
        ("MaxSLPips", "Maximum allowed SL size in pips",
         "80", "80", "80", "80",
         "Rejects setups with SL wider than 80 pips"),
        ("MinSLPips", "Minimum required SL size in pips",
         "15", "15", "15", "15",
         "Rejects setups with SL tighter than 15 pips"),
    ]),
    ("  [23] DAILY PROFIT LOCK", [
        ("UseProfitLock", "Enable automatic profit protection for the day",
         "true", "true", "true", "true",
         "Locks gains after reaching profit target"),
        ("ProfitLock_ReduceR", "Reduce position size after +XR daily profit",
         "3.0", "3.0", "3.0", "3.0",
         "At 3R daily profit, reduce risk by ReducePct%"),
        ("ProfitLock_StopR", "Stop trading after +XR daily profit",
         "5.0", "5.0", "5.0", "5.0",
         "At 5R daily profit, no more trades for the day"),
        ("ProfitLock_ReducePct", "Reduce risk by this % when ProfitLock_ReduceR hit",
         "50.0", "50.0", "50.0", "50.0",
         "Trade at 50% of normal risk after hitting reduce level"),
    ]),
    ("  [24] DAILY LOSS PROTECTION  ← Turn OFF for backtesting", [
        ("UseLossProtect", "Enable automatic loss protection for the day",
         "true", "true", "true", "true",
         "Critical for risk control; turn OFF for backtesting"),
        ("LossProtect_ReduceR", "Reduce size after -XR daily loss",
         "2.0", "2.0", "2.0", "2.0",
         "At 2R daily loss, reduce risk by 50%"),
        ("LossProtect_StopR", "Stop all trading after -XR daily loss",
         "3.0", "3.0", "3.0", "3.0",
         "At 3R daily loss, no more trades for the day"),
        ("MaxWeeklyLossR", "Maximum weekly loss in R before stopping",
         "6.0", "6.0", "6.0", "6.0",
         "Weekly circuit breaker; 6R max weekly drawdown"),
    ]),
    ("  TRADE CLOSE RULES", [
        ("CloseOnFriday", "Automatically close all open trades on Friday",
         "true", "true", "true", "true",
         "Avoids weekend gap risk on XAUUSD"),
        ("FridayCloseHour", "GMT hour on Friday to close all trades",
         "14", "14", "14", "14",
         "14:00 GMT = NY open Friday; good liquidity exit"),
        ("CooldownMinutes", "Minimum minutes between two trade entries",
         "30", "30", "15", "15",
         "Prevents over-trading; 15 min in active sessions"),
    ]),
    ("  [26] VISUAL CHART TOOLS", [
        ("DrawPDH_PDL", "Draw Previous Day High and Low lines on chart",
         "true", "true", "true", "true",
         "Key daily reference levels for liquidity"),
        ("DrawPWH_PWL", "Draw Previous Week High and Low lines on chart",
         "true", "true", "true", "true",
         "Weekly reference levels for major sweeps"),
        ("DrawSessionRanges", "Draw Asian and London session range boxes",
         "true", "true", "true", "true",
         "Shows ICT dealing range visually"),
        ("DrawFVGZones", "Highlight Fair Value Gap zones on chart",
         "true", "true", "true", "true",
         "Visualizes entry zones"),
        ("DrawOBZones", "Highlight Order Block zones on chart",
         "true", "true", "true", "true",
         "Visualizes OB entry zones"),
        ("DrawLiqSweeps", "Mark detected liquidity sweep points",
         "true", "true", "true", "true",
         "Shows where sweeps occurred"),
        ("DrawMSSLines", "Mark Market Structure Shift levels",
         "true", "true", "true", "true",
         "Shows MSS confirmation levels"),
        ("DrawTargets", "Show projected TP1/TP2/TP3 target lines",
         "true", "true", "true", "true",
         "Visualizes reward targets on chart"),
        ("ColorPDH", "Color for Previous Day High line",
         "Lime", "Lime", "Lime", "Lime",
         "Default: Lime green"),
        ("ColorPDL", "Color for Previous Day Low line",
         "Tomato", "Tomato", "Tomato", "Tomato",
         "Default: Tomato red"),
        ("ColorPWH", "Color for Previous Week High line",
         "Aqua", "Aqua", "Aqua", "Aqua",
         "Default: Aqua blue"),
        ("ColorPWL", "Color for Previous Week Low line",
         "Orange", "Orange", "Orange", "Orange",
         "Default: Orange"),
        ("ColorFVGBull", "Fill color for Bullish FVG zones",
         "0,80,0", "0,80,0", "0,80,0", "0,80,0",
         "Dark green fill for bullish FVGs"),
        ("ColorFVGBear", "Fill color for Bearish FVG zones",
         "80,0,0", "80,0,0", "80,0,0", "80,0,0",
         "Dark red fill for bearish FVGs"),
        ("ColorOBBull", "Fill color for Bullish Order Block zones",
         "0,50,100", "0,50,100", "0,50,100", "0,50,100",
         "Dark blue fill for bullish OBs"),
        ("ColorOBBear", "Fill color for Bearish Order Block zones",
         "80,30,0", "80,30,0", "80,30,0", "80,30,0",
         "Brown fill for bearish OBs"),
    ]),
    ("  [27] DEBUG PANEL", [
        ("ShowPanel", "Show the ICT ATLAS debug panel on chart",
         "true", "true", "true", "true",
         "Highly recommended — shows real-time EA state"),
        ("PanelX", "Panel horizontal position from left edge (pixels)",
         "12", "12", "12", "12",
         "Adjust if panel overlaps chart elements"),
        ("PanelY", "Panel vertical position from top edge (pixels)",
         "30", "30", "30", "30",
         "Adjust if panel overlaps chart elements"),
        ("ShowStatPanel", "Show statistics section at bottom of panel",
         "true", "true", "true", "true",
         "Tracks daily P/L, win rate, trade count"),
        ("DebugLogs", "Print detailed FVG and filter decisions to Experts log",
         "false", "false", "false", "false",
         "Enable only when diagnosing why trades are not firing"),
    ]),
    ("  [29] AGGRESSIVE MODE  ← Enables high-frequency ICT trading", [
        ("AggressiveMode", "Master switch for high-frequency / relaxed-filter mode",
         "false", "false", "false", "true",
         "Enable for 15-30 trades/day. Relaxes all secondary filters."),
        ("AggrMinScore", "Minimum confluence score in aggressive mode",
         "55", "55", "65", "55",
         "Lower than MinScore — accepts more setups in aggressive"),
        ("AggrMSSOptional", "MSS gives half-credit (not hard required) in aggressive",
         "true", "true", "true", "true",
         "Removes hard MSS block; many valid setups lack textbook MSS"),
        ("AggrDispOptional", "Displacement is optional (0 pts, no fail) in aggressive",
         "false", "false", "false", "false",
         "Keep false — displacement is still a useful quality filter"),
        ("AggrDispMinBodyPct", "Relaxed body % threshold for displacement in aggressive",
         "0.40", "0.40", "0.50", "0.40",
         "Conservative = 0.60; Aggressive = 0.40 (accepts more candles)"),
        ("AggrDispMinATRMulti", "Relaxed ATR multiplier for displacement in aggressive",
         "0.60", "0.60", "0.80", "0.60",
         "Conservative = 1.3×ATR; Aggressive = 0.6×ATR"),
        ("AggrRequireFVG", "Require FVG entry zone in aggressive mode",
         "false", "false", "false", "false",
         "Set false in aggressive — allows entry without FVG retrace"),
        ("AggrStrictPremDisc", "Enforce Premium/Discount zone in aggressive mode",
         "false", "false", "false", "false",
         "Conservative = enforce P/D; Aggressive = treat as bonus points"),
        ("AggrADRMaxPct", "Relaxed ADR completion cap in aggressive mode",
         "1.20", "1.20", "1.00", "1.20",
         "1.20 = entry allowed even if price moved 120% of normal ADR"),
        ("AggrStrictKillzone", "Require active killzone in aggressive mode",
         "false", "false", "false", "false",
         "Set false to allow entries outside London/NY windows"),
        ("AllowContinuation", "Allow trades without a sweep when bias aligned",
         "true", "true", "true", "true",
         "Continuation trades: bias-aligned moves without a fresh sweep"),
        ("ExpandKillzones", "Expand all session windows by ±2 hours",
         "false", "false", "false", "true",
         "More time in-session = more valid entries in aggressive"),
        ("ScalperMode", "Ultra-aggressive scalper mode (further relaxed thresholds)",
         "false", "false", "false", "false",
         "Experimental. Uses ScalperMinScore=40; bypasses most filters"),
        ("ScalperMinScore", "Minimum score when ScalperMode=true",
         "40", "40", "40", "40",
         "Only active when ScalperMode=true; very low threshold"),
    ]),
]

SUMMARY_HEADERS = ["Input Name", "Description", "$50 acct", "$100 acct",
                   "$1,000 acct", "$10,000 acct", "Notes"]

SUMMARY_DATA = [
    ("RiskMode", "Risk calculation method", "RISK_LOT", "RISK_LOT", "RISK_PCT", "RISK_PCT", ""),
    ("RiskPercent / FixedLot", "Risk per trade", "0.01 lot", "0.01 lot", "1.0%", "0.5%", ""),
    ("MaxLotCap", "Lot size ceiling", "0.01", "0.02", "0.10", "0.50", ""),
    ("MaxSLPips", "Maximum stop loss size", "80 pips", "80 pips", "80 pips", "80 pips", ""),
    ("MinSLPips", "Minimum stop loss size", "15 pips", "15 pips", "15 pips", "15 pips", ""),
    ("LossProtect_StopR", "Daily loss limit in R", "3R", "3R", "3R", "3R", ""),
    ("MaxWeeklyLossR", "Weekly loss limit in R", "6R", "6R", "6R", "6R", ""),
    ("MinScore", "Minimum confluence score", "70", "70", "80", "85", ""),
    ("MaxTradesPerDay", "Max trades per day", "5", "5", "10", "10", ""),
    ("AggressiveMode", "High-frequency mode", "false", "false", "false", "optional", ""),
]


def apply_cell(cell, value="", fill=None, font=None, alignment=None, border=True):
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = ALIGN_LEFT
    if border:
        cell.border = BORDER


def style_row_cells(ws, row_num, fill, font=None):
    """Apply fill/font/border to all 7 cells in a data row."""
    for col in range(1, 8):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER


def write_section_header(ws, row_num, text):
    ws.merge_cells(f"A{row_num}:G{row_num}")
    cell = ws.cell(row=row_num, column=1)
    apply_cell(cell, value=text, fill=FILL_SECTION, font=FONT_SECTION,
               alignment=ALIGN_CENTER)
    # Apply border to merged cells
    for col in range(2, 8):
        c = ws.cell(row=row_num, column=col)
        c.fill = FILL_SECTION
        c.font = FONT_SECTION
        c.alignment = ALIGN_CENTER
        c.border = BORDER


def write_data_row(ws, row_num, row_data, is_even):
    fill = FILL_EVEN if is_even else FILL_ODD
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        apply_cell(cell, value=value, fill=fill, font=FONT_DATA, alignment=ALIGN_LEFT)


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ICT ATLAS EA Guide"

    # ── Set column widths ────────────────────────────────────────────────────
    for col_num, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1)
    apply_cell(title_cell,
               value="ICT ATLAS EA V1.0  —  Input Settings Guide by Account Size",
               fill=FILL_TITLE, font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 22
    for col in range(2, 8):
        c = ws.cell(row=1, column=col)
        c.fill = FILL_TITLE; c.border = BORDER

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    sub_cell = ws.cell(row=2, column=1)
    apply_cell(sub_cell,
               value="Created by: RATTANA CHHORM  |  Symbol: XAUUSD M15  |  Timeframe: M15",
               fill=FILL_SUBTITLE, font=FONT_SUBTITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[2].height = 16
    for col in range(2, 8):
        c = ws.cell(row=2, column=col)
        c.fill = FILL_SUBTITLE; c.border = BORDER

    # ── Row 3: Empty ──────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Row 4: Headers ────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        apply_cell(cell, value=header, fill=FILL_HEADER, font=FONT_HEADER,
                   alignment=ALIGN_CENTER)

    # ── Row 5: Warning ────────────────────────────────────────────────────────
    ws.merge_cells("A5:G5")
    warn_cell = ws.cell(row=5, column=1)
    apply_cell(warn_cell,
               value="  ⚠  XAUUSD minimum viable account: ~$500.  "
                     "With aggressive mode enabled, use $200+ with 0.01 fixed lot.  "
                     "Risk % mode requires at least $1,000.",
               fill=FILL_WARNING, font=FONT_WARNING, alignment=ALIGN_LEFT)
    ws.row_dimensions[5].height = 20
    for col in range(2, 8):
        c = ws.cell(row=5, column=col)
        c.fill = FILL_WARNING; c.border = BORDER

    # ── Row 6: Empty ──────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 8

    # ── Sections ──────────────────────────────────────────────────────────────
    current_row = 7
    data_row_counter = 0  # global counter for even/odd alternation

    for section_name, rows in SECTIONS:
        # Section header
        write_section_header(ws, current_row, section_name)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Data rows
        for row_data in rows:
            data_row_counter += 1
            is_even = (data_row_counter % 2 == 0)
            write_data_row(ws, current_row, list(row_data), is_even)
            ws.row_dimensions[current_row].height = 30
            current_row += 1

    # ── Empty row before summary ──────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    # ── Summary section header ─────────────────────────────────────────────────
    write_section_header(ws, current_row,
                         "  QUICK SUMMARY TABLE — Recommended values by account size")
    ws.row_dimensions[current_row].height = 16
    current_row += 1

    # ── Summary header row ────────────────────────────────────────────────────
    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        apply_cell(cell, value=header, fill=FILL_HEADER, font=FONT_HEADER,
                   alignment=ALIGN_CENTER)
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # ── Summary data rows ─────────────────────────────────────────────────────
    for idx, row_data in enumerate(SUMMARY_DATA, start=1):
        is_even = (idx % 2 == 0)
        write_data_row(ws, current_row, list(row_data), is_even)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_PATH)
    total_rows = ws.max_row
    print(f"File saved: {OUTPUT_PATH}")
    print(f"Total rows written: {total_rows}")
    print(f"Sections written: {len(SECTIONS)}")

    # Quick verification
    wb2 = openpyxl.load_workbook(OUTPUT_PATH)
    ws2 = wb2.active
    print(f"Verification — reloaded max_row: {ws2.max_row}, max_col: {ws2.max_column}")
    wb2.close()


if __name__ == "__main__":
    main()

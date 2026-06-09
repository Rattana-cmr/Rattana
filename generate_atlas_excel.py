#!/usr/bin/env python3
"""Generate ICT_ATLAS_EA_Input_Guide.xlsx — ALL sections, dark-theme."""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
BG_DARK   = "0D0F1A"
BG_PANEL  = "12141F"
BG_HDR    = "1C1F33"
BG_COL    = "0A2240"
BG_GREEN  = "0A1F0A"

FG_WHITE  = "FFFFFF"
FG_GOLD   = "FFD700"
FG_GREEN  = "00E676"
FG_CYAN   = "00E5FF"
FG_GRAY   = "AAAAAA"
FG_SILVER = "DDDDDD"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=FG_WHITE, bold=False, size=10, italic=False):
    return Font(name="Consolas", color=color, bold=bold, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color="444466"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

ws_guide = wb.active
ws_guide.title = "Input Guide"
ws_acct  = wb.create_sheet("Account Profiles")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INPUT GUIDE
# ══════════════════════════════════════════════════════════════════════════════

ws = ws_guide
ws.sheet_view.showGridLines = False

col_widths = [2, 30, 16, 16, 48, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def wc(ws, row, col, value, bg, fg, bold=False, size=10, h="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(fg, bold, size, italic)
    c.alignment = align(h, "center", wrap); c.border = thin_border()
    return c

def fill_row(ws, row, bg, cols=6):
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        try:
            if c.value is None: c.value = ""
        except AttributeError:
            pass  # merged cell
        c.fill = fill(bg); c.border = thin_border()

def section_hdr(ws, row, section_id, title):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2, value=f"◆  {section_id}  {title}")
    c.fill = fill(BG_HDR); c.font = font(FG_CYAN, bold=True, size=10)
    c.alignment = align("left"); c.border = thin_border()
    fill_row(ws, row, BG_HDR)

def col_hdr(ws, row):
    ws.row_dimensions[row].height = 18
    headers = ["", "Input Parameter", "Type", "Default", "Description / ICT Context", "Recommendation"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(BG_COL); c.font = font(FG_GOLD, bold=True, size=9)
        c.alignment = align("center"); c.border = thin_border()

def param_row(ws, row, name, ptype, default, desc, rec="", highlight=False):
    ws.row_dimensions[row].height = 32
    bg = "2A2000" if highlight else BG_PANEL
    bg_r = BG_GREEN if rec and rec not in ("—", "") else BG_PANEL
    wc(ws, row, 1, "", BG_DARK, FG_WHITE)
    wc(ws, row, 2, name,    bg, FG_SILVER, size=9)
    wc(ws, row, 3, ptype,   bg, FG_CYAN,   size=9, h="center")
    wc(ws, row, 4, default, bg, FG_GOLD,   size=9, h="center")
    wc(ws, row, 5, desc,    bg, FG_SILVER, size=8, wrap=True)
    wc(ws, row, 6, rec,     bg_r, FG_GREEN, size=9, wrap=True)

def spacer(ws, row, h=6):
    ws.row_dimensions[row].height = h
    for col in range(1, 7):
        c = ws.cell(row=row, column=col, value="")
        c.fill = fill(BG_DARK); c.border = thin_border()

r = 1

# Title
ws.row_dimensions[r].height = 40
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="  ICT ATLAS EA V1.0  —  Complete Input Parameter Guide  (All Sections)")
c.fill = fill("1A1030"); c.font = font(FG_GOLD, bold=True, size=14)
c.alignment = align("center"); c.border = thin_border()
r += 1

ws.row_dimensions[r].height = 18
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="  Created by RATTANA CHHORM  ·  XAUUSD M15  ·  ICT 2022 Mentorship  ·  Sections: SYMBOL PRESET → [28] DEBUG & RELAXED MODE")
c.fill = fill("12101E"); c.font = font(FG_GRAY, italic=True, size=9)
c.alignment = align("center"); c.border = thin_border()
r += 1

spacer(ws, r); r += 1
col_hdr(ws, r); r += 1

# ══ SYMBOL PRESET ════════════════════════════════════════
section_hdr(ws, r, "SYMBOL PRESET", "Auto-configure for each instrument"); r += 1
param_row(ws, r, "SymPreset", "Enum", "SYM_XAUUSD",
    "Selects pip factor, spread thresholds, and SL defaults for the symbol.\n"
    "SYM_AUTO = detect by broker digits  |  SYM_XAUUSD = Gold (2 decimal places)\n"
    "SYM_EURUSD/GBPUSD/USDJPY = major FX  |  SYM_BTCUSD = crypto\n"
    "Set this FIRST before adjusting any other inputs.",
    "SYM_XAUUSD for Gold trading", highlight=True); r += 1
spacer(ws, r); r += 1

# ══ [01] BIAS ENGINE ════════════════════════════════════
section_hdr(ws, r, "[01]", "BIAS ENGINE — Higher-Timeframe Narrative Filter"); r += 1
param_row(ws, r, "UseBiasEngine", "bool", "true",
    "Master switch. When ON the EA checks Weekly/Daily/H4 bias before any trade.\n"
    "ICT Rule: only trade WITH the higher-timeframe institutional narrative.\n"
    "Disabling produces more trades but far lower quality.",
    "true — always ON"); r += 1
param_row(ws, r, "RequireWeeklyBias", "bool", "true",
    "Require Weekly chart bias to match trade direction.\n"
    "Weekly is the most important timeframe for ICT — defines the primary draw on liquidity.",
    "true — most critical filter"); r += 1
param_row(ws, r, "RequireDailyBias", "bool", "true",
    "Require Daily bias to match. Daily is the active institutional narrative.\n"
    "NEUTRAL daily can still get partial score if H4 confirms.",
    "true — keep ON"); r += 1
param_row(ws, r, "RequireH4Bias", "bool", "false",
    "Also require H4 bias match. Adds extra precision but reduces trade count.\n"
    "Good for conservative large accounts.",
    "false = default. Set true for $10k+ high-precision mode"); r += 1
param_row(ws, r, "BiasSwingLookback", "int", "5",
    "Bars each side used to detect swing highs/lows for bias calculation.\n"
    "Higher value = fewer but more significant structure points.",
    "5 bars standard. Range: 3–10"); r += 1
spacer(ws, r); r += 1

# ══ [02] LIQUIDITY ENGINE ════════════════════════════════
section_hdr(ws, r, "[02]", "LIQUIDITY ENGINE — Sweep Confirmation"); r += 1
param_row(ws, r, "UseLiquidityEngine", "bool", "true",
    "ICT Core Concept: price MUST sweep a liquidity pool (stop hunt) before reversal entry.\n"
    "Targets: PDH, PDL, PWH, PWL, EQH, EQL, Asian range highs/lows.\n"
    "Disabling this removes the foundational ICT premise.",
    "true — never disable", highlight=True); r += 1
param_row(ws, r, "UseEQHEQL", "bool", "true",
    "Detect Equal Highs (EQH) and Equal Lows (EQL) as liquidity targets.\n"
    "ICT: double tops/bottoms = resting stops that institutions will target.",
    "true — Gold forms many EQH/EQL patterns"); r += 1
param_row(ws, r, "EQHTolerance", "double/pips", "5.0",
    "Max distance in pips between two highs/lows to be considered 'equal'.\n"
    "Too small misses EQH. Too large creates false positives.",
    "5 pips for XAUUSD. Range: 3–8"); r += 1
param_row(ws, r, "LiqLookbackBars", "int", "50",
    "How many M15 bars back to scan for liquidity levels.\n"
    "50 bars ≈ 12.5 hours of price history on M15.",
    "50 bars. Range: 30–100"); r += 1
param_row(ws, r, "SweepWickMinPips", "int", "2",
    "Minimum wick extension beyond the level to confirm a sweep.\n"
    "Prevents triggers from candles that merely touch the level.",
    "2 pips for XAUUSD M15"); r += 1
spacer(ws, r); r += 1

# ══ [03] MARKET STRUCTURE ENGINE ════════════════════════
section_hdr(ws, r, "[03]", "MARKET STRUCTURE ENGINE — MSS / BOS Detection"); r += 1
param_row(ws, r, "UseMSSFilter", "bool", "true",
    "MSS = Market Structure Shift. After a sweep, price must close above/below\n"
    "the nearest internal swing high/low to confirm a reversal.\n"
    "This is THE ICT entry trigger — sweep → MSS → entry.",
    "true — core signal, always ON", highlight=True); r += 1
param_row(ws, r, "UseBOSFilter", "bool", "false",
    "BOS = Break of Structure (trend continuation in same direction).\n"
    "Enabling allows trend-following trades in addition to reversals.",
    "false for pure ICT reversal model"); r += 1
param_row(ws, r, "MSSwingLookback", "int", "3",
    "Bars each side to identify the swing pivot used in MSS detection.\n"
    "Smaller = more sensitive, larger = only major structure only.",
    "3 bars for M15 balanced sensitivity"); r += 1
param_row(ws, r, "MSSLookbackBars", "int", "30",
    "How many bars back to look for the MSS swing level.\n"
    "Must be larger than typical sweep-to-MSS distance.",
    "30 bars covers most M15 ICT setups"); r += 1
spacer(ws, r); r += 1

# ══ [04] DISPLACEMENT ENGINE ════════════════════════════
section_hdr(ws, r, "[04]", "DISPLACEMENT ENGINE — Institutional Move Confirmation"); r += 1
param_row(ws, r, "UseDispFilter", "bool", "true",
    "Displacement = strong, fast directional move (1–3 candles) creating an FVG.\n"
    "ICT: displacement MUST follow the MSS to confirm institutional participation.\n"
    "Without displacement = weak setup, likely retail-driven.",
    "true — confirms institutions are driving", highlight=True); r += 1
param_row(ws, r, "DispMinBodyPct", "double", "0.60",
    "Minimum candle body as fraction of total range (0–1).\n"
    "0.60 = body must be ≥60% of high-low range.\n"
    "Filters doji and indecision candles — requires decisive move.",
    "0.60 standard. Range: 0.50–0.75"); r += 1
param_row(ws, r, "DispMinATRMulti", "double", "1.3",
    "Displacement candle range must exceed X × ATR(14).\n"
    "Ensures the candle is larger than average — true displacement.",
    "1.3 for XAUUSD M15. Range: 1.0–2.0"); r += 1
param_row(ws, r, "DispLookbackBars", "int", "5",
    "How many bars back to find the displacement candle.\n"
    "Displacement usually occurs within 1–3 bars of the MSS.",
    "5 bars. Range: 3–10"); r += 1
spacer(ws, r); r += 1

# ══ [05] FVG ENGINE ══════════════════════════════════════
section_hdr(ws, r, "[05]", "FVG ENGINE — Fair Value Gap Entry Zones"); r += 1
param_row(ws, r, "UseFVGFilter", "bool", "true",
    "FVG = 3-candle imbalance where price moved too fast, leaving an unfilled gap.\n"
    "ICT: institutions re-visit FVGs to fill remaining orders.\n"
    "The FVG zone is your entry zone — price returning into it = entry trigger.",
    "true — primary ICT entry model", highlight=True); r += 1
param_row(ws, r, "UseFVGEntry", "bool", "true",
    "Allow entries when price retraces into a valid unmitigated FVG.\n"
    "Entry is triggered at FVG zone (bottom/top) or CE midpoint.",
    "true — activates FVG entry trigger"); r += 1
param_row(ws, r, "UseIFVG", "bool", "true",
    "IFVG = Inverse FVG. A filled bullish FVG that flips bearish (or vice versa).\n"
    "ICT advanced concept for secondary refined entry targets.",
    "true — adds extra entry precision"); r += 1
param_row(ws, r, "UseCEentry", "bool", "true",
    "CE = Consequent Encroachment (50% midpoint of the FVG).\n"
    "More precise than the full FVG — price often taps exactly the CE before reversing.",
    "true — tighter entries, better RR"); r += 1
param_row(ws, r, "FVGMaxAgeBars", "int", "50",
    "Max age in bars before an FVG is considered stale and removed.\n"
    "Old FVGs lose their institutional significance.",
    "50 bars ≈ 12.5 hours on M15. Range: 20–100"); r += 1
param_row(ws, r, "MaxFVGsTracked", "int", "20",
    "Maximum active FVGs tracked simultaneously.",
    "20 standard. Increase to 30 for busy markets"); r += 1
param_row(ws, r, "UseFVGM5", "bool", "false",
    "Also scan the M5 chart for FVGs. Ultra-precise entries when M15 setup is confirmed.",
    "false default. Enable for precision scalping entries"); r += 1
spacer(ws, r); r += 1

# ══ [06] PD ARRAY ENGINE ═════════════════════════════════
section_hdr(ws, r, "[06]", "PD ARRAY ENGINE — Order Zones Tracking"); r += 1
param_row(ws, r, "UsePDAEngine", "bool", "true",
    "Master switch for PD Array detection.\n"
    "PD Arrays = all ICT entry zone types: FVG, OB, Breaker, etc.",
    "true — required for zone entries"); r += 1
param_row(ws, r, "PDA_UseFVG", "bool", "true",
    "Include Fair Value Gaps as PD Arrays.", "true — primary entry zones"); r += 1
param_row(ws, r, "PDA_UseIFVG", "bool", "true",
    "Include Inverse FVGs (filled FVGs that flip polarity).", "true"); r += 1
param_row(ws, r, "PDA_UseOB", "bool", "true",
    "Include Order Blocks (last opposing candle before displacement).\n"
    "OBs represent where institutions placed their orders.",
    "true — classic ICT entry zone"); r += 1
param_row(ws, r, "PDA_UseBreaker", "bool", "true",
    "Include Breaker Blocks (former OBs that have been broken and flipped).\n"
    "High-probability — former support becomes resistance and vice versa.",
    "true — adds A+ setups"); r += 1
param_row(ws, r, "PDA_UseMitigation", "bool", "false",
    "Include Mitigation Blocks (areas with remaining unmitigated orders).",
    "false default"); r += 1
param_row(ws, r, "PDA_UseLiqVoid", "bool", "false",
    "Include Liquidity Voids (thin areas price will revisit quickly).",
    "false default"); r += 1
param_row(ws, r, "PDA_UseNDOG", "bool", "false",
    "NDOG = New Day Opening Gap. Price often fills these gaps same day.",
    "false default. Useful on assets with overnight gaps"); r += 1
param_row(ws, r, "PDA_UseNWOG", "bool", "false",
    "NWOG = New Week Opening Gap. Gold often has weekend gaps.",
    "false default. Enable for Gold Monday opens"); r += 1
param_row(ws, r, "OBLookbackBars", "int", "10",
    "How many bars back to scan for Order Block formation.",
    "10 bars standard. Range: 5–20"); r += 1
param_row(ws, r, "MaxPDAsTracked", "int", "15",
    "Maximum total PD Arrays tracked (FVG + OB + Breaker combined).",
    "15 standard. Increase to 25 for busy markets"); r += 1
spacer(ws, r); r += 1

# ══ [07] PREMIUM/DISCOUNT ════════════════════════════════
section_hdr(ws, r, "[07]", "PREMIUM / DISCOUNT ENGINE — Zone Filter"); r += 1
param_row(ws, r, "UsePremDiscFilter", "bool", "true",
    "ICT Core: only BUY from discount (<50% of dealing range),\n"
    "only SELL from premium (>50%). Disabling allows entries anywhere.",
    "true — critical ICT rule", highlight=True); r += 1
param_row(ws, r, "DiscountZone", "double (0–1)", "0.50",
    "Maximum % position in dealing range for BUY entries.\n"
    "0.50 = must be in lower half. 0.40 = stricter (lower 40%).",
    "0.50 standard ICT"); r += 1
param_row(ws, r, "PremiumZone", "double (0–1)", "0.50",
    "Minimum % from top for SELL entries.\n"
    "Works symmetrically with DiscountZone.",
    "0.50 standard ICT"); r += 1
param_row(ws, r, "DealingRangeLookback", "int (H4 bars)", "60",
    "H4 bars used to define the high/low anchor for the dealing range.\n"
    "60 H4 bars ≈ 10 trading days.",
    "60 bars ≈ 2 weeks. Range: 20–100"); r += 1
spacer(ws, r); r += 1

# ══ [08] SESSION & KILLZONE ══════════════════════════════
section_hdr(ws, r, "[08]", "SESSION & KILLZONE ENGINE — Trade Timing"); r += 1
param_row(ws, r, "UseSessionFilter", "bool", "true",
    "Only trade during killzones — the 1–3 hour windows when institutions are active.\n"
    "Outside killzones = noise trading. This filter is essential for ICT.",
    "true — always ON", highlight=True); r += 1
param_row(ws, r, "BrokerGMTOffset", "int", "0",
    "Manual GMT offset for broker server time. Ignored when AutoGMTOffset=true.",
    "0 if AutoGMTOffset=true (recommended)"); r += 1
param_row(ws, r, "AutoGMTOffset", "bool", "true",
    "Automatically detect GMT offset from broker server clock. Handles DST changes.",
    "true — always use auto-detect"); r += 1
param_row(ws, r, "SessionAsian", "bool", "false",
    "Trade Asian killzone (00:00–07:00 GMT). Lower liquidity, more false sweeps.",
    "false for XAUUSD. Enable for JPY pairs"); r += 1
param_row(ws, r, "SessionLondon", "bool", "true",
    "Trade London killzone (07:00–10:00 GMT). High liquidity, strong Gold trends.",
    "true — prime killzone for Gold", highlight=True); r += 1
param_row(ws, r, "SessionNewYork", "bool", "true",
    "Trade NY AM killzone (13:00–16:00 GMT). Highest volume for Gold.",
    "true — most important session for XAUUSD", highlight=True); r += 1
param_row(ws, r, "SessionNYPM", "bool", "false",
    "Trade NY PM session (18:00–20:00 GMT). Low volume, consolidation risk.",
    "false for XAUUSD"); r += 1
param_row(ws, r, "AsianStartHour", "int (GMT)", "0",
    "Asian killzone start in GMT.", "0 = midnight GMT"); r += 1
param_row(ws, r, "AsianEndHour", "int (GMT)", "7",
    "Asian killzone end in GMT.", "7 = 7AM GMT"); r += 1
param_row(ws, r, "LondonStartHour", "int (GMT)", "7",
    "London killzone start in GMT.", "7 = London open"); r += 1
param_row(ws, r, "LondonEndHour", "int (GMT)", "10",
    "London killzone end in GMT.", "10 = 3-hour window"); r += 1
param_row(ws, r, "NYStartHour", "int (GMT)", "13",
    "New York AM killzone start (13:00 GMT = 8:00 AM EST).", "13 = NY open"); r += 1
param_row(ws, r, "NYEndHour", "int (GMT)", "16",
    "New York AM killzone end.", "16 = covers NY morning"); r += 1
param_row(ws, r, "NYPMStartHour", "int (GMT)", "18",
    "NY PM session start.", "18"); r += 1
param_row(ws, r, "NYPMEndHour", "int (GMT)", "20",
    "NY PM session end.", "20"); r += 1
spacer(ws, r); r += 1

# ══ [09] POWER OF 3 ══════════════════════════════════════
section_hdr(ws, r, "[09]", "POWER OF 3 ENGINE — AMD Pattern Confirmation"); r += 1
param_row(ws, r, "UsePO3Filter", "bool", "false",
    "PO3 = Accumulation → Manipulation (fake sweep) → Distribution (real move).\n"
    "Enable for extra confluence — confirms the AMD cycle is complete.",
    "false default. Enable for additional confirmation"); r += 1
param_row(ws, r, "PO3ManipMinPips", "double/pips", "10.0",
    "Minimum size of the manipulation (fake) sweep in pips.\n"
    "Smaller sweeps may not represent true PO3 manipulation phase.",
    "10 pips for XAUUSD M15. Range: 5–20"); r += 1
spacer(ws, r); r += 1

# ══ [10] SMT DIVERGENCE ══════════════════════════════════
section_hdr(ws, r, "[10]", "SMT DIVERGENCE ENGINE — Correlated Asset Check"); r += 1
param_row(ws, r, "UseSMTFilter", "bool", "false",
    "SMT = when XAUUSD makes a new swing high but XAGUSD fails to confirm,\n"
    "it signals a divergence — the move is likely to reverse.\n"
    "Powerful filter but requires reliable correlated symbol in broker.",
    "false default. Enable if XAGUSD is available"); r += 1
param_row(ws, r, "SMTSymbol", "string", "XAGUSD",
    "The correlated symbol for comparison.\n"
    "XAUUSD ↔ XAGUSD is the classic Gold/Silver pair for SMT.",
    "XAGUSD for Gold. Must be in Market Watch"); r += 1
param_row(ws, r, "SMTLookbackBars", "int", "5",
    "Bars to compare for the SMT swing divergence check.",
    "5 bars standard"); r += 1
spacer(ws, r); r += 1

# ══ [11] ADR FILTER ══════════════════════════════════════
section_hdr(ws, r, "[11]", "ADR FILTER — Daily Range Exhaustion Block"); r += 1
param_row(ws, r, "UseADRFilter", "bool", "true",
    "Block entries when today's move has already used most of the ADR.\n"
    "Prevents entering late when daily range is near completion.",
    "true — prevents poor entries late in the day"); r += 1
param_row(ws, r, "ADRPeriod", "int (days)", "14",
    "Days to calculate the average daily range.",
    "14 days. Range: 10–20"); r += 1
param_row(ws, r, "ADRMaxPct", "double (0–1)", "0.80",
    "Block when today's range exceeds this fraction of ADR.\n"
    "0.80 = stop new trades when 80% of average daily move is used.",
    "0.80 for XAUUSD. Use 0.70 for conservative mode"); r += 1
spacer(ws, r); r += 1

# ══ [12] NEWS FILTER ══════════════════════════════════════
section_hdr(ws, r, "[12]", "NEWS FILTER — High-Impact Event Blackout"); r += 1
param_row(ws, r, "UseNewsFilter", "bool", "true",
    "Block trading around high-impact events.\n"
    "CRITICAL for Gold: NFP, CPI, FOMC, Fed speeches cause extreme volatility.",
    "true — essential for Gold trading", highlight=True); r += 1
param_row(ws, r, "NewsBlockBefore", "int (minutes)", "30",
    "Minutes before news to halt new entries.",
    "30 min. Use 60 for FOMC/NFP"); r += 1
param_row(ws, r, "NewsBlockAfter", "int (minutes)", "30",
    "Minutes after news to resume trading.\n"
    "Price often makes false moves immediately post-news.",
    "30 min standard. Use 60 for major events"); r += 1
param_row(ws, r, "NewsTime1–8", "string HH:MM", "(blank)",
    "Enter high-impact news times in HH:MM GMT format.\n"
    "Common: NFP=13:30  CPI=13:30  FOMC=19:00  Fed=varies\n"
    "Check Forex Factory every week for the schedule.\n"
    "Up to 8 events can be blocked per day.",
    "Fill daily from Forex Factory calendar", highlight=True); r += 1
spacer(ws, r); r += 1

# ══ [13] MARKET CONDITION ════════════════════════════════
section_hdr(ws, r, "[13]", "MARKET CONDITION FILTER — Trend / Range / Choppy"); r += 1
param_row(ws, r, "UseConditionFilter", "bool", "true",
    "ADX-based market condition detection.\n"
    "Prevents trading in choppy/ranging markets where ICT setups fail more.",
    "true — improves win rate"); r += 1
param_row(ws, r, "TradeInTrend", "bool", "true",
    "Allow trades when ADX shows trending (ADX > CondADXTrend).\n"
    "ICT setups work best in trending environments.",
    "true — ideal ICT condition"); r += 1
param_row(ws, r, "TradeInRanging", "bool", "false",
    "Allow trades in ranging markets (ADX between choppy and trend thresholds).",
    "false default. Enable carefully with tight TP1"); r += 1
param_row(ws, r, "TradeInChoppy", "bool", "false",
    "Allow trades in choppy conditions (ADX < CondADXChoppy). Very risky.",
    "false — keep OFF always"); r += 1
param_row(ws, r, "CondADXPeriod", "int", "14",
    "ADX period for trend strength.", "14 standard"); r += 1
param_row(ws, r, "CondADXTrend", "double", "25.0",
    "ADX above this = trending market.", "25.0 classic threshold"); r += 1
param_row(ws, r, "CondADXChoppy", "double", "18.0",
    "ADX below this = choppy (avoid).", "18.0 conservative chop filter"); r += 1
spacer(ws, r); r += 1

# ══ [15-16] SPREAD & SLIPPAGE ════════════════════════════
section_hdr(ws, r, "[15-16]", "SPREAD & SLIPPAGE FILTERS"); r += 1
param_row(ws, r, "MaxSpreadPips", "int (pips)", "50",
    "Block entries when spread exceeds this threshold.\n"
    "XAUUSD normal spread: 2–5 pips. News spikes: 30–100 pips.\n"
    "0 = no spread check (not recommended).",
    "50 pips permissive. Use 30 for tighter control"); r += 1
param_row(ws, r, "MaxSlippagePips", "int (pips)", "5",
    "Maximum order execution slippage. If fill deviates more, order is rejected.",
    "5 pips standard. 10 during news"); r += 1
spacer(ws, r); r += 1

# ══ [17] CORRELATION FILTER ══════════════════════════════
section_hdr(ws, r, "[17]", "CORRELATION FILTER — Macro Confirmation"); r += 1
param_row(ws, r, "UseCorrelFilter", "bool", "false",
    "Check a correlated symbol for directional alignment.\n"
    "Gold is inversely correlated with DXY (Dollar Index).",
    "false default. Enable for advanced macro confirmation"); r += 1
param_row(ws, r, "CorrelSymbol", "string", "DXY",
    "Symbol to check for correlation.\n"
    "DXY = Dollar Index (primary Gold macro driver). Must be in Market Watch.",
    "DXY for Gold trading"); r += 1
spacer(ws, r); r += 1

# ══ [18] CONFLUENCE SCORING ══════════════════════════════
section_hdr(ws, r, "[18]", "CONFLUENCE SCORING SYSTEM — Quality Gate (Max 125 pts)"); r += 1
param_row(ws, r, "UseScoringSystem", "bool", "true",
    "Master switch. When ON each setup is scored out of 125 points.\n"
    "Only setups above MinScore are taken. The quality gate separates good from great.",
    "true — the most important filter", highlight=True); r += 1
param_row(ws, r, "MinScore", "int (0–125)", "80",
    "Minimum confluence score to enter a trade.\n"
    "80/125 = A-grade setups only. 60 = B-grade allowed. 50 = relaxed mode.",
    "80 standard. 90+ for aggressive quality filter"); r += 1
param_row(ws, r, "ScoreWeeklyBias", "int", "15",
    "Points for Weekly bias alignment. Highest weight = most important timeframe.",
    "15 — major structural weight"); r += 1
param_row(ws, r, "ScoreDailyBias", "int", "15",
    "Points for Daily bias alignment.", "15 — equal to Weekly"); r += 1
param_row(ws, r, "ScoreLiqSweep", "int", "20",
    "Points for confirmed liquidity sweep. Highest single score = core ICT trigger.",
    "20 — the primary trigger"); r += 1
param_row(ws, r, "ScoreMSS", "int", "20",
    "Points for MSS confirmation.", "20 — equal weight to sweep"); r += 1
param_row(ws, r, "ScoreDisplacement", "int", "15",
    "Points for displacement candle.", "15 — institutional confirmation"); r += 1
param_row(ws, r, "ScoreFVG", "int", "10",
    "Points for entry inside an FVG zone.", "10 — entry precision"); r += 1
param_row(ws, r, "ScoreKillzone", "int", "10",
    "Points for being inside an active killzone.", "10 — timing filter"); r += 1
param_row(ws, r, "ScoreSMT", "int", "5",
    "Bonus points for SMT divergence confirmation.", "5 — optional bonus"); r += 1
param_row(ws, r, "ScoreADR", "int", "5",
    "Points when ADR range is not exhausted.", "5 — room-to-move bonus"); r += 1
param_row(ws, r, "ScorePO3", "int", "5",
    "Bonus points for PO3 pattern completion.", "5 — AMD bonus"); r += 1
param_row(ws, r, "ScorePremDisc", "int", "5",
    "Points for being in premium (sell) or discount (buy) zone.", "5 — zone bonus"); r += 1
spacer(ws, r); r += 1

# ══ [19] TRADE QUALITY GRADES ════════════════════════════
section_hdr(ws, r, "[19]", "TRADE QUALITY GRADES — A+, A, B Filter"); r += 1
param_row(ws, r, "AllowedGrades", "Enum", "GRADES_A_UP",
    "Sets the minimum grade for trade execution.\n"
    "GRADES_APLUS = A+ only (score ≥ 100)  |  GRADES_A_UP = A+ and A (≥80)\n"
    "GRADES_B_UP = A+, A, B (≥60)  |  GRADES_ALL = all grades",
    "GRADES_A_UP — balanced quality/frequency", highlight=True); r += 1
param_row(ws, r, "GradeAPlus", "int", "100",
    "Score threshold for A+ grade (elite setup — nearly perfect confluence).",
    "100"); r += 1
param_row(ws, r, "GradeA", "int", "80",
    "Score threshold for A grade (standard high-quality ICT setup).",
    "80"); r += 1
param_row(ws, r, "GradeB", "int", "60",
    "Score threshold for B grade (partial confluence — use only in GRADES_B_UP).",
    "60"); r += 1
spacer(ws, r); r += 1

# ══ [25] RISK MANAGEMENT ═════════════════════════════════
section_hdr(ws, r, "[25]", "RISK MANAGEMENT — Lot Sizing & Trade Limits"); r += 1
param_row(ws, r, "RiskMode", "Enum", "RISK_PCT",
    "RISK_PCT = risk fixed % of account per trade (scales with growth).\n"
    "RISK_LOT = fixed lot size regardless of SL distance.",
    "RISK_PCT — always recommended", highlight=True); r += 1
param_row(ws, r, "RiskPercent", "double (%)", "0.5",
    "Percentage of balance to risk per trade.\n"
    "$50 → 0.5% = $0.25  |  $100 → 0.5% = $0.50  |  $1k → 0.5% = $5\n"
    "Lower % = longer survival during drawdown. Proven edge first, then scale.",
    "0.5% conservative start. Scale after 50+ profitable trades"); r += 1
param_row(ws, r, "FixedLotSize", "double", "0.01",
    "Fixed lot when RiskMode = RISK_LOT. 0.01 = micro lot = 1,000 units.",
    "0.01 for small accounts in RISK_LOT mode"); r += 1
param_row(ws, r, "MaxLotCap", "double", "0.50",
    "Hard ceiling on lot size — prevents oversizing on large accounts or wide SL.",
    "0.50 small/mid accounts. Scale to 2.0+ for $10k+"); r += 1
param_row(ws, r, "MaxTradesPerDay", "int", "10",
    "Maximum trades per calendar day. Prevents overtrading after a winning streak.",
    "10 permissive. Use 3–5 for conservative daily cap"); r += 1
param_row(ws, r, "MaxConsecLosses", "int", "5",
    "Pause EA after this many consecutive losses. Emotional circuit breaker.",
    "5 standard. Use 3 for aggressive protection"); r += 1
spacer(ws, r); r += 1

# ══ [22] ADVANCED TRADE MANAGEMENT ═══════════════════════
section_hdr(ws, r, "[22]", "ADVANCED TRADE MANAGEMENT — TP / SL / Trailing Stop"); r += 1
param_row(ws, r, "TP1_RR", "double (R)", "1.0",
    "Take Profit 1 in multiples of risk. 1.0R = close at 1:1 (safety lock partial).",
    "1.0R standard"); r += 1
param_row(ws, r, "TP2_RR", "double (R)", "2.0",
    "Take Profit 2. Main profit target.",
    "2.0R minimum ICT target"); r += 1
param_row(ws, r, "TP3_RR", "double (R)", "3.0",
    "Runner target — remaining position rides to 3R.",
    "3.0R. Extend to 5.0 in strong trends"); r += 1
param_row(ws, r, "TP1_ClosePct", "double (%)", "40.0",
    "% of position to close at TP1.",
    "40% at TP1 locks nearly half expected profit early"); r += 1
param_row(ws, r, "TP2_ClosePct", "double (%)", "40.0",
    "% of position to close at TP2. Remaining becomes the runner.",
    "40% at TP2 — 20% runner remains"); r += 1
param_row(ws, r, "UseBreakeven", "bool", "true",
    "Move SL to breakeven after TP1. Trade can no longer become a loss.",
    "true — essential post-TP1 protection", highlight=True); r += 1
param_row(ws, r, "BreakevenBufferPips", "int (pips)", "5",
    "Extra pips above entry when moving SL to breakeven. Covers spread noise.",
    "5 pips for XAUUSD"); r += 1
param_row(ws, r, "UseTrailingStop", "bool", "false",
    "Enable trailing stop on runner after TP2.",
    "false default. Enable on strong trending days"); r += 1
param_row(ws, r, "TrailStartPips", "int (pips)", "50",
    "Trailing stop activates only after price moves this far beyond TP2.",
    "50 pips for XAUUSD runner"); r += 1
param_row(ws, r, "TrailStepPips", "int (pips)", "15",
    "How many pips price must advance before trailing stop moves.",
    "15 pips balanced step"); r += 1
param_row(ws, r, "SLBufferPips", "int (pips)", "10",
    "Extra pips added to calculated SL to avoid premature stops from spread/wicks.",
    "10 pips for XAUUSD"); r += 1
param_row(ws, r, "MaxSLPips", "int (pips)", "80",
    "Max SL size — skip trade if calculated SL exceeds this.",
    "80 pips for XAUUSD M15 setups"); r += 1
param_row(ws, r, "MinSLPips", "int (pips)", "15",
    "Min SL size — expand if calculated SL is too small (prevents micro-SL stops).",
    "15 pips covers Gold intraday noise"); r += 1
spacer(ws, r); r += 1

# ══ [23] DAILY PROFIT LOCK ═══════════════════════════════
section_hdr(ws, r, "[23]", "DAILY PROFIT LOCK — Protect the Good Days"); r += 1
param_row(ws, r, "UseProfitLock", "bool", "true",
    "ICT principle: once you have a good day, protect it.\n"
    "Reduces risk or stops trading when daily profit targets are met.",
    "true — protects gains from reversal"); r += 1
param_row(ws, r, "ProfitLock_ReduceR", "double (R)", "3.0",
    "After this many R daily profit, reduce position size.",
    "3.0R — after a great day, play smaller"); r += 1
param_row(ws, r, "ProfitLock_StopR", "double (R)", "5.0",
    "After this many R daily profit, STOP trading for the day.",
    "5.0R — bank exceptional gains and go"); r += 1
param_row(ws, r, "ProfitLock_ReducePct", "double (%)", "50.0",
    "Reduce risk by this % when ProfitLock_ReduceR is hit.",
    "50% — halve size for remaining trades"); r += 1
spacer(ws, r); r += 1

# ══ [24] DAILY LOSS PROTECTION ═══════════════════════════
section_hdr(ws, r, "[24]", "DAILY LOSS PROTECTION — Stop-Out Circuit Breaker"); r += 1
param_row(ws, r, "UseLossProtect", "bool", "true",
    "Automatically reduces or halts trading when daily losses reach thresholds.\n"
    "Prevents one bad day from destroying the account. NON-NEGOTIABLE.",
    "true — never disable", highlight=True); r += 1
param_row(ws, r, "LossProtect_ReduceR", "double (R)", "2.0",
    "After this many R daily loss, reduce risk by 50%.",
    "2.0R — trigger size reduction after 2 bad trades"); r += 1
param_row(ws, r, "LossProtect_StopR", "double (R)", "3.0",
    "After this many R daily loss, stop trading for the day entirely.",
    "3.0R — maximum daily loss. Do NOT exceed."); r += 1
param_row(ws, r, "MaxWeeklyLossR", "double (R)", "6.0",
    "Max weekly loss in R before EA halts until next Monday.",
    "6.0R = 2× daily stop loss"); r += 1
spacer(ws, r); r += 1

# ══ TRADE CLOSE RULES ════════════════════════════════════
section_hdr(ws, r, "CLOSE RULES", "Weekend Close / Trade Cooldown"); r += 1
param_row(ws, r, "CloseOnFriday", "bool", "true",
    "Close all open trades on Friday before the weekend.\n"
    "Avoids weekend gap risk — Gold frequently gaps on Sunday open.",
    "true — always close Gold before weekend", highlight=True); r += 1
param_row(ws, r, "FridayCloseHour", "int (GMT)", "14",
    "GMT hour on Friday to close all trades. 14:00 GMT = 10:00 AM EST.",
    "14 GMT — safe pre-weekend close"); r += 1
param_row(ws, r, "CooldownMinutes", "int", "15",
    "Minimum minutes between new trade entries.\n"
    "Prevents immediately re-entering after a close.",
    "15 min standard. Increase to 30 after a loss"); r += 1
spacer(ws, r); r += 1

# ══ [26] VISUAL CHART TOOLS ══════════════════════════════
section_hdr(ws, r, "[26]", "VISUAL CHART TOOLS — Drawing & Colour Settings"); r += 1
param_row(ws, r, "DrawPDH_PDL", "bool", "true",
    "Draw Previous Day High/Low lines — key liquidity reference levels.", "true"); r += 1
param_row(ws, r, "DrawPWH_PWL", "bool", "true",
    "Draw Previous Week High/Low lines — macro liquidity reference.", "true"); r += 1
param_row(ws, r, "DrawSessionRanges", "bool", "true",
    "Highlight Asian/London session ranges — Asian range is common manipulation target.", "true"); r += 1
param_row(ws, r, "DrawFVGZones", "bool", "true",
    "Draw FVG rectangles on chart — shows your active entry zones.", "true"); r += 1
param_row(ws, r, "DrawOBZones", "bool", "true",
    "Draw Order Block rectangles on chart.", "true"); r += 1
param_row(ws, r, "DrawLiqSweeps", "bool", "true",
    "Mark confirmed liquidity sweep points with arrows.", "true"); r += 1
param_row(ws, r, "DrawMSSLines", "bool", "true",
    "Draw MSS level lines where market structure shifted.", "true"); r += 1
param_row(ws, r, "DrawTargets", "bool", "true",
    "Project TP1/TP2/TP3 target levels on chart for active trades.", "true"); r += 1
param_row(ws, r, "ColorPDH", "color", "clrLime",
    "Previous Day High line colour.", "clrLime = green (supply)"); r += 1
param_row(ws, r, "ColorPDL", "color", "clrTomato",
    "Previous Day Low line colour.", "clrTomato = red (demand)"); r += 1
param_row(ws, r, "ColorPWH", "color", "clrAqua",
    "Previous Week High colour.", "clrAqua = cyan"); r += 1
param_row(ws, r, "ColorPWL", "color", "clrOrange",
    "Previous Week Low colour.", "clrOrange"); r += 1
param_row(ws, r, "ColorFVGBull", "color", "C'0,80,0'",
    "Bullish FVG fill colour.", "Dark green"); r += 1
param_row(ws, r, "ColorFVGBear", "color", "C'80,0,0'",
    "Bearish FVG fill colour.", "Dark red"); r += 1
param_row(ws, r, "ColorOBBull", "color", "C'0,50,100'",
    "Bullish Order Block fill colour.", "Dark blue"); r += 1
param_row(ws, r, "ColorOBBear", "color", "C'80,30,0'",
    "Bearish Order Block fill colour.", "Dark brown-red"); r += 1
spacer(ws, r); r += 1

# ══ [27] DEBUG PANEL ═════════════════════════════════════
section_hdr(ws, r, "[27]", "DEBUG PANEL — On-Chart Status Dashboard"); r += 1
param_row(ws, r, "ShowPanel", "bool", "true",
    "Show on-chart debug panel.\n"
    "Click [-] to collapse. Click [+] to expand. Drag title to move.\n"
    "Shows bias, liquidity, MSS, displacement, FVG, score, fail reasons, risk state.",
    "true — keep visible during development"); r += 1
param_row(ws, r, "PanelX", "int (pixels)", "12",
    "X position from left edge of chart.", "12"); r += 1
param_row(ws, r, "PanelY", "int (pixels)", "30",
    "Y position from top of chart.", "30"); r += 1
param_row(ws, r, "ShowStatPanel", "bool", "true",
    "Show statistics section (trades, win rate, profit factor, avg RR).", "true"); r += 1
spacer(ws, r); r += 1

# ══ [28] DEBUG & RELAXED MODE ════════════════════════════
section_hdr(ws, r, "[28]", "DEBUG & RELAXED MODE — Diagnosis Tools"); r += 1
param_row(ws, r, "DebugLogs", "bool", "false",
    "Print detailed engine logs to MT5 Journal tab.\n"
    "Shows: MSS swing level + break level, displacement body%/ATR ratio,\n"
    "OB detection results, complete score breakdown per component.\n"
    "TURN OFF after diagnosis — adds Journal spam on every bar.",
    "false production. true for investigating NO TRADE situations", highlight=True); r += 1
param_row(ws, r, "UseRelaxedMode", "bool", "false",
    "Relaxed Mode — lowers filters for testing whether issue is filtering vs detection.\n"
    "When ON: uses RelaxedMinScore, makes MSS/Disp/Daily optional (per switches below).\n"
    "Panel shows [RELAXED] indicator next to symbol.\n"
    "DO NOT use in live trading — for diagnosis only.",
    "false production. true to test what would trade", highlight=True); r += 1
param_row(ws, r, "RelaxedMinScore", "int", "50",
    "Minimum score when UseRelaxedMode = true.\n"
    "50 allows setups with just sweep + killzone + bias to qualify.",
    "50 for testing. Lower to 30 for maximum relaxation"); r += 1
param_row(ws, r, "RelaxedMSSOptional", "bool", "true",
    "In relaxed mode: MSS not required for trade entry.",
    "true — skip MSS requirement in diagnosis"); r += 1
param_row(ws, r, "RelaxedDispOptional", "bool", "true",
    "In relaxed mode: Displacement not required.",
    "true — skip displacement requirement"); r += 1
param_row(ws, r, "RelaxedDailyOptional", "bool", "true",
    "In relaxed mode: Daily bias mismatch does not fail the setup.",
    "true — ignore daily bias opposition"); r += 1
param_row(ws, r, "RelaxedWeeklyOptional", "bool", "false",
    "In relaxed mode: also make weekly bias optional.\n"
    "Keep false unless you suspect weekly bias detection has issues.",
    "false — keep weekly bias requirement even in relaxed mode"); r += 1
param_row(ws, r, "AllowFVGOnlyEntry", "bool", "false",
    "Allow entries with Sweep + FVG alone — no MSS or Displacement required.\n"
    "SIGNIFICANTLY increases trade frequency but reduces quality.\n"
    "Suitable if you want 10–20 trades/day but accept lower win rate.",
    "false strict mode. true for high-frequency FVG mode"); r += 1
param_row(ws, r, "AllowOBEntry", "bool", "false",
    "Allow Order Block entries without the full ICT sequence.\n"
    "OB is detected from any sweep direction — no displacement required.",
    "false strict mode. true to enable OB-only entries"); r += 1
param_row(ws, r, "ExpandKillzones", "bool", "false",
    "Expand killzone hours to include Asian session + NY extended to 18:00 GMT.\n"
    "Significantly increases trade frequency (wider trading window).\n"
    "Asian session quality is lower for Gold — expect more false signals.",
    "false strict mode. true to expand trading hours"); r += 1

# Footer
spacer(ws, r); r += 1
ws.row_dimensions[r].height = 18
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1,
            value="  ★  All sections covered: SYMBOL PRESET → [28] DEBUG & RELAXED MODE  ·  Total: 95+ parameters  ★")
c.fill = fill("1A1030"); c.font = font(FG_GOLD, italic=True, size=9)
c.alignment = align("center"); c.border = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ACCOUNT PROFILES
# ══════════════════════════════════════════════════════════════════════════════

ws2 = ws_acct
ws2.sheet_view.showGridLines = False

col_w2 = [2, 30, 16, 16, 16, 16, 24]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

def w2(row, col, value, bg, fg, bold=False, size=10, h="left", wrap=False):
    c = ws2.cell(row=row, column=col, value=value)
    c.fill = fill(bg); c.font = font(fg, bold, size)
    c.alignment = align(h, "center", wrap); c.border = thin_border()
    return c

def s2_spacer(row):
    ws2.row_dimensions[row].height = 6
    for col in range(1, 8):
        c = ws2.cell(row=row, column=col, value="")
        c.fill = fill(BG_DARK); c.border = thin_border()

r2 = 1

ws2.row_dimensions[r2].height = 40
ws2.merge_cells(f"A{r2}:G{r2}")
c = ws2.cell(row=r2, column=1, value="  ICT ATLAS EA V1.0  —  Recommended Settings by Account Size")
c.fill = fill("1A1030"); c.font = font(FG_GOLD, bold=True, size=14)
c.alignment = align("center"); c.border = thin_border()
r2 += 1

ws2.row_dimensions[r2].height = 18
ws2.merge_cells(f"A{r2}:G{r2}")
c = ws2.cell(row=r2, column=1, value="  XAUUSD M15  ·  ICT 2022 Mentorship  ·  Created by RATTANA CHHORM")
c.fill = fill("12101E"); c.font = font(FG_GRAY, italic=True, size=9)
c.alignment = align("center"); c.border = thin_border()
r2 += 1

s2_spacer(r2); r2 += 1

ws2.row_dimensions[r2].height = 22
hdrs2 = ["", "Input Parameter", "Section", "$50 Account", "$100 Account", "$1,000 Account", "$10,000 Account"]
for i, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=r2, column=i, value=h)
    c.fill = fill(BG_COL); c.font = font(FG_GOLD, bold=True, size=9)
    c.alignment = align("center"); c.border = thin_border()
r2 += 1

def acct_row(row, name, section, v50, v100, v1k, v10k, highlight=False):
    ws2.row_dimensions[row].height = 22
    bg = "2A2000" if highlight else BG_PANEL
    w2(row, 1, "", BG_DARK, FG_WHITE)
    w2(row, 2, name, bg, FG_SILVER, size=9)
    w2(row, 3, section, bg, FG_GRAY, size=8, h="center")
    for col, val in enumerate([v50, v100, v1k, v10k], 4):
        bg_v = BG_GREEN if val not in ("—", "", None) else BG_PANEL
        w2(row, col, val, bg_v, FG_GREEN, size=9, h="center")

def acct_section(row, title):
    ws2.row_dimensions[row].height = 18
    ws2.merge_cells(f"B{row}:G{row}")
    c = ws2.cell(row=row, column=2, value=f"◆  {title}")
    c.fill = fill(BG_HDR); c.font = font(FG_CYAN, bold=True, size=10)
    c.alignment = align("left"); c.border = thin_border()
    for col in [1] + list(range(3, 8)):
        try:
            ws2.cell(row=row, column=col, value="").fill = fill(BG_HDR)
        except AttributeError:
            ws2.cell(row=row, column=col).fill = fill(BG_HDR)

acct_section(r2, "CRITICAL — Risk & Position Sizing"); r2 += 1
acct_row(r2, "RiskPercent", "[25]", "0.5%", "0.5%", "1.0%", "1.0–2.0%", True); r2 += 1
acct_row(r2, "RiskMode", "[25]", "RISK_PCT", "RISK_PCT", "RISK_PCT", "RISK_PCT"); r2 += 1
acct_row(r2, "FixedLotSize (backup)", "[25]", "0.01", "0.01", "0.05", "0.10"); r2 += 1
acct_row(r2, "MaxLotCap", "[25]", "0.02", "0.03", "0.25", "2.00"); r2 += 1
acct_row(r2, "MaxTradesPerDay", "[25]", "3", "3", "5", "10"); r2 += 1
acct_row(r2, "MaxConsecLosses", "[25]", "2", "3", "3", "5"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "Stop Loss & Take Profit"); r2 += 1
acct_row(r2, "MaxSLPips", "[22]", "50", "60", "80", "100"); r2 += 1
acct_row(r2, "MinSLPips", "[22]", "10", "10", "15", "15"); r2 += 1
acct_row(r2, "SLBufferPips", "[22]", "5", "7", "10", "10"); r2 += 1
acct_row(r2, "TP1_RR", "[22]", "1.0", "1.0", "1.0", "1.0"); r2 += 1
acct_row(r2, "TP2_RR", "[22]", "2.0", "2.0", "2.0", "2.5"); r2 += 1
acct_row(r2, "TP3_RR (runner)", "[22]", "3.0", "3.0", "3.0", "4.0"); r2 += 1
acct_row(r2, "TP1_ClosePct", "[22]", "50%", "50%", "40%", "40%"); r2 += 1
acct_row(r2, "TP2_ClosePct", "[22]", "50%", "50%", "40%", "40%"); r2 += 1
acct_row(r2, "UseBreakeven", "[22]", "true", "true", "true", "true", True); r2 += 1
acct_row(r2, "BreakevenBufferPips", "[22]", "3", "5", "5", "7"); r2 += 1
acct_row(r2, "UseTrailingStop", "[22]", "false", "false", "false", "true"); r2 += 1
acct_row(r2, "TrailStartPips", "[22]", "—", "—", "—", "50"); r2 += 1
acct_row(r2, "TrailStepPips", "[22]", "—", "—", "—", "15"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "Daily Protection"); r2 += 1
acct_row(r2, "UseLossProtect", "[24]", "true", "true", "true", "true", True); r2 += 1
acct_row(r2, "LossProtect_ReduceR", "[24]", "1.5", "2.0", "2.0", "2.5"); r2 += 1
acct_row(r2, "LossProtect_StopR", "[24]", "2.5", "3.0", "3.0", "4.0"); r2 += 1
acct_row(r2, "MaxWeeklyLossR", "[24]", "4.0", "5.0", "6.0", "8.0"); r2 += 1
acct_row(r2, "UseProfitLock", "[23]", "true", "true", "true", "true"); r2 += 1
acct_row(r2, "ProfitLock_ReduceR", "[23]", "2.0", "3.0", "3.0", "4.0"); r2 += 1
acct_row(r2, "ProfitLock_StopR", "[23]", "3.0", "4.0", "5.0", "6.0"); r2 += 1
acct_row(r2, "ProfitLock_ReducePct", "[23]", "50%", "50%", "50%", "50%"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "Trade Quality Filters"); r2 += 1
acct_row(r2, "MinScore", "[18]", "90", "85", "80", "75"); r2 += 1
acct_row(r2, "AllowedGrades", "[19]", "APLUS", "GRADES_A_UP", "GRADES_A_UP", "GRADES_B_UP"); r2 += 1
acct_row(r2, "UseScoringSystem", "[18]", "true", "true", "true", "true"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "Core ICT Filters (identical for all account sizes)"); r2 += 1
for name, sec in [("UseBiasEngine","[01]"),("UseLiquidityEngine","[02]"),("UseMSSFilter","[03]"),
                  ("UseDispFilter","[04]"),("UseFVGFilter","[05]"),("UsePremDiscFilter","[07]"),
                  ("UseSessionFilter","[08]"),("UseNewsFilter","[12]"),("CloseOnFriday","Close"),
                  ("UseBreakeven","[22]")]:
    acct_row(r2, name, sec, "true","true","true","true"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "Optional Filters (enable as account and confidence grows)"); r2 += 1
acct_row(r2, "RequireH4Bias", "[01]", "true", "true", "false", "false"); r2 += 1
acct_row(r2, "UsePO3Filter", "[09]", "false", "false", "false", "true"); r2 += 1
acct_row(r2, "UseSMTFilter", "[10]", "false", "false", "false", "true"); r2 += 1
acct_row(r2, "UseCorrelFilter", "[17]", "false", "false", "false", "true"); r2 += 1
acct_row(r2, "UseTrailingStop", "[22]", "false", "false", "false", "true"); r2 += 1
s2_spacer(r2); r2 += 1

acct_section(r2, "High-Frequency Mode (15–30 trades/day — HIGHER RISK)"); r2 += 1
acct_row(r2, "AllowFVGOnlyEntry", "[28]", "—", "—", "true", "true"); r2 += 1
acct_row(r2, "ExpandKillzones", "[28]", "—", "—", "true", "true"); r2 += 1
acct_row(r2, "MinScore (HF mode)", "[18]", "—", "—", "60", "65"); r2 += 1
acct_row(r2, "AllowedGrades (HF)", "[19]", "—", "—", "GRADES_B_UP", "GRADES_B_UP"); r2 += 1
acct_row(r2, "MaxTradesPerDay (HF)", "[25]", "—", "—", "15", "25"); r2 += 1
acct_row(r2, "RequireWeeklyBias (HF)", "[01]", "—", "—", "false", "false"); r2 += 1
acct_row(r2, "RequireDailyBias (HF)", "[01]", "—", "—", "false", "false"); r2 += 1
acct_row(r2, "SessionAsian (HF)", "[08]", "—", "—", "true", "true"); r2 += 1
acct_row(r2, "Expected WR (estimate)", "—", "—", "—", "45–55%", "45–55%"); r2 += 1
s2_spacer(r2); r2 += 1

# Notes
for txt, fg_c in [
    ("★  $50 Account: Ultra-conservative. 0.5% risk. A+ trades only. Goal: survive and learn the model.", FG_GOLD),
    ("★  $100 Account: Conservative. 0.5% risk. A+ and A trades. Prove edge before scaling.", FG_CYAN),
    ("★  $1,000 Account: Standard. 1% risk. Full ICT model active. Scale only after 50+ trades.", FG_GREEN),
    ("★  $10,000 Account: Advanced. 1–2% risk. Optional filters ON. Trail runners on trend days.", "DDDDDD"),
    ("⚠  High-Frequency Mode: Relaxes ICT requirements. Win rate drops to ~45–55%. Use only with proven edge.", "FF9900"),
    ("⚠  15–30 trades/day on ONE symbol is very aggressive. Consider trading 3–5 symbols simultaneously.", "FF9900"),
]:
    ws2.row_dimensions[r2].height = 20
    ws2.merge_cells(f"A{r2}:G{r2}")
    c = ws2.cell(row=r2, column=1, value="  " + txt)
    c.fill = fill(BG_HDR); c.font = font(fg_c, size=9, italic=True)
    c.alignment = align("left"); c.border = thin_border()
    r2 += 1

out = "/home/user/Rattana/ICT_ATLAS_EA_Input_Guide.xlsx"
wb.save(out)
print(f"Saved: {out}")

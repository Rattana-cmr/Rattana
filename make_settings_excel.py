#!/usr/bin/env python3
"""Generate EA Input Settings Excel file for ICT SMC EA V1.1 and SMC SCALPER V1.0"""

from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────
C_TITLE_BG   = "1A1A3E"   # dark navy
C_TITLE_FG   = "FFD700"   # gold
C_HDR_BG     = "2E2E5E"   # medium navy
C_HDR_FG     = "FFFFFF"   # white
C_GRP_BG     = "3A3A6A"   # section header
C_GRP_FG     = "FFD700"   # gold
C_ROW_A      = "F0F0FF"   # light lavender (alternating row A)
C_ROW_B      = "FFFFFF"   # white (alternating row B)
C_GOLD_CELL  = "FFF8DC"   # cornsilk — default value highlight
C_GREEN_CELL = "E8FFE8"   # light green — recommended
C_RED_CELL   = "FFE8E8"   # light red — warning
C_BLUE_CELL  = "E8F0FF"   # light blue — info
C_NOTE_BG    = "FFFACD"   # lemon chiffon — notes
C_BORDER     = "8080C0"   # soft purple border


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(color="000000", bold=False, size=9, name="Calibri", italic=False):
    return Font(color=color, bold=bold, size=size, name=name, italic=italic)

def make_border(style="thin"):
    s = Side(style=style, color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def make_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def make_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_title(ws, row, text, colspan, bg=C_TITLE_BG, fg=C_TITLE_FG, size=13):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=colspan)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = make_fill(bg)
    c.font      = make_font(fg, bold=True, size=size)
    c.alignment = make_center()
    c.border    = make_border()
    ws.row_dimensions[row].height = 28

def write_header(ws, row, cols, col_widths=None):
    for ci, (colname, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=colname)
        c.fill      = make_fill(C_HDR_BG)
        c.font      = make_font(C_HDR_FG, bold=True, size=9)
        c.alignment = make_center()
        c.border    = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 22

def write_section(ws, row, text, colspan):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=colspan)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = make_fill(C_GRP_BG)
    c.font      = make_font(C_GRP_FG, bold=True, size=9)
    c.alignment = make_left()
    c.border    = make_border()
    ws.row_dimensions[row].height = 18

def write_row(ws, row, values, row_bg=None, cell_colors=None):
    """Write a data row. cell_colors is dict {col_index: hex_color}"""
    bg = row_bg if row_bg else (C_ROW_A if row % 2 == 0 else C_ROW_B)
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        col_bg = (cell_colors.get(ci) if cell_colors else None) or bg
        c.fill      = make_fill(col_bg)
        c.font      = make_font("1A1A3E", size=9)
        c.alignment = make_left()
        c.border    = make_border()
    ws.row_dimensions[row].height = 16

def write_note_row(ws, row, text, colspan):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=colspan)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = make_fill(C_NOTE_BG)
    c.font      = make_font("7B5A00", italic=True, size=8)
    c.alignment = make_left()
    c.border    = make_border()
    ws.row_dimensions[row].height = 14


# ══════════════════════════════════════════════════════════════════════
#  SHEET 1 — ICT SMC EA V1.1
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "ICT SMC EA V1.1"
ws1.freeze_panes = "A4"

COLS1 = [
    ("Parameter",          22),
    ("Default Value",      14),
    ("$100",               12),
    ("$200–$300",          12),
    ("$500",               12),
    ("$1,000",             12),
    ("Type",               12),
    ("Description",        52),
]
N1 = len(COLS1)

r = 1
write_title(ws1, r, "ICT SMC EA V1.1  —  Input Settings Guide  |  Created by: RATTANA CHHORM", N1)
r += 1
write_title(ws1, r,
    "Symbol: XAUUSD  |  Timeframe: M15  |  Magic Number: 888777  |  Strategy: ICT Smart Money Concepts",
    N1, bg="2A2A5A", fg="C8C8FF", size=9)
r += 1
write_header(ws1, r, COLS1)
r += 1

# Helper — each entry: (param, default, $100, $200-300, $500, $1000, type, description)
# color_map col indices: 3=$100, 4=$200, 5=$500, 6=$1000
def ea1_row(ws, row, param, default, v100, v200, v500, v1000, typ, desc, hl=None):
    cc = {3: C_GREEN_CELL, 4: C_GREEN_CELL, 5: C_GREEN_CELL, 6: C_GREEN_CELL,
          2: C_GOLD_CELL}
    if hl == "warn":
        for k in [3,4,5,6]: cc[k] = C_RED_CELL
    elif hl == "info":
        for k in [3,4,5,6]: cc[k] = C_BLUE_CELL
    write_row(ws, row, [param, default, v100, v200, v500, v1000, typ, desc],
              cell_colors=cc)

# ─── RISK MANAGEMENT ────────────────────────────────────────────────
write_section(ws1, r, "  ═══  RISK MANAGEMENT  ═══", N1); r+=1
ea1_row(ws1,r,"RiskMode","RISK_FIXED_PCT","RISK_FIXED_PCT","RISK_FIXED_PCT","RISK_FIXED_PCT","RISK_FIXED_PCT","Enum","How lot size is calculated. FIXED_PCT = % of balance (recommended). FIXED_LOT = fixed size. DYNAMIC_EQ = uses equity."); r+=1
ea1_row(ws1,r,"RiskPercent","0.5","0.5","0.5","0.5","0.5–1.0","Double (%)","% of balance risked per trade. Keep at 0.5% for all small accounts. Never exceed 2%."); r+=1
ea1_row(ws1,r,"FixedLot","0.0","0.0","0.0","0.0","0.0","Double","Fixed lot override. Set 0.0 to use RiskPercent. Only set a number if you want a locked lot size."); r+=1
ea1_row(ws1,r,"MaxDailyLossPercent","10.0","5.0","5.0","5.0","10.0","Double (%)","EA stops trading for the day if total loss exceeds this %. Use 5% for small accounts.",hl="warn"); r+=1
ea1_row(ws1,r,"MaxTradesPerDay","10","3","5","5","10","Integer","Maximum number of trades allowed per day. ICT SMC is selective — rarely needs more than 3–5."); r+=1
ea1_row(ws1,r,"RewardRiskRatio","2.0","2.0","2.0","2.0","2.0","Double","TP = SL distance × this value. 2.0 = 1:2 R:R. Do not go below 1.5."); r+=1
ea1_row(ws1,r,"MaxLotLimit","0.10","0.01","0.02","0.05","0.10","Double","Hard ceiling on lot size. Safety cap — prevents a bug from opening huge lots.",hl="warn"); r+=1
ea1_row(ws1,r,"MinRewardRiskRatio","2.0","2.0","2.0","2.0","2.0","Double","Skip trade if achievable R:R is below this. Keep at 2.0."); r+=1

# ─── TAKE PROFIT MODE ───────────────────────────────────────────────
write_section(ws1, r, "  ═══  TAKE PROFIT MODE  ═══", N1); r+=1
ea1_row(ws1,r,"TPMode","TP_FIXED_RR","TP_FIXED_RR","TP_FIXED_RR","TP_FIXED_RR","TP_FIXED_RR","Enum","TP calculation method. FIXED_RR = fixed ratio. ATR = ATR multiplier. HYBRID = more conservative of the two.",hl="info"); r+=1
ea1_row(ws1,r,"ATRMultiplierTP","3.0","3.0","3.0","3.0","3.0","Double","ATR multiplier for TP. Only used when TPMode = TP_ATR or TP_HYBRID."); r+=1

# ─── PARTIAL TAKE PROFIT ─────────────────────────────────────────────
write_section(ws1, r, "  ═══  PARTIAL TAKE PROFIT  ═══", N1); r+=1
ea1_row(ws1,r,"UsePartialTP","true","true","true","true","true","Boolean","Close 50% of position at 1R, then move SL to breakeven. Highly recommended — locks profit."); r+=1
ea1_row(ws1,r,"PartialClosePercent","50.0","50.0","50.0","50.0","50.0","Double (%)","% of position to close at partial TP level. Keep at 50%."); r+=1
ea1_row(ws1,r,"PartialCloseRR","1.0","1.0","1.0","1.0","1.0","Double","R:R level to trigger partial close. 1.0 = when price reaches 1:1, close half."); r+=1

# ─── TRADE FILTERS ────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  TRADE FILTERS  ═══", N1); r+=1
ea1_row(ws1,r,"UseTimeFilter","true","true","true","true","true","Boolean","Restrict trades to session hours defined in the Sessions section."); r+=1
ea1_row(ws1,r,"MaxSpreadPoints","50","50","50","50","50","Integer","Skip trade if broker spread exceeds this (in points). 50 pts = 0.5 pip for XAUUSD."); r+=1
ea1_row(ws1,r,"MinStopDistance","20","20","20","20","20","Integer","Minimum SL distance in points — broker protection. Keep at 20."); r+=1
ea1_row(ws1,r,"MaxConsecutiveLosses","10","3","5","5","10","Integer","Stop trading after N losses in a row. Use smaller value for small accounts.",hl="warn"); r+=1
ea1_row(ws1,r,"MaxDailyLossTrades","3","3","3","3","3","Integer","Stop trading after N losing trades in one day (FIX T)."); r+=1
ea1_row(ws1,r,"ResetLossStreakDaily","true","true","true","true","true","Boolean","Reset consecutive loss counter at start of each new day."); r+=1

# ─── ICT STRUCTURE FILTERS ────────────────────────────────────────────
write_section(ws1, r, "  ═══  ICT STRUCTURE FILTERS  ═══", N1); r+=1
ea1_row(ws1,r,"UseMSSFilter","true","true","true","true","true","Boolean","Require H1 Market Structure Shift before entry. Core ICT filter. Keep true for quality."); r+=1
ea1_row(ws1,r,"UseBOSFilter","true","true","true","true","true","Boolean","Require M15 Break of Structure — must agree with MSS direction. Keep true."); r+=1
ea1_row(ws1,r,"RequireLiquiditySweep","false","false","false","false","false","Boolean","Require a stop hunt before entry. false = more trades. true = higher precision only."); r+=1
ea1_row(ws1,r,"UseSMTFilter","false","false","false","false","false","Boolean","Smart Money Tool — compare to correlated symbol. Keep false unless XAGUSD data available."); r+=1
ea1_row(ws1,r,"SMTSymbol","XAGUSD","XAGUSD","XAGUSD","XAGUSD","XAGUSD","String","Correlated symbol for SMT divergence check. Only used if UseSMTFilter = true."); r+=1

# ─── NEWS FILTER ─────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  NEWS FILTER  ═══", N1); r+=1
ea1_row(ws1,r,"UseNewsFilter","false","false","false","false","false","Boolean","Block trades around high-impact news (MT5 Calendar). Keep false for backtesting."); r+=1
ea1_row(ws1,r,"NewsBlockBeforeMin","30","30","30","30","30","Integer","Block trading N minutes before a news event."); r+=1
ea1_row(ws1,r,"NewsBlockAfterMin","30","30","30","30","30","Integer","Block trading N minutes after a news event."); r+=1

# ─── TRADE QUALITY SCORE ──────────────────────────────────────────────
write_section(ws1, r, "  ═══  TRADE QUALITY SCORE  ═══", N1); r+=1
ea1_row(ws1,r,"MinimumTradeScore","70","70","70","70","70","Integer (0–100)","Minimum score for entry. 70 = only trades scoring 70/100 or above. 0 = disable scoring.",hl="info"); r+=1

# ─── SESSIONS ────────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  SESSIONS (GMT TIME)  ═══", N1); r+=1
ea1_row(ws1,r,"AutoDetectGMT","true","true","true","true","true","Boolean","Auto-detect broker GMT offset. Keep true."); r+=1
ea1_row(ws1,r,"BrokerGMTOffset","0","0","0","0","0","Integer","Manual GMT offset. Only used if AutoDetectGMT = false."); r+=1
ea1_row(ws1,r,"SessionSydney","false","false","false","false","false","Boolean","Sydney session 22:00–02:00 GMT. Low XAUUSD volume. Keep false."); r+=1
ea1_row(ws1,r,"SessionTokyo","false","false","false","false","false","Boolean","Tokyo session 00:00–04:00 GMT. Keep false for XAUUSD."); r+=1
ea1_row(ws1,r,"SessionLondon","true","true","true","true","true","Boolean","London session 08:00–12:00 GMT. Best session for XAUUSD. Keep true."); r+=1
ea1_row(ws1,r,"SessionNewYork","true","true","true","true","true","Boolean","New York session 13:00–17:00 GMT. Second best. Keep true."); r+=1
ea1_row(ws1,r,"OverlapLondonNY","true","true","true","true","true","Boolean","London + NY overlap 13:00–16:00 GMT. Highest volume. Keep true."); r+=1
ea1_row(ws1,r,"OverlapTokyoLondon","false","false","false","false","false","Boolean","Tokyo + London overlap 07:00–08:00 GMT. Keep false."); r+=1

# ─── STOP LOSS ────────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  STOP LOSS  ═══", N1); r+=1
ea1_row(ws1,r,"SLBufferPips","15","15","15","15","15","Integer","Extra buffer beyond swing high/low for SL (pips). Prevents SL from sitting on exact wick."); r+=1
ea1_row(ws1,r,"UseTrailingStop","false","false","false","false","false","Boolean","Move SL in profit direction. Keep false — UsePartialTP already handles SL→BE."); r+=1
ea1_row(ws1,r,"TrailingStartPips","30","30","30","30","30","Integer","Pips in profit before trailing activates. Only used if UseTrailingStop = true."); r+=1
ea1_row(ws1,r,"TrailingStepPips","10","10","10","10","10","Integer","Trail step in pips. Only used if UseTrailingStop = true."); r+=1

# ─── POSITION MANAGEMENT ─────────────────────────────────────────────
write_section(ws1, r, "  ═══  POSITION MANAGEMENT  ═══", N1); r+=1
ea1_row(ws1,r,"CloseOnFriday","true","true","true","true","true","Boolean","Close all positions at Friday cutoff — prevents weekend gap risk."); r+=1
ea1_row(ws1,r,"FridayCloseHour","14","14","14","14","14","Integer","GMT hour to close positions on Friday. 14:00 GMT (FIX P)."); r+=1
ea1_row(ws1,r,"UseBreakeven","false","false","false","false","false","Boolean","Move SL to entry at trigger level. Keep false — UsePartialTP already does SL→BE."); r+=1
ea1_row(ws1,r,"BreakevenTriggerPips","40","40","40","40","40","Integer","Pips in profit before SL moves to entry. Only used if UseBreakeven = true."); r+=1

# ─── SWING DETECTION ──────────────────────────────────────────────────
write_section(ws1, r, "  ═══  SWING DETECTION  ═══", N1); r+=1
ea1_row(ws1,r,"SwingLookbackBarsH1","50","50","50","50","50","Integer","H1 bars to scan for swing highs/lows (OTE zone). Keep at 50."); r+=1
ea1_row(ws1,r,"SwingConfirmBarsH1","3","3","3","3","3","Integer","Bars each side to confirm H1 swing. Keep at 3."); r+=1
ea1_row(ws1,r,"SwingLookbackBarsM15","30","30","30","30","30","Integer","M15 bars to scan for SL swing. Keep at 30."); r+=1
ea1_row(ws1,r,"SwingConfirmBarsM15","5","5","5","5","5","Integer","Bars each side to confirm M15 swing. Keep at 5."); r+=1
ea1_row(ws1,r,"MaxSwingDistancePips","500","500","500","500","500","Integer","Max distance from swing to current price. Prevents using stale swings."); r+=1
ea1_row(ws1,r,"MaxSLPips","30","30","30","30","30","Integer","Maximum SL size in pips (FIX O). Trade skipped if SL would be wider."); r+=1
ea1_row(ws1,r,"MinSLPips","10","10","10","10","10","Integer","Minimum SL size in pips (FIX Q). Trade skipped if SL too tight."); r+=1
ea1_row(ws1,r,"ShowSwingLines","true","true","true","true","true","Boolean","Draw swing high/low lines on chart. Visual only."); r+=1

# ─── ICT TWINS MODEL ─────────────────────────────────────────────────
write_section(ws1, r, "  ═══  ICT TWINS MODEL  ═══", N1); r+=1
ea1_row(ws1,r,"UseTwinsModel","true","true","true","true","true","Boolean","Enable ICT Twins entry model (OTE + 1M trigger). Core entry logic. Keep true."); r+=1
ea1_row(ws1,r,"HTFLevelMinutes","15","15","15","15","15","Integer","HTF reference timeframe in minutes. 15 = M15. Keep at 15."); r+=1
ea1_row(ws1,r,"OTEMinPercent","0.65","0.65","0.65","0.65","0.65","Double","Lower edge of OTE zone — 65% Fibonacci retracement of H1 swing (FIX J)."); r+=1
ea1_row(ws1,r,"OTEMaxPercent","0.75","0.75","0.75","0.75","0.75","Double","Upper edge of OTE zone — 75% Fibonacci retracement of H1 swing (FIX J)."); r+=1
ea1_row(ws1,r,"OTESweetSpotPercent","0.705","0.705","0.705","0.705","0.705","Double","Ideal sweet spot inside OTE zone. Used for scoring."); r+=1
ea1_row(ws1,r,"MinFVGsRequired","0","0","0","0","0","Integer","Minimum 1-minute FVGs inside OTE zone. 0 = do not require any."); r+=1
ea1_row(ws1,r,"HTFToleranceATRMulti","2","2","2","2","2","Integer","ATR multiplier for HTF level tolerance."); r+=1
ea1_row(ws1,r,"HTFLevelRequired","false","false","false","false","false","Boolean","Require price near Daily/Weekly level (FIX U). false = auto-pass → more trades."); r+=1
ea1_row(ws1,r,"ShowOTEZone","true","true","true","true","true","Boolean","Draw OTE zone box on chart. Visual only. Set false if chart looks cluttered."); r+=1
ea1_row(ws1,r,"MinH1RangePips","50","50","50","50","50","Integer","Minimum H1 swing range to qualify. Filters out tiny, unreliable swings."); r+=1

# ─── MSS / BOS / LIQUIDITY ────────────────────────────────────────────
write_section(ws1, r, "  ═══  MSS / BOS / LIQUIDITY  ═══", N1); r+=1
ea1_row(ws1,r,"MSSLookbackBars","30","30","30","30","30","Integer","H1 bars to scan for Market Structure Shift. Keep at 30."); r+=1
ea1_row(ws1,r,"MSSConfirmBars","3","3","3","3","3","Integer","Swing confirm bars for MSS detection."); r+=1
ea1_row(ws1,r,"BOSLookbackBars","20","20","20","20","20","Integer","M15 bars to scan for Break of Structure."); r+=1
ea1_row(ws1,r,"LiquidityLookbackBars","50","50","50","50","50","Integer","M15 bars to scan for liquidity sweep."); r+=1
ea1_row(ws1,r,"LiquidityWickPips","3","3","3","3","3","Integer","Min wick beyond liquidity level to count as a sweep."); r+=1

# ─── SYMBOL PRESET ────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  SYMBOL PRESET  ═══", N1); r+=1
ea1_row(ws1,r,"SymbolPreset","PRESET_AUTO","PRESET_AUTO","PRESET_AUTO","PRESET_AUTO","PRESET_AUTO","Enum","Auto-selects OTE range and SL limits for the symbol. Keep PRESET_AUTO for XAUUSD.",hl="info"); r+=1

# ─── OPTIMIZATION MODE ────────────────────────────────────────────────
write_section(ws1, r, "  ═══  OPTIMIZATION MODE  ═══", N1); r+=1
ea1_row(ws1,r,"OptMode","OPT_BALANCED","OPT_CONSERVATIVE","OPT_CONSERVATIVE","OPT_BALANCED","OPT_BALANCED","Enum","CONSERVATIVE: score≥80, risk×0.5. BALANCED: score≥70, risk as-set. AGGRESSIVE: score≥55, risk×2.",hl="warn"); r+=1

# ─── LOGGING ──────────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  LOGGING  ═══", N1); r+=1
ea1_row(ws1,r,"EnableScreenshot","true","true","true","true","true","Boolean","Save chart screenshot on every trade open/close. Set false to save disk space."); r+=1
ea1_row(ws1,r,"EnableCSVLog","true","true","true","true","true","Boolean","Export every trade to a CSV file. Saved to MT5 Common folder."); r+=1

# ─── DEBUG ────────────────────────────────────────────────────────────
write_section(ws1, r, "  ═══  DEBUG  ═══", N1); r+=1
ea1_row(ws1,r,"PostTradeCooldownMin","30","30","30","30","30","Integer","Minutes to wait after a trade closes before next entry. Prevents overtrading."); r+=1
ea1_row(ws1,r,"UseDailyTrendFilter","true","true","true","true","true","Boolean","Only trade in direction of daily trend. Adds D1 bias check."); r+=1
ea1_row(ws1,r,"BestHoursOnly","true","true","true","true","true","Boolean","Only trade 08:30–15:00 GMT — London open + NY morning (FIX N). Keep true."); r+=1
ea1_row(ws1,r,"ForceTrades","false","false","false","false","false","Boolean","Bypass ALL filters. TEST ONLY. NEVER set true on live account.",hl="warn"); r+=1
ea1_row(ws1,r,"DebugMode","false","false","false","false","false","Boolean","Print extra diagnostic messages in Experts tab. Set true only when troubleshooting."); r+=1
ea1_row(ws1,r,"RelaxedMode","false","false","false","false","false","Boolean","Widens OTE zone ±2%, disables FVG requirement. More trades but lower quality. Keep false.",hl="warn"); r+=1

# Final notes row
write_note_row(ws1, r,
    "  ⚑  COLOR KEY:  Gold = Default Value  |  Green = Recommended for that capital tier  |  "
    "Blue = Info/mode setting  |  Red = Risk/warning setting",
    N1); r+=1
write_note_row(ws1, r,
    "  ⚑  OptMode effect:  CONSERVATIVE → MinScore raised to 80, Risk halved to 0.25%  |  "
    "BALANCED → MinScore 70, Risk as set  |  AGGRESSIVE → MinScore 55, Risk doubled (max 2%)",
    N1); r+=1


# ══════════════════════════════════════════════════════════════════════
#  SHEET 2 — SMC SCALPER V1.0
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("SMC SCALPER V1.0")
ws2.freeze_panes = "A4"

COLS2 = [
    ("Parameter",          24),
    ("Default Value",      16),
    ("$100",               12),
    ("$200–$300",          12),
    ("$500",               12),
    ("$1,000",             12),
    ("Type",               14),
    ("Description",        52),
]
N2 = len(COLS2)

r = 1
write_title(ws2, r, "SMC SCALPER V1.0  —  Input Settings Guide  |  Created by: RATTANA CHHORM", N2)
r += 1
write_title(ws2, r,
    "Symbol: XAUUSD  |  Timeframe: M1 / M5  |  Magic Number: 999888  |  Strategy: EMA Crossover + RSI + ATR Scalper",
    N2, bg="2A2A5A", fg="C8C8FF", size=9)
r += 1
write_header(ws2, r, COLS2)
r += 1

def ea2_row(ws, row, param, default, v100, v200, v500, v1000, typ, desc, hl=None):
    cc = {3: C_GREEN_CELL, 4: C_GREEN_CELL, 5: C_GREEN_CELL, 6: C_GREEN_CELL,
          2: C_GOLD_CELL}
    if hl == "warn":
        for k in [3,4,5,6]: cc[k] = C_RED_CELL
    elif hl == "info":
        for k in [3,4,5,6]: cc[k] = C_BLUE_CELL
    write_row(ws, row, [param, default, v100, v200, v500, v1000, typ, desc],
              cell_colors=cc)

# ─── SIGNAL SETTINGS ─────────────────────────────────────────────────
write_section(ws2, r, "  ═══  SIGNAL SETTINGS  ═══", N2); r+=1
ea2_row(ws2,r,"SignalMode","EMA_CROSS","EMA_CROSS","EMA_CROSS","EMA_CROSS","EMA_CROSS","Enum","EMA_CROSS = crossover only (medium). EMA_SLOPE = every bar same slope (highest trades). PRICE_ACTION = engulfing (lowest).",hl="info"); r+=1
ea2_row(ws2,r,"EMAFast","5","5","5","5","5","Integer","Fast EMA period. Crossover signal uses fast vs slow."); r+=1
ea2_row(ws2,r,"EMASlow","13","13","13","13","13","Integer","Slow EMA period. Fast crosses slow = entry signal."); r+=1
ea2_row(ws2,r,"EMATrend","50","50","50","50","50","Integer","Trend EMA period. Only buy above this, only sell below this."); r+=1
ea2_row(ws2,r,"UseTrendFilterSC","true","true","true","true","true","Boolean","Only trade in EMA trend direction. false = more trades but lower quality."); r+=1
ea2_row(ws2,r,"RSIPeriodSC","14","14","14","14","14","Integer","RSI period for momentum filter."); r+=1
ea2_row(ws2,r,"RSIBuyMin","45.0","45.0","45.0","45.0","45.0","Double","Minimum RSI for BUY — confirms upward momentum."); r+=1
ea2_row(ws2,r,"RSIBuyMax","75.0","75.0","75.0","75.0","75.0","Double","Maximum RSI for BUY — avoids overbought entries."); r+=1
ea2_row(ws2,r,"RSISellMin","25.0","25.0","25.0","25.0","25.0","Double","Minimum RSI for SELL — avoids oversold entries."); r+=1
ea2_row(ws2,r,"RSISellMax","55.0","55.0","55.0","55.0","55.0","Double","Maximum RSI for SELL — confirms downward momentum."); r+=1
ea2_row(ws2,r,"UseRSIFilterSC","true","true","true","true","true","Boolean","Enable RSI momentum filter. false = more trades but more noise."); r+=1
ea2_row(ws2,r,"ATRPeriodSC","14","14","14","14","14","Integer","ATR period — measures market volatility for SL/TP sizing."); r+=1

# ─── STOP LOSS / TAKE PROFIT ─────────────────────────────────────────
write_section(ws2, r, "  ═══  STOP LOSS / TAKE PROFIT  ═══", N2); r+=1
ea2_row(ws2,r,"SLMultiATR","0.8","0.8","0.8","0.8","0.8","Double","SL = ATR × this value. 0.8 = 80% of current ATR. Adapts to volatility."); r+=1
ea2_row(ws2,r,"TPMultiATR","1.6","1.6","1.6","1.6","1.6","Double","TP = ATR × this value. 1.6 / 0.8 = 2:1 R:R. Keep ratio = TPMulti / SLMulti = 2.0."); r+=1

# ─── PARTIAL TAKE PROFIT ─────────────────────────────────────────────
write_section(ws2, r, "  ═══  PARTIAL TAKE PROFIT  ═══", N2); r+=1
ea2_row(ws2,r,"UsePartialSC","true","true","true","true","true","Boolean","Close 50% at 1R then SL to breakeven. Locks profit on winning trades."); r+=1
ea2_row(ws2,r,"PartialPctSC","50.0","50.0","50.0","50.0","50.0","Double (%)","% of position to close at partial TP trigger."); r+=1
ea2_row(ws2,r,"PartialRRSC","1.0","1.0","1.0","1.0","1.0","Double","R:R level that triggers partial close. 1.0 = at 1:1 R:R."); r+=1

# ─── RISK MANAGEMENT ─────────────────────────────────────────────────
write_section(ws2, r, "  ═══  RISK MANAGEMENT  ═══", N2); r+=1
ea2_row(ws2,r,"RiskModeSC","SC_RISK_FIXED_PCT","SC_RISK_FIXED_PCT","SC_RISK_FIXED_PCT","SC_RISK_FIXED_PCT","SC_RISK_FIXED_PCT","Enum","Risk calculation mode. FIXED_PCT = % of balance (recommended)."); r+=1
ea2_row(ws2,r,"RiskPctSC","0.5","0.5","0.5","0.5","0.5–1.0","Double (%)","% of balance risked per trade. Keep 0.5% for all small accounts."); r+=1
ea2_row(ws2,r,"FixedLotSC","0.0","0.0","0.0","0.0","0.0","Double","Fixed lot override. 0.0 = use RiskPctSC calculation."); r+=1
ea2_row(ws2,r,"MaxLotSC","0.10","0.01","0.02","0.05","0.10","Double","Hard ceiling on lot size.",hl="warn"); r+=1
ea2_row(ws2,r,"MaxDailyLossPctSC","10.0","5.0","5.0","5.0","10.0","Double (%)","EA stops trading the day if daily loss exceeds this %.",hl="warn"); r+=1
ea2_row(ws2,r,"MaxTradesPerDaySC","30","20","25","30","30","Integer","Maximum trades per day. Scalper can trade frequently — keep at 30 for max signals."); r+=1
ea2_row(ws2,r,"MaxConsecLossesSC","5","3","3","5","5","Integer","Stop after N consecutive losses.",hl="warn"); r+=1

# ─── TRADE FILTERS ────────────────────────────────────────────────────
write_section(ws2, r, "  ═══  TRADE FILTERS  ═══", N2); r+=1
ea2_row(ws2,r,"MaxSpreadSC","50","50","50","50","50","Integer","Skip if spread exceeds this (points). 50 pts = 0.5 pip for XAUUSD."); r+=1
ea2_row(ws2,r,"MinATRPipsSC","1.5","1.5","1.5","1.5","1.5","Double","Skip if ATR below this (pips). Avoids flat/choppy market entries."); r+=1
ea2_row(ws2,r,"MaxSLPipsSC","20","20","20","20","20","Integer","Skip if calculated SL exceeds this in pips."); r+=1
ea2_row(ws2,r,"MinSLPipsSC","2","2","2","2","2","Integer","Skip if calculated SL is below this — avoids broker stops-level issues."); r+=1
ea2_row(ws2,r,"CooldownBarsSC","2","2","2","2","2","Integer","Wait N bars after a trade before next entry. 2 bars = 2 minutes on M1."); r+=1
ea2_row(ws2,r,"OneAtATimeSC","true","true","true","true","true","Boolean","Only 1 open position at a time. Recommended for controlled risk."); r+=1

# ─── SESSIONS ────────────────────────────────────────────────────────
write_section(ws2, r, "  ═══  SESSIONS (GMT TIME)  ═══", N2); r+=1
ea2_row(ws2,r,"AutoGMT_SC","true","true","true","true","true","Boolean","Auto-detect broker GMT offset. Keep true."); r+=1
ea2_row(ws2,r,"GMTOffsetSC","0","0","0","0","0","Integer","Manual GMT offset. Only used if AutoGMT_SC = false."); r+=1
ea2_row(ws2,r,"SydneySC","false","false","false","false","false","Boolean","Sydney session 22:00–01:00 GMT. Low volume. Keep false."); r+=1
ea2_row(ws2,r,"TokyoSC","false","false","false","false","false","Boolean","Tokyo session 00:00–04:00 GMT. Keep false for XAUUSD."); r+=1
ea2_row(ws2,r,"LondonSC","true","true","true","true","true","Boolean","London session 08:00–12:00 GMT. Best for XAUUSD scalping."); r+=1
ea2_row(ws2,r,"NewYorkSC","true","true","true","true","true","Boolean","New York session 13:00–17:00 GMT. High volatility — good for scalping."); r+=1
ea2_row(ws2,r,"OverlapSC","true","true","true","true","true","Boolean","London + NY overlap 13:00–16:00 GMT. Highest volume. Keep true."); r+=1
ea2_row(ws2,r,"AllHoursSC","false","false","false","false","false","Boolean","Trade 24 hours — ignores session settings. Set true for maximum signals.",hl="info"); r+=1

# ─── POSITION MANAGEMENT ─────────────────────────────────────────────
write_section(ws2, r, "  ═══  POSITION MANAGEMENT  ═══", N2); r+=1
ea2_row(ws2,r,"CloseFridaySC","true","true","true","true","true","Boolean","Close all positions at Friday cutoff to avoid weekend gap risk."); r+=1
ea2_row(ws2,r,"FridayHourSC","14","14","14","14","14","Integer","GMT hour to close positions on Friday. 14:00 GMT."); r+=1
ea2_row(ws2,r,"TrailStopSC","false","false","false","false","false","Boolean","Enable ATR trailing stop. Keep false — UsePartialSC already handles SL→BE."); r+=1
ea2_row(ws2,r,"TrailATRMulti","0.5","0.5","0.5","0.5","0.5","Double","Trail = ATR × this. Only used if TrailStopSC = true."); r+=1

# ─── LOGGING ──────────────────────────────────────────────────────────
write_section(ws2, r, "  ═══  LOGGING  ═══", N2); r+=1
ea2_row(ws2,r,"ScreenshotSC","true","true","true","true","true","Boolean","Save chart screenshot on every trade open/close."); r+=1
ea2_row(ws2,r,"CSVLogSC","true","true","true","true","true","Boolean","Export trades to CSV file. Saved to MT5 Common folder."); r+=1

# ─── DEBUG ────────────────────────────────────────────────────────────
write_section(ws2, r, "  ═══  DEBUG  ═══", N2); r+=1
ea2_row(ws2,r,"ShowArrowsSC","true","true","true","true","true","Boolean","Draw buy/sell arrows on chart for visual confirmation."); r+=1
ea2_row(ws2,r,"DebugSC","false","false","false","false","false","Boolean","Print detailed debug messages in MT5 Experts tab."); r+=1
ea2_row(ws2,r,"ForceSC","false","false","false","false","false","Boolean","Bypass ALL filters — TEST ONLY. NEVER true on live account.",hl="warn"); r+=1

# Notes
write_note_row(ws2, r,
    "  ⚑  FOR MORE TRADES:  SignalMode=EMA_SLOPE + AllHoursSC=true + UseTrendFilterSC=false + UseRSIFilterSC=false + CooldownBarsSC=1",
    N2); r+=1
write_note_row(ws2, r,
    "  ⚑  FOR BETTER QUALITY:  SignalMode=EMA_CROSS + UseTrendFilterSC=true + UseRSIFilterSC=true + CooldownBarsSC=3",
    N2); r+=1
write_note_row(ws2, r,
    "  ⚑  COLOR KEY:  Gold = Default Value  |  Green = Recommended for that capital tier  |  "
    "Blue = Info/mode setting  |  Red = Risk/warning setting",
    N2); r+=1


# ══════════════════════════════════════════════════════════════════════
#  SHEET 3 — Quick Comparison
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Quick Comparison")

COLS3 = [
    ("Feature",            28),
    ("ICT SMC EA V1.1",    30),
    ("SMC SCALPER V1.0",   30),
]
N3 = len(COLS3)

r = 1
write_title(ws3, r, "ICT SMC EA V1.1  vs  SMC SCALPER V1.0  —  Quick Comparison", N3, size=12)
r += 1
write_header(ws3, r, COLS3)
r += 1

def cmp_row(ws, row, feat, v1, v2):
    for ci, val in enumerate([feat, v1, v2], 1):
        c = ws.cell(row=row, column=ci, value=val)
        cell_bg = C_GOLD_CELL if ci == 1 else (C_BLUE_CELL if ci == 2 else C_GREEN_CELL)
        c.fill = make_fill(cell_bg)
        c.font = make_font("1A1A3E", size=9)
        c.alignment = make_left()
        c.border = make_border()
    ws.row_dimensions[row].height = 16

comparisons = [
    ("Timeframe",               "M15",                          "M1 or M5"),
    ("Strategy type",           "ICT Smart Money Concepts",     "EMA Crossover Scalper"),
    ("Magic Number",            "888777",                       "999888"),
    ("Expected trades/month",   "1–8 (very selective)",         "200–500 (high frequency)"),
    ("Expected trades/day",     "0–1 per day",                  "10–25 per day"),
    ("Entry signal",            "OTE zone + MSS + BOS",         "EMA(5/13) crossover"),
    ("Trend filter",            "MSS on H1 + D1 trend",         "EMA(50) direction"),
    ("Momentum filter",         "Trade Quality Score (0–100)",  "RSI(14) range check"),
    ("SL basis",                "M15 swing low/high + buffer",  "ATR × multiplier"),
    ("TP basis",                "Fixed R:R or ATR or Hybrid",   "ATR × multiplier"),
    ("Partial TP",              "50% at 1R → SL to BE",         "50% at 1R → SL to BE"),
    ("Trailing stop",           "Optional",                     "Optional (ATR-based)"),
    ("Session filter",          "London + New York",            "London + New York"),
    ("News filter",             "Yes (MT5 Calendar API)",       "No"),
    ("Score gate",              "Min 70/100",                   "No score gate"),
    ("Liquidity sweep filter",  "Optional",                     "No"),
    ("Symbol Presets",          "XAUUSD/BTCUSD/EURUSD/GBPUSD", "Any symbol"),
    ("Optimization Mode",       "Conservative/Balanced/Aggressive", "Manual filter toggle"),
    ("CSV logging",             "Yes",                          "Yes"),
    ("Screenshot logging",      "Yes",                          "Yes"),
    ("Dashboard panel",         "Yes — 7-step state machine",   "Yes — signal + stats"),
    ("Can run together?",       "YES — different magic numbers","YES — different magic numbers"),
    ("Best for",                "Quality ICT setups, patient trader",
                                "High frequency, active trading"),
    ("Recommended account",     "$500+ (selective entries)",    "$100+ (any size)"),
    ("Risk per trade",          "0.5% (ICT default)",           "0.5% (scalper default)"),
]

for feat, v1, v2 in comparisons:
    bg = C_ROW_A if r % 2 == 0 else C_ROW_B
    for ci, val in enumerate([feat, v1, v2], 1):
        c = ws3.cell(row=r, column=ci, value=val)
        cell_bg = C_GOLD_CELL if ci == 1 else (C_BLUE_CELL if ci == 2 else C_GREEN_CELL)
        c.fill      = make_fill(cell_bg)
        c.font      = make_font("1A1A3E", bold=(ci==1), size=9)
        c.alignment = make_left()
        c.border    = make_border()
    ws3.row_dimensions[r].height = 16
    r += 1

write_note_row(ws3, r,
    "  ⚑  Both EAs can run simultaneously on the same chart. They use different Magic Numbers (888777 vs 999888) so they never interfere.",
    N3); r+=1

# ── Save ──────────────────────────────────────────────────────────────
out = "/home/user/Rattana/EA_Input_Settings_Guide.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheet 1 rows: {ws1.max_row}  Sheet 2 rows: {ws2.max_row}  Sheet 3 rows: {ws3.max_row}")

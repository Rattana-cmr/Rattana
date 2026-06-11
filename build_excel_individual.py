#!/usr/bin/env python3
"""Generate individual Excel settings files for ICT SMC EA V1.0 – V1.4"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── style helpers ──────────────────────────────────────────────────────
def f(c):  return PatternFill(fill_type='solid', fgColor=c)
def fn(c='FFFFFF', b=False, sz=10): return Font(color=c, bold=b, size=sz, name='Calibri')
def al(h='left', v='center', w=True): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def bd(c='303050'):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)

HDR='0D2137'; GRP='12243A'; ODD='161625'; EVN='1C1C2E'
NEW='0A2A0A'; WARN='2A0A0A'
TW='FFFFFF'; TG='FFD700'; TGR='00DD77'; TR='FF6060'
TB='64AAFF'; TSL='B0B0C0'; TOR='FFA040'
C50='1C1C00'; C100='001C00'

def wc(ws, r, c, val, bg=ODD, fc=TW, bold=False, sz=10, h='left', wrap=True):
    cell = ws.cell(row=r, column=c, value=str(val) if val is not None else '')
    cell.fill = f(bg); cell.font = fn(fc, bold, sz)
    cell.alignment = al(h, 'center', wrap); cell.border = bd()

# ══════════════════════════════════════════════════════════════════════
# MASTER PARAMETER TABLE
# Each row: (group, name, type, comment_from_code, what_it_does,
#            v10, v11, v12, v13, v14,   <-- default per version (None = not present)
#            rec50, note50, rec100, note100, warning)
# ══════════════════════════════════════════════════════════════════════
PARAMS = [
  # ── TRADING STYLE ─────────────────────────────────────────────────
  ('TRADING STYLE','TradingStyle','Enum','Trading style preset',
   'Master preset that overrides many individual filters and thresholds.\n'
   '• CONSERVATIVE  — score≥80, all ICT strict, CD 30 min  → 1-5 trades/day\n'
   '• BALANCED      — respects input toggles, score≥70, CD 20 min  → 3-8 trades/day\n'
   '• AGGRESSIVE    — MSS/Trend OFF, OTE 55-90%, score≥50, CD 10 min  → 8-15 trades/day\n'
   '• SMART ACTIVE  (V1.3+) — all ICT ON, improved detection, score≥55  → 8-15 trades/day\n'
   '• SMART ACTIVE+ (V1.4)  — all ICT ON, OTE 55-90%, score≥45  → 15-25 trades/day\n'
   '• ULTRA ACTIVE  — filters minimal, max frequency',
   None, None, 'STYLE_BALANCED', 'STYLE_SMART_ACTIVE', 'STYLE_SMART_ACTIVE',
   'STYLE_CONSERVATIVE','Strictest entries, protects small $50 capital from overtrading.',
   'STYLE_BALANCED','Good balance of frequency and quality for $100.',
   'Overrides many inputs below. Settings marked * may not take effect.'),

  # ── RISK MANAGEMENT ───────────────────────────────────────────────
  ('RISK MANAGEMENT','RiskMode','Enum','Risk mode',
   'How position size is calculated:\n'
   '• RISK_FIXED_PCT  — lot = (balance × risk%) ÷ (SL pips × pip value)\n'
   '• RISK_FIXED_LOT  — always use FixedLot (ignores RiskPercent)\n'
   '• RISK_DYNAMIC_EQ — like FIXED_PCT but based on equity, not balance',
   None,'RISK_FIXED_PCT','RISK_FIXED_PCT','RISK_FIXED_PCT','RISK_FIXED_PCT',
   'RISK_FIXED_LOT','Fixed 0.01 lot on $50. % mode can calculate below broker minimum lot.',
   'RISK_FIXED_PCT','Scales correctly as $100 account grows.',
   'Not in V1.0'),

  ('RISK MANAGEMENT','RiskPercent','%','Risk per trade (%)',
   'Percentage of balance to risk per trade.\n'
   'Only used when RiskMode = RISK_FIXED_PCT or RISK_DYNAMIC_EQ.\n'
   'Example: $100 × 1.0% = $1.00 max risk per trade.',
   '0.5','0.5','0.5','0.5','0.5',
   '1.0','Not used in FIXED_LOT mode. Keep ≥1.0 as fallback.',
   '1.0','$100 × 1.0% = $1 risk per trade.',
   ''),

  ('RISK MANAGEMENT','FixedLot','lot','Fixed lot size (0 = use risk%)',
   'Fixed lot used when RiskMode = RISK_FIXED_LOT.\n'
   '0.0 = use RiskPercent calculation instead.\n'
   'Minimum lot on most brokers: 0.01',
   '0.0','0.0','0.0','0.0','0.0',
   '0.01','Set 0.01 + RiskMode=RISK_FIXED_LOT for predictable micro risk.',
   '0.01','Use 0.01 as safety floor, or 0.0 to let RiskPercent calculate.',
   ''),

  ('RISK MANAGEMENT','MaxDailyLossPercent','%','Max daily loss (%)',
   'EA stops trading for the day when daily loss reaches this % of balance.\n'
   '$50 × 6% = $3.00 daily stop  |  $100 × 6% = $6.00 daily stop.\n'
   'Resets at midnight broker time.',
   '10.0','10.0','10.0','10.0','10.0',
   '6.0','$50 × 6% = $3 max/day. Default 10% = $5 which is too large for a micro start.',
   '6.0','$100 × 6% = $6 daily stop. Keeps drawdown manageable.',
   ''),

  ('RISK MANAGEMENT','MaxTradesPerDay','int','Max trades per day',
   'Maximum new trades allowed per day. Counter resets at midnight.\n'
   'Match to your TradingStyle target:\n'
   '  Conservative 1-5 | Balanced 3-8 | SmartActive 8-15 | SmartActive+ 15-25',
   '10','10','25','25','25',
   '5','Limit to 5/day on $50 — reduces daily exposure on small capital.',
   '10','10/day for V1.0-V1.2. Raise to 15-20 for V1.3/V1.4 SmartActive.',
   ''),

  ('RISK MANAGEMENT','RewardRiskRatio','R:R','Reward/Risk ratio',
   'Target profit relative to stop loss. 2.0 = TP is 2× the SL distance.\n'
   'At 2:1 R:R you break even with only 34% win rate.',
   '2.0','2.0','2.0','2.0','2.0',
   '2.0','Never go below 2.0 on $50.',
   '2.0','Keep 2.0. Do not reduce below 1.5.',
   ''),

  ('RISK MANAGEMENT','MaxLotLimit','lot','Hard lot ceiling (scale with account)',
   'Hard ceiling on position size regardless of risk calculation.\n'
   'Prevents oversized positions from glitches or very tight SL signals.',
   '0.10','0.10','0.10','0.10','0.10',
   '0.01','Lock at 0.01 — never more than minimum lot on $50.',
   '0.02','0.02 allows slightly larger positions as $100 grows.',
   ''),

  ('RISK MANAGEMENT','MinRewardRiskRatio','R:R','Minimum R:R before placing trade',
   'Trade skipped if actual calculated R:R is below this value.\n'
   'Rejects setups where SL is too wide vs target TP.\n'
   'Small float tolerance (-0.001) avoids rounding false-skips.',
   '2.0','2.0','2.0','2.0','2.0',
   '2.0','Always require 2:1 minimum.',
   '1.8','Can relax to 1.8 to allow slightly more setups.',
   ''),

  # ── TAKE PROFIT MODE ──────────────────────────────────────────────
  ('TAKE PROFIT MODE','TPMode','Enum','TP calculation mode',
   'How take profit is calculated:\n'
   '• TP_FIXED_RR — TP = entry ± (SL distance × RewardRiskRatio)\n'
   '• TP_ATR      — TP = entry ± (ATR(14) × ATRMultiplierTP)\n'
   '• TP_HYBRID   — TP = min(FixedRR, ATR)  most conservative',
   None,'TP_FIXED_RR','TP_FIXED_RR','TP_FIXED_RR','TP_FIXED_RR',
   'TP_FIXED_RR','Predictable and simple. Best for small accounts.',
   'TP_FIXED_RR','Keep Fixed R:R for consistent trade math.',
   'Not in V1.0'),

  ('TAKE PROFIT MODE','ATRMultiplierTP','×','ATR multiplier (TP_ATR / TP_HYBRID)',
   'ATR multiplier used only when TPMode = TP_ATR or TP_HYBRID.\n'
   'TP = current ATR(14) × this value. Adapts to market volatility.',
   None,'3.0','3.0','3.0','3.0',
   '3.0','Keep default. Only relevant if you use TP_ATR mode.',
   '3.0','Keep default.',
   'Not in V1.0. Only used when TPMode ≠ TP_FIXED_RR'),

  # ── PARTIAL TAKE PROFIT ───────────────────────────────────────────
  ('PARTIAL TAKE PROFIT','UsePartialTP','bool','Close part at 1R, rest to full TP',
   'Closes PartialClosePercent% of position at PartialCloseRR × R:R,\n'
   'then moves SL to breakeven for the remaining portion.\n'
   'Example: 50% closed at 1R → remaining 50% runs risk-free to 2R TP.',
   None,'true','true','true','true',
   'true','Highly recommended for $50 — locks in 50% profit at 1R, remainder runs free.',
   'true','Recommended for all account sizes.',
   'Not in V1.0'),

  ('PARTIAL TAKE PROFIT','PartialClosePercent','%','% of position to close at 1R',
   'Percentage of the open position to close at the partial TP level.\n'
   '50.0 = close half the position when 1R profit is reached.',
   None,'50.0','50.0','50.0','50.0',
   '50.0','Standard 50% partial close.',
   '50.0','Keep 50%.',
   ''),

  ('PARTIAL TAKE PROFIT','PartialCloseRR','R','R:R level to trigger partial close',
   'R:R ratio at which partial close triggers.\n'
   '1.0 = close partial when profit equals 1× the SL distance.\n'
   'After partial close, SL moves to breakeven.',
   None,'1.0','1.0','1.0','1.0',
   '1.0','Trigger at 1R — ensures 50% is locked even if remainder hits breakeven.',
   '1.0','Standard. Adjust to 1.5 to run a bit further before locking.',
   ''),

  # ── TRADE FILTERS ─────────────────────────────────────────────────
  ('TRADE FILTERS','UseTimeFilter','bool','Use trading hours',
   'When true: EA only trades during active sessions configured below.\n'
   'When false: EA trades 24/5 regardless of time.',
   'true','true','true','true','true',
   'true','Always ON — avoids low-liquidity hours.',
   'true','Keep ON.',
   ''),

  ('TRADE FILTERS','MaxSpreadPoints','pts','Max spread (0 = disable)',
   'Maximum spread (bid-ask) in broker points allowed before a trade.\n'
   'XAUUSD typical spread: 20-40 pts normal, 80-200 pts during news.\n'
   '0 = disable check.',
   '50','50','80','80','80',
   '80','Covers normal gold spreads. 80 pts raised from 50 in V1.2+.',
   '80','Appropriate for XAUUSD.',
   'Default is 50 in V1.0-V1.1, raised to 80 in V1.2+'),

  ('TRADE FILTERS','MinStopDistance','pts','Min SL distance in points',
   'Minimum SL distance from entry in broker points.\n'
   'Brokers have a freeze level — orders too close to price are rejected.',
   '20','20','20','20','20',
   '20','Keep default 20 points.',
   '20','Keep default.',
   ''),

  ('TRADE FILTERS','MaxConsecutiveLosses','int','Halt after N consecutive losses',
   'EA stops trading after N consecutive losing trades.\n'
   'Counter resets daily if ResetLossStreakDaily = true.',
   '10','10','10','10','10',
   '5','Lower to 5 on $50 — stops EA before 5 losses damage the account.',
   '7','7 consecutive losses is a solid circuit breaker for $100.',
   ''),

  ('TRADE FILTERS','MaxDailyLossTrades','int','Max losing trades per day',
   'Stop trading for the day after N losing trades.\n'
   'FIX T: prevents clusters of consecutive daily losses.',
   '3','3','3','3','3',
   '2','Stop after 2 losing trades on $50 — preserves capital for tomorrow.',
   '3','Standard 3 losing-trade daily stop.',
   ''),

  ('TRADE FILTERS','ResetLossStreakDaily','bool','Reset streak daily (true) or lifetime (false)',
   'true = MaxConsecutiveLosses counter resets every midnight.\n'
   'false = counter accumulates for the lifetime of the EA session.',
   'true','true','true','true','true',
   'true','Reset daily — fresh start each morning.',
   'true','Keep true.',
   ''),

  # ── ICT STRUCTURE FILTERS ─────────────────────────────────────────
  ('ICT STRUCTURE FILTERS','UseMSSFilter','bool','Require MSS (H1) before entry',
   'Require a Market Structure Shift on H1 before entering.\n'
   'MSS = H1 price breaks above a previous swing high (bullish)\n'
   'or below a swing low (bearish) — confirms direction change.\n'
   '* Overridden OFF by STYLE_AGGRESSIVE and STYLE_ULTRA_ACTIVE.',
   None,'true','true','true','true',
   'true','Keep ON — MSS confirms the move is genuine, not noise.',
   'true','Always ON for quality signals.',
   'Not in V1.0. Overridden by TradingStyle in V1.2+'),

  ('ICT STRUCTURE FILTERS','UseBOSFilter','bool','Require BOS (M15) confirmation',
   'Require Break of Structure on M15 for lower-timeframe confirmation.\n'
   'BOS = M15 price breaks above recent swing high after bullish MSS.',
   None,'true','true','true','true',
   'true','Keep ON.',
   'true','Keep ON.',
   'Not in V1.0'),

  ('ICT STRUCTURE FILTERS','RequireLiquiditySweep','bool','Require liquidity sweep',
   'Require a liquidity sweep (stop hunt) before entry.\n'
   'false = optional (recommended). true = must detect sweep — very rare,\n'
   'reduces frequency to ~1-2 trades per week.',
   None,'false','false','false','false',
   'false','Keep false — enabling reduces trades dramatically.',
   'false','Keep false unless you want very high-rarity setups.',
   'Not in V1.0'),

  ('ICT STRUCTURE FILTERS','UseSMTFilter','bool','SMT divergence (default OFF)',
   'SMT divergence filter. Compares price structure between this symbol\n'
   'and SMTSymbol. BUY if this symbol makes lower low but SMTSymbol does not.\n'
   'Usually: XAUUSD vs XAGUSD.',
   None,'false','false','false','false',
   'false','Keep OFF for simplicity.',
   'false','Keep OFF unless you specifically trade SMT setups.',
   'Not in V1.0'),

  ('ICT STRUCTURE FILTERS','SMTSymbol','string','Correlated symbol for SMT',
   'Correlated symbol used for SMT divergence. Only active when UseSMTFilter=true.\n'
   'Gold → XAGUSD (silver).',
   None,'XAGUSD','XAGUSD','XAGUSD','XAGUSD',
   'XAGUSD','Keep default.',
   'XAGUSD','Keep default.',
   'Symbol must exist on your broker'),

  # ── NEWS FILTER ───────────────────────────────────────────────────
  ('NEWS FILTER','UseNewsFilter','bool','Block trading during high-impact news',
   'Block trading before/after high-impact economic news.\n'
   'Uses MT5 Calendar API. Requires Allow WebRequest in MT5.',
   None,'false','false','false','false',
   'false','Disable in Strategy Tester. Enable for live trading.',
   'false','Optional but recommended for live (NFP, FOMC, CPI).',
   'Not in V1.0. Requires Allow WebRequest enabled in MT5'),

  ('NEWS FILTER','NewsBlockBeforeMin','min','Block N min before news',
   'Minutes before a high-impact news event to stop accepting new trades.',
   None,'30','30','30','30',
   '30','Standard 30-min pre-news block.',
   '30','Keep 30 minutes.',
   ''),

  ('NEWS FILTER','NewsBlockAfterMin','min','Block N min after news',
   'Minutes after a high-impact news event before resuming trades.\n'
   'Spreads remain elevated for 15-30 min post-news.',
   None,'30','30','30','30',
   '30','30-min post-news cooldown.',
   '30','Keep 30 minutes.',
   ''),

  # ── TRADE QUALITY SCORE ───────────────────────────────────────────
  ('TRADE QUALITY SCORE','UseTradeScore','bool','Enable minimum score gate',
   'When true: trade is skipped if score < MinimumTradeScore.\n'
   'When false: score shown on panel but never blocks trades.\n'
   'Only in V1.2+. In V1.1 set MinimumTradeScore=0 to disable.',
   None,None,'true','true','true',
   'true','Keep ON — score filter is your quality gate on small accounts.',
   'true','Keep ON.',
   'Not in V1.0 or V1.1'),

  ('TRADE QUALITY SCORE','MinimumTradeScore','0-100','Min score/100 to allow trade',
   'Minimum score (0-100) required to place a trade.\n'
   'Score built from: session quality, spread, OTE precision,\n'
   'liquidity presence, FVG count, D1/H4 trend alignment.\n'
   'Overridden by TradingStyle preset in V1.2+.',
   None,'70','70','70','70',
   '70','70 filters low-quality setups without being too restrictive.',
   '65','Slightly relax to 65 for more trades on $100.',
   'Not in V1.0. Overridden by TradingStyle in V1.2+'),

  # ── SESSIONS ──────────────────────────────────────────────────────
  ('SESSIONS (GMT)','AutoDetectGMT','bool','Auto-detect GMT offset (DST-aware)',
   'Auto-detect broker GMT offset: offset = (TimeCurrent - TimeGMT) / 3600.\n'
   'Works correctly on live. In Strategy Tester always shows GMT+0 — this is normal.',
   None,'true','true','true','true',
   'true','Keep ON for live trading.',
   'true','Keep ON.',
   'Not in V1.0. Always GMT+0 in Strategy Tester — expected, not a bug'),

  ('SESSIONS (GMT)','BrokerGMTOffset','hrs','Broker GMT offset',
   'Manual GMT offset when AutoDetectGMT = false.\n'
   'EU brokers: GMT+2 (winter) / GMT+3 (summer DST).\n'
   'In V1.0 this is always manual (no auto-detect).',
   '0','0','0','0','0',
   '0','Leave 0 when AutoDetectGMT = true.',
   '0','Leave 0 when AutoDetectGMT = true.',
   ''),

  ('SESSIONS (GMT)','SessionSydney','bool','Sydney (22:00-07:00 GMT)',
   'Allow trading during Sydney session. Low XAUUSD volume.',
   'false','false','false','false','false',
   'false','Keep OFF — low gold liquidity.',
   'false','Keep OFF.',
   ''),

  ('SESSIONS (GMT)','SessionTokyo','bool','Tokyo (00:00-09:00 GMT)',
   'Allow trading during Tokyo session. Moderate XAUUSD activity.',
   'false','false','false','false','false',
   'false','Keep OFF on $50.',
   'false','Can enable for V1.4 SmartActive+ expanded coverage.',
   ''),

  ('SESSIONS (GMT)','SessionLondon','bool','London (08:00-17:00 GMT)',
   'Allow trading during London session. Highest liquidity period.\n'
   'Best session for XAUUSD and FX.',
   'true','true','true','true','true',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','SessionNewYork','bool','New York (13:00-22:00 GMT)',
   'Allow trading during New York session. High volatility, strong moves.',
   'true','true','true','true','true',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','OverlapLondonNY','bool','London+NY Overlap (13:00-17:00 GMT)',
   'Allow trading during London/NY overlap. Best 4 hours of the day\n'
   '— highest liquidity AND volatility combined.',
   'true','true','true','true','true',
   'true','Definitely ON — highest quality ICT entries here.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','OverlapTokyoLondon','bool','Tokyo+London Overlap (08:00-09:00 GMT)',
   'Allow trading during Tokyo/London overlap. Only 1 hour.\n'
   'Already covered by SessionLondon.',
   'false','false','false','false','false',
   'false','Keep OFF.',
   'false','Keep OFF.',
   ''),

  # ── STOP LOSS ─────────────────────────────────────────────────────
  ('STOP LOSS','SLBufferPips','pips','SL buffer in pips behind swing',
   'Extra pip buffer beyond the swing high/low for SL placement.\n'
   'Example: swing low at 2000.00, buffer=15 → SL ~15 pips below.\n'
   'Prevents SL from being hit by noise at the exact swing level.',
   '15','15','15','15','15',
   '10','Reduce to 10 pips on $50 — tightens SL and reduces $ loss per trade.',
   '15','Default 15 pips provides good buffer against stop hunts.',
   ''),

  ('STOP LOSS','UseTrailingStop','bool','Enable trailing stop',
   'When true: SL moves up (BUY) or down (SELL) after TrailingStartPips profit.\n'
   'UsePartialTP is a more consistent alternative.',
   'false','false','false','false','false',
   'false','Keep OFF — UsePartialTP at 1R is more effective.',
   'false','Keep OFF unless backtest proves improvement.',
   ''),

  ('STOP LOSS','TrailingStartPips','pips','Start trailing after N pips profit',
   'Profit in pips before trailing stop activates.\n'
   'Only used when UseTrailingStop = true.',
   '30','30','30','30','30',
   '30','Only relevant if UseTrailingStop = true.',
   '30','Only relevant if UseTrailingStop = true.',
   'Only used when UseTrailingStop = true'),

  ('STOP LOSS','TrailingStepPips','pips','Trail by N pips',
   'How many pips to trail behind current price.\n'
   'Smaller = tighter; Larger = more room.',
   '10','10','10','10','10',
   '10','Only relevant if UseTrailingStop = true.',
   '10','Only relevant if UseTrailingStop = true.',
   'Only used when UseTrailingStop = true'),

  # ── POSITION MANAGEMENT ───────────────────────────────────────────
  ('POSITION MANAGEMENT','CloseOnFriday','bool','Close positions on Friday',
   'Auto-close all open positions on Friday at FridayCloseHour GMT.\n'
   'Prevents weekend gap risk — price can jump 50-200 pips over weekends.',
   'true','true','true','true','true',
   'true','Always ON — weekend gaps can easily wipe $50.',
   'true','Always ON.',
   ''),

  ('POSITION MANAGEMENT','FridayCloseHour','GMT hr','Friday close at 14:00 GMT',
   'GMT hour on Friday when positions are closed.\n'
   '14:00 GMT = before thin NY afternoon and approaching weekend gap risk.',
   '14','14','14','14','14',
   '14','14:00 GMT is optimal.',
   '14','Keep 14:00 GMT.',
   ''),

  ('POSITION MANAGEMENT','UseBreakeven','bool','Move SL to breakeven',
   'Move SL to entry price after BreakevenTriggerPips profit.\n'
   'UsePartialTP is a more elegant alternative.',
   'false','false','false','false','false',
   'false','Keep OFF — UsePartialTP at 1R is better.',
   'false','Keep OFF.',
   'Conflicts with UsePartialTP'),

  ('POSITION MANAGEMENT','BreakevenTriggerPips','pips','Pips profit to trigger breakeven',
   'Pips of profit required before SL moves to breakeven.\n'
   'Only used when UseBreakeven = true.',
   '40','40','40','40','40',
   '30','Only relevant if UseBreakeven = true.',
   '40','Only relevant if UseBreakeven = true.',
   'Only used when UseBreakeven = true'),

  # ── SWING DETECTION ───────────────────────────────────────────────
  ('SWING DETECTION','SwingLookbackBarsH1','bars','H1 bars to scan (OTE)',
   'H1 bars to scan when finding swing highs/lows for OTE zone.\n'
   '50 bars ≈ 2 days of H1 history.',
   '50','50','50','50','50',
   '50','Keep default.',
   '50','Keep default.',
   ''),

  ('SWING DETECTION','SwingConfirmBarsH1','bars','Bars each side to confirm H1 swing',
   'Bars on each side required to confirm an H1 swing.\n'
   '3 = swing must be highest/lowest of 7 bars total.\n'
   'In V1.3+ overridden to 2 (SmartActive) or 1 (SmartActive+) by style.',
   '3','3','3','3','3',
   '3','Default fine — overridden by TradingStyle in V1.3+.',
   '3','Keep default.',
   'Overridden by effMSSConfirm in V1.3+'),

  ('SWING DETECTION','SwingLookbackBarsM15','bars','M15 bars to scan (SL)',
   'M15 bars to scan for the SL anchor swing.\n'
   '30 bars ≈ 7.5 hours of M15.',
   '30','30','30','30','30',
   '30','Keep default.',
   '30','Keep default.',
   ''),

  ('SWING DETECTION','SwingConfirmBarsM15','bars','Bars each side to confirm M15 swing',
   'Bars each side to confirm an M15 swing (used for SL placement).\n'
   '5 = swing must be highest/lowest of 11 bars.',
   '5','5','5','5','5',
   '5','Keep default.',
   '5','Keep default.',
   ''),

  ('SWING DETECTION','MaxSwingDistancePips','pips','Max swing distance in pips (0=disabled)',
   'Maximum allowed H1 swing range in pips.\n'
   'Rejects abnormally large swings. 0 = no maximum.',
   '500','500','500','500','500',
   '300','Tighter at 300 pips on $50.',
   '500','Keep default.',
   ''),

  ('SWING DETECTION','MaxSLPips','pips','Max SL in pips',
   'Maximum SL distance in pips. Trade skipped if SL > this.\n'
   'FIX O: prevents large SL from consuming disproportionate % of small accounts.',
   '30','30','30','30','30',
   '20','20 pips max on $50 keeps loss per trade tightly controlled.',
   '25','25-30 pips appropriate for $100.',
   ''),

  ('SWING DETECTION','MinSLPips','pips','Min SL in pips',
   'Minimum SL distance in pips.\n'
   'FIX Q: skips trades with SL so tight it lives inside the spread.',
   '10','10','10','10','10',
   '10','Keep default.',
   '10','Keep default.',
   ''),

  ('SWING DETECTION','ShowSwingLines','bool','Draw M15 swing SL lines',
   'Draw M15 swing high/low lines on chart.\n'
   'Shows where EA anchored the SL.',
   'true','true','true','true','true',
   'true','Keep ON for visual monitoring.',
   'true','Keep ON.',
   ''),

  # ── ICT TWINS MODEL ───────────────────────────────────────────────
  ('ICT TWINS MODEL','UseTwinsModel','bool','Enable full ICT TWINS model',
   'Enable the full ICT Twins sequential entry model.\n'
   'Steps: HTF Level → MSS → BOS → Liquidity → OTE Zone → CISD → 1M Trigger.\n'
   'This IS the core strategy. Must always be ON.',
   'true','true','true','true','true',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('ICT TWINS MODEL','HTFLevelMinutes','min','HTF timeframe (15, 30, 60)',
   'Timeframe for HTF reference level detection.\n'
   '15 = M15. EA checks proximity to M15 FVGs, prev-day H/L, H1/H4 levels.',
   '15','15','15','15','15',
   '15','Keep default M15.',
   '15','Keep default.',
   ''),

  ('ICT TWINS MODEL','OTEMinPercent','0-1','OTE minimum (was 0.62, now 0.65)',
   'OTE zone minimum as % of H1 swing range.\n'
   '0.65 = 65% retracement into the swing.\n'
   'Overridden by TradingStyle in V1.2+.',
   '0.65','0.65','0.65','0.65','0.65',
   '0.65','Keep default. Overridden by TradingStyle in V1.2+ anyway.',
   '0.65','Keep default.',
   'Overridden by TradingStyle in V1.2+'),

  ('ICT TWINS MODEL','OTEMaxPercent','0-1','OTE maximum (was 0.79, now 0.75)',
   'OTE zone maximum as % of H1 swing range.\n'
   '0.75 = creates classic 65-75% OTE window (10% wide).\n'
   'V1.3 SmartActive widens to 60-85%; V1.4 SmartActive+ widens to 55-90%.',
   '0.75','0.75','0.75','0.75','0.75',
   '0.75','Keep default.',
   '0.75','Keep default.',
   'Overridden by TradingStyle in V1.2+'),

  ('ICT TWINS MODEL','OTESweetSpotPercent','0-1','OTE sweet spot (70.5%)',
   'The golden ratio point within the OTE zone (70.5%).\n'
   'Entries near 70.5% get bonus quality score points.',
   '0.705','0.705','0.705','0.705','0.705',
   '0.705','Keep default.',
   '0.705','Keep default.',
   ''),

  ('ICT TWINS MODEL','MinFVGsRequired','int','Min 1-min FVGs required (0=disabled)',
   'Minimum 1-minute Fair Value Gaps required near entry.\n'
   '0 = FVG check disabled (recommended).',
   '0','0','0','0','0',
   '0','Keep 0 — FVG requirement reduces frequency too much.',
   '0','Keep 0.',
   ''),

  ('ICT TWINS MODEL','HTFToleranceATRMulti','×','HTF tolerance = N x ATR(14)',
   'Price must be within N × ATR(14) of an HTF level to qualify.\n'
   '2 = within 2 ATRs of the reference level.',
   '2','2','2','2','2',
   '2','Default 2 ATRs is balanced.',
   '2','Keep default.',
   ''),

  ('ICT TWINS MODEL','HTFLevelRequired','bool','Require HTF key level',
   'Require price near an HTF reference level before the MSS step.\n'
   'FIX U: default false (auto-pass Step 1).\n'
   'true = strictest mode. In trending markets EA may not trade for days.',
   'false','false','false','false','false',
   'false','Keep false. Setting true on trending gold = EA stops trading.',
   'false','Keep false.',
   'Setting true on strong trends = EA will not trade'),

  ('ICT TWINS MODEL','ShowOTEZone','bool','Draw OTE zone on chart',
   'Draw OTE zone rectangle on chart.\n'
   'Shows exactly where EA expects price to enter.',
   'true','true','true','true','true',
   'true','Keep ON for visual confirmation.',
   'true','Keep ON.',
   ''),

  ('ICT TWINS MODEL','MinH1RangePips','pips','Min H1 swing range in pips (0=disabled)',
   'Minimum H1 swing range (high to low) in pips.\n'
   'Filters out sideways/choppy markets. 0 = no minimum.',
   '50','50','50','50','50',
   '30','Relax to 30 pips on $50 — 50 pip default reduces frequency too much.',
   '40','40 pips balances quality vs frequency for $100.',
   ''),

  ('ICT TWINS MODEL','UseH1RangeFilter','bool','Toggle H1 range check',
   'Toggle the H1 range minimum pip check.\n'
   'false = skip MinH1RangePips check entirely.',
   None,None,'true','true','true',
   'true','Keep ON.',
   'true','Keep ON.',
   'Not in V1.0 and V1.1'),

  ('ICT TWINS MODEL','MinH1RangeATRMulti','×','H1 range >= N * H1-ATR (SmartActive)',
   'H1 swing range must be ≥ (H1 ATR × this multiplier).\n'
   'Active in STYLE_SMART_ACTIVE and STYLE_SMART_ACTIVE_PLUS.\n'
   'Replaces fixed pip check with a volatility-adaptive check.\n'
   '0.8 = range ≥ 80% of one H1 ATR period.',
   None,None,None,'0.8','0.8',
   '0.8','Keep default. Filters extreme chop while allowing normal volatility.',
   '0.7','Relax slightly to 0.7 for slightly more trades on $100.',
   'Only in V1.3 and V1.4'),

  # ── MSS / BOS / LIQUIDITY ─────────────────────────────────────────
  ('MSS / BOS / LIQUIDITY','MSSLookbackBars','bars','H1 bars to scan for MSS',
   'H1 bars to look back when detecting Market Structure Shift.\n'
   '30 H1 bars ≈ 30 hours of history.',
   None,'30','30','30','30',
   '30','Keep default.',
   '30','Keep default.',
   'Not in V1.0'),

  ('MSS / BOS / LIQUIDITY','MSSConfirmBars','bars','Swing confirm bars (each side)',
   'Bars each side to confirm an H1 swing for MSS detection.\n'
   'In V1.3+: overridden to 2 (SmartActive) or 1 (SmartActive+) by style.',
   None,'3','3','3','3',
   '3','Default — overridden by TradingStyle anyway.',
   '3','Keep default.',
   'Not in V1.0. Overridden by effMSSConfirm in V1.3+'),

  ('MSS / BOS / LIQUIDITY','BOSLookbackBars','bars','M15 bars to scan for BOS',
   'M15 bars to look back when detecting Break of Structure.\n'
   '20 M15 bars ≈ 5 hours of M15.',
   None,'20','20','20','20',
   '20','Keep default.',
   '20','Keep default.',
   'Not in V1.0'),

  ('MSS / BOS / LIQUIDITY','LiquidityLookbackBars','bars','M15 bars to scan for sweep',
   'M15 bars to scan for liquidity levels (swing highs/lows).\n'
   'Used to detect if price swept these levels before entry.',
   None,'50','50','50','50',
   '50','Keep default.',
   '50','Keep default.',
   'Not in V1.0'),

  ('MSS / BOS / LIQUIDITY','LiquidityWickPips','pips','Min wick beyond liquidity level',
   'Minimum wick size beyond a liquidity level to confirm a sweep.\n'
   'Wick ≥ 3 pips beyond swing = liquidity sweep confirmed.',
   None,'3','3','3','3',
   '3','Keep default 3 pips.',
   '3','Keep default.',
   'Not in V1.0'),

  # ── SYMBOL PRESET ─────────────────────────────────────────────────
  ('SYMBOL PRESET','SymbolPreset','Enum','Auto-detect from symbol name',
   'Pre-configured OTE and SL parameters for known symbols:\n'
   '• PRESET_AUTO   — auto-detects from symbol name (recommended)\n'
   '• PRESET_XAUUSD — Gold: OTE 65-75%, MaxSL 30 pips\n'
   '• PRESET_BTCUSD — Bitcoin: OTE 62-78%, MaxSL 80 pips\n'
   '• PRESET_EURUSD — Euro: OTE 62-79%, MaxSL 25 pips\n'
   '• PRESET_GBPUSD — Pound: OTE 62-79%, MaxSL 30 pips',
   None,'PRESET_AUTO','PRESET_AUTO','PRESET_AUTO','PRESET_AUTO',
   'PRESET_AUTO','Auto detects XAUUSD from symbol name.',
   'PRESET_AUTO','Keep auto.',
   'Not in V1.0'),

  # ── OPTIMIZATION MODE ─────────────────────────────────────────────
  ('OPTIMIZATION MODE','OptMode','Enum','Conservative / Balanced / Aggressive',
   'Secondary optimisation layer (applied after TradingStyle in V1.2+):\n'
   '• OPT_CONSERVATIVE — score threshold +10, risk 0.25%\n'
   '• OPT_BALANCED     — no change from defaults\n'
   '• OPT_AGGRESSIVE   — score threshold -15, risk ×2',
   None,'OPT_BALANCED','OPT_BALANCED','OPT_BALANCED','OPT_BALANCED',
   'OPT_CONSERVATIVE','Extra score filtering helps on $50.',
   'OPT_BALANCED','Balanced is fine for $100.',
   'Not in V1.0. In V1.2+ TradingStyle takes precedence'),

  # ── LOGGING ───────────────────────────────────────────────────────
  ('LOGGING','EnableScreenshot','bool','Save screenshot on trade open/close',
   'Save a chart screenshot on every trade open and close.\n'
   'Files saved to MT5/MQL5/Files/ folder.',
   None,'true','true','true','true',
   'true','Keep ON for trade journal.',
   'true','Keep ON.',
   'Not in V1.0'),

  ('LOGGING','EnableCSVLog','bool','Export trades to CSV',
   'Export every trade to a CSV file.\n'
   'Columns: time, symbol, direction, entry, SL, TP, lot, score, result.',
   None,'true','true','true','true',
   'true','Keep ON — CSV lets you analyse patterns.',
   'true','Keep ON.',
   'Not in V1.0'),

  # ── DEBUG / ADVANCED ──────────────────────────────────────────────
  ('DEBUG / ADVANCED','PostTradeCooldownMin','min','Minutes to wait after trade',
   'Minutes to wait after closing a trade before next entry.\n'
   'Prevents back-to-back trades in the same move.\n'
   'Overridden by TradingStyle: Conservative=30, Balanced=20, SmartActive=10, SmartActive+=5.',
   '30','30','20','20','20',
   '30','30 min on V1.0-V1.1. Let TradingStyle control on V1.2+.',
   '20','20 min default on V1.2+.',
   'Overridden by TradingStyle in V1.2+'),

  ('DEBUG / ADVANCED','UseDailyTrendFilter','bool','Only trade in D1 trend direction',
   'Only trade in the direction of the D1 trend.\n'
   'Uses D1 candle direction + H4 50-EMA + H4 200-EMA alignment.\n'
   'BUY requires 3/3 bullish D1 candles (FIX M).',
   'true','true','true','true','true',
   'true','Keep ON — trading with D1 trend is core ICT principle.',
   'true','Keep ON.',
   'Overridden to OFF by STYLE_AGGRESSIVE'),

  ('DEBUG / ADVANCED','BestHoursOnly','bool','Trade 08:30-15:00 GMT',
   'FIX N: Restrict trading to 08:30-15:00 GMT.\n'
   'Highest-quality ICT entry hours.\n'
   'Overridden to false by STYLE_SMART_ACTIVE_PLUS (V1.4) for all-session coverage.',
   'true','true','true','true','true',
   'true','Keep ON for $50 — focus on best-quality hours only.',
   'true','Keep ON unless using V1.4 SmartActive+.',
   'Overridden to false by STYLE_SMART_ACTIVE_PLUS in V1.4'),

  ('DEBUG / ADVANCED','ForceTrades','bool','NEVER true on live accounts',
   '⚠ TESTING ONLY: bypasses ALL filters and forces trade entries.\n'
   'Used in development to verify order execution.\n'
   'NEVER set to true on live accounts or real backtests.',
   'false','false','false','false','false',
   'false','ALWAYS false. Never change this on live.',
   'false','ALWAYS false.',
   '⚠ DANGER: Setting true = random trades with no logic applied'),

  ('DEBUG / ADVANCED','DebugMode','bool','Verbose logging',
   'Enable verbose logging to MT5 journal.\n'
   'Prints step-by-step logic for every tick.\n'
   'Useful for diagnosing why EA is not trading.',
   'false','false','false','false','false',
   'false','Keep OFF in production. Enable temporarily to diagnose issues.',
   'false','Keep OFF.',
   'Generates massive journal output — can slow Strategy Tester'),

  ('DEBUG / ADVANCED','RelaxedMode','bool','Relaxed testing',
   'Testing mode applying relaxed parameters.\n'
   'Only for verifying setup works in Strategy Tester.\n'
   'Not for real backtests or live trading.',
   'false','false','false','false','false',
   'false','Always false in production.',
   'false','Always false.',
   ''),
]

# ── version metadata ───────────────────────────────────────────────────
VERSIONS = {
  'V1.0': {
    'name':'ICT SMC EA V1.0',
    'color':'1565C0', 'tab':'174080',
    'default_style':'(no preset)',
    'new_features':[
      'Base ICT Twins model: HTF Level → OTE → CISD → 1M Trigger',
      'H1 + M15 dual-timeframe swing detection',
      'D1 trend filter (3/3 bullish candles for BUY — FIX M)',
      'BestHoursOnly: 08:30-15:00 GMT (FIX N)',
      'OTE zone tightened to 65-75% (FIX J)',
      'Friday cut-off at 14:00 GMT (FIX P)',
      'MaxSLPips = 30, MinSLPips = 10 (FIX O / FIX Q)',
      'MaxDailyLossTrades circuit breaker (FIX T)',
      'HTFLevelRequired = false auto-pass (FIX U)',
      'Draggable dashboard panel',
    ],
    'param_col':5,  # index in (grp,name,type,comment,what,v10,v11,v12,v13,v14,...)
  },
  'V1.1': {
    'name':'ICT SMC EA V1.1',
    'color':'00695C', 'tab':'004D40',
    'default_style':'(no preset)',
    'new_features':[
      'Market Structure Shift (MSS) filter on H1',
      'Break of Structure (BOS) filter on M15',
      'Liquidity Sweep detection',
      'SMT Divergence filter (optional)',
      'Economic News Filter (MT5 Calendar API)',
      'Dynamic Risk Mode: FIXED_PCT / FIXED_LOT / DYNAMIC_EQ',
      'Partial Take Profit: 50% at 1R → SL to breakeven',
      'ATR Dynamic TP: FIXED_RR / ATR / HYBRID',
      'Broker Protection Layer (stops, freeze, margin check)',
      'Auto GMT Detection (DST-aware)',
      'Trade Quality Score (0-100)',
      'Symbol Presets: XAUUSD / BTCUSD / EURUSD / GBPUSD',
      'Optimization Mode: Conservative / Balanced / Aggressive',
      'CSV Performance Log + Screenshot on every trade',
      'Advanced Dashboard V1.1 (7-step sequence + score)',
    ],
    'param_col':6,
  },
  'V1.2': {
    'name':'ICT SMC EA V1.2',
    'color':'558B2F', 'tab':'33691E',
    'default_style':'STYLE_BALANCED',
    'new_features':[
      'TradingStyle preset: Conservative / Balanced / Aggressive / Ultra Active',
      'Filter Rejection Statistics — per-filter blocked count on panel',
      'UseTradeScore toggle (can fully disable score gate)',
      'UseH1RangeFilter toggle (can disable H1 range minimum)',
      'Effective override system (style overrides individual inputs)',
      'MaxSpreadPoints raised 50→80 for XAUUSD',
      'MaxTradesPerDay raised 10→25 for UltraActive',
      'PostTradeCooldownMin changed 30→20 base',
    ],
    'param_col':7,
  },
  'V1.3': {
    'name':'ICT SMC EA V1.3',
    'color':'E65100', 'tab':'BF360C',
    'default_style':'STYLE_SMART_ACTIVE',
    'new_features':[
      'STYLE_SMART_ACTIVE: all ICT filters ON, improved detection, target 8-15 trades/day',
      'Improved MSS detection: effMSSConfirm=2 bars, current-bar check, displacement candle',
      'Improved BOS detection: effBOSConf=2 bars, current-bar check',
      'Enhanced 1M entry trigger: pin-bar + 3-bar momentum + adaptive effBodyThresh=0.60',
      'OTE SmartActive: 60-85% (wider window, vs default 65-75%)',
      'ATR-adaptive H1 range filter: MinH1RangeATRMulti input added',
      'Cumulative filter rejection counters (never reset — full backtest view)',
      'PrintFilterSummary() printed to journal on deinit with TOP BLOCKER identification',
      'Daily rejection stats reset each day; cumulative stats span entire backtest',
    ],
    'param_col':8,
  },
  'V1.4': {
    'name':'ICT SMC EA V1.4',
    'color':'6A1B9A', 'tab':'4A148C',
    'default_style':'STYLE_SMART_ACTIVE',
    'new_features':[
      'STYLE_SMART_ACTIVE_PLUS: all ICT ON, OTE 55-90%, score≥45, target 15-25 trades/day',
      'MSS confirmation: 1 bar (fastest valid structural confirmation)',
      'BOS confirmation: 1 bar (earliest valid BOS)',
      'BestHoursOnly = false — covers London, New York, Tokyo, Sydney sessions',
      'ATR range threshold relaxed to 0.6× H1-ATR (vs 0.8× for SmartActive)',
      'Body threshold: 55% for CISD/momentum triggers',
      'Cooldown: 5 minutes (vs 10 min for SmartActive)',
      'Panel: SmartActive+ shown in blue, SmartActive in green, UltraActive in red',
    ],
    'param_col':9,
  },
}

# ── build one Excel per version ────────────────────────────────────────
for ver_key, vmeta in VERSIONS.items():
    vcol = vmeta['param_col']   # index 5-9 in PARAMS tuple for default value
    out_path = f"/home/user/Rattana/ICT_SMC_EA_{ver_key}_Settings.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ─────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = f"{ver_key} Overview"
    ws_ov.sheet_properties.tabColor = vmeta['tab']
    ws_ov.column_dimensions['A'].width = 80

    def ov_row(ws, r, text, bg=ODD, fc=TW, bold=False, sz=10):
        c = ws.cell(row=r, column=1, value=text)
        c.fill = f(bg); c.font = fn(fc, bold, sz)
        c.alignment = al('left','center',False); ws.row_dimensions[r].height = 17

    ov_row(ws_ov,1, vmeta['name'], HDR, TG, True, 16)
    ov_row(ws_ov,2, 'ICT Smart Money Concepts EA  |  Created by RATTANA CHHORM', HDR, TSL, False, 11)
    ov_row(ws_ov,3, '', HDR)
    ov_row(ws_ov,4, f"Default TradingStyle: {vmeta['default_style']}", GRP, TOR, True, 11)
    ov_row(ws_ov,5, '', GRP)
    ov_row(ws_ov,6, f'WHAT\'S NEW IN {ver_key}', GRP, TG, True, 12)
    ov_row(ws_ov,7, '', GRP)
    for i, feat in enumerate(vmeta['new_features'], 8):
        ov_row(ws_ov, i, f'  ✓  {feat}', NEW if i%2==0 else ODD, TGR, False, 10)
    r_after = 8 + len(vmeta['new_features'])
    ov_row(ws_ov, r_after, '', ODD)
    ov_row(ws_ov, r_after+1, 'ACCOUNT NOTES', GRP, TG, True, 12)
    ov_row(ws_ov, r_after+2, '', GRP)
    ov_row(ws_ov, r_after+3, '  $50  → RiskMode=RISK_FIXED_LOT · FixedLot=0.01 · MaxLotLimit=0.01 · MaxDailyLossPercent=6.0 · MaxDailyLossTrades=2 · MaxTradesPerDay=5', C50, TG, True, 10)
    ov_row(ws_ov, r_after+4, '        MaxSLPips=20 · MinH1RangePips=30 · SLBufferPips=10 · TradingStyle=CONSERVATIVE (V1.2+)', C50, TG, False, 10)
    ov_row(ws_ov, r_after+5, '', ODD)
    ov_row(ws_ov, r_after+6, '  $100 → RiskMode=RISK_FIXED_PCT · RiskPercent=1.0 · MaxLotLimit=0.02 · MaxDailyLossPercent=6.0 · MaxDailyLossTrades=3 · MaxTradesPerDay=10', C100, TGR, True, 10)
    ov_row(ws_ov, r_after+7, '         MaxSLPips=25 · MinH1RangePips=40 · SLBufferPips=15 · TradingStyle=BALANCED (V1.2+)', C100, TGR, False, 10)

    # ── Sheet 2: All Settings ─────────────────────────────────────────
    ws_set = wb.create_sheet(f"{ver_key} All Settings")
    ws_set.sheet_properties.tabColor = vmeta['tab']
    col_w = [20, 26, 8, 14, 50, 18, 28, 28]
    col_h = ['Group','Parameter','Type','Default Value','What It Does / How It Works',
             'Code Comment','$50 Recommended','$100 Recommended']
    for i,(h,w) in enumerate(zip(col_h,col_w),1):
        ws_set.column_dimensions[get_column_letter(i)].width = w
        wc(ws_set,1,i,h,HDR,TG,True,11,'center')
    ws_set.row_dimensions[1].height = 22
    ws_set.freeze_panes = 'A2'

    cur_grp = None
    r = 2
    for row in PARAMS:
        grp = row[0]; name = row[1]; typ = row[2]; code_cmt = row[3]
        what = row[4]; default_val = row[vcol]
        rec50 = row[10]; note50 = row[11]; rec100 = row[12]; note100 = row[13]; warn = row[14]

        if default_val is None:
            continue   # parameter not in this version

        if grp != cur_grp:
            for c in range(1,9):
                wc(ws_set, r, c, grp if c==1 else '', GRP, TG, True, 10)
            ws_set.row_dimensions[r].height = 18
            r += 1
            cur_grp = grp

        row_bg = WARN if '⚠' in warn else (ODD if r%2==0 else EVN)
        # highlight newly added params
        if ver_key=='V1.1' and row[5] is None:  row_bg = NEW
        elif ver_key=='V1.2' and row[5] is None and row[6] is None: row_bg = NEW
        elif ver_key=='V1.3' and row[5] is None and row[6] is None and row[7] is None: row_bg = NEW
        elif ver_key=='V1.4' and row[5] is None and row[6] is None and row[7] is None and row[8] is None: row_bg = NEW

        wc(ws_set, r, 1, grp,         row_bg, TSL, False, 9)
        wc(ws_set, r, 2, name,        row_bg, TW,  True,  10)
        wc(ws_set, r, 3, typ,         row_bg, TSL, False, 9, 'center')
        wc(ws_set, r, 4, default_val, row_bg, TGR, True,  10, 'center')
        wc(ws_set, r, 5, what,        row_bg, TSL, False, 9)
        wc(ws_set, r, 6, code_cmt,    row_bg, TB,  False, 9)
        wc(ws_set, r, 7, f'{rec50}\n{note50}',   C50,  TG,  False, 9)
        wc(ws_set, r, 8, f'{rec100}\n{note100}', C100, TGR, False, 9)
        ws_set.row_dimensions[r].height = 55
        r += 1

    # ── Sheet 3: $50 Quick Start ──────────────────────────────────────
    for acct, amount, bg_h, bg_c, fc_c in [
        ('$50', 50, '252500', C50, TG),
        ('$100', 100, '002500', C100, TGR),
    ]:
        ws_q = wb.create_sheet(f"{ver_key} {acct} Quick Start")
        ws_q.sheet_properties.tabColor = vmeta['tab']
        cw = [20, 26, 8, 14, 20, 36, 50]
        ch = ['Group','Parameter','Type','Default','Recommended','Why This Setting','What It Does']
        for i,(h,w) in enumerate(zip(ch,cw),1):
            ws_q.column_dimensions[get_column_letter(i)].width = w
            wc(ws_q,1,i,h,HDR,TG,True,11,'center')
        ws_q.row_dimensions[1].height = 22
        ws_q.freeze_panes = 'A2'

        # intro banner
        ws_q.merge_cells('A2:G2')
        if amount==50:
            intro = f'{vmeta["name"]}  |  $50 Account  |  FixedLot=0.01  |  MaxDailyLoss=6%($3)  |  MaxTrades=5/day'
        else:
            intro = f'{vmeta["name"]}  |  $100 Account  |  RiskPercent=1.0%  |  MaxDailyLoss=6%($6)  |  MaxTrades=10/day'
        wc(ws_q, 2, 1, intro, bg_h, fc_c, True, 11, 'center', False)
        ws_q.row_dimensions[2].height = 20

        cur_grp = None
        r = 3
        for row in PARAMS:
            grp = row[0]; name = row[1]; typ = row[2]
            default_val = row[vcol]
            what = row[4]
            rec50 = row[10]; note50 = row[11]; rec100 = row[12]; note100 = row[13]
            warn  = row[14]

            if default_val is None:
                continue

            rec   = rec50   if amount==50 else rec100
            note  = note50  if amount==50 else note100

            if grp != cur_grp:
                for c in range(1,8):
                    wc(ws_q, r, c, grp if c==1 else '', GRP, TG, True, 10)
                ws_q.row_dimensions[r].height = 16
                r += 1
                cur_grp = grp

            changed = (str(rec) != str(default_val))
            row_bg = WARN if '⚠' in warn else (bg_h if changed else (ODD if r%2==0 else EVN))
            fc_name = fc_c if changed else TW
            wc(ws_q, r, 1, grp,          row_bg, TSL,    False, 9)
            wc(ws_q, r, 2, name,         row_bg, fc_name,True,  10)
            wc(ws_q, r, 3, typ,          row_bg, TSL,    False, 9, 'center')
            wc(ws_q, r, 4, default_val,  row_bg, TSL,    False, 9, 'center')
            wc(ws_q, r, 5, rec,          bg_c,   fc_c,   True,  10, 'center')
            wc(ws_q, r, 6, note,         row_bg, TSL,    False, 9)
            wc(ws_q, r, 7, what,         row_bg, C'B0B0B0' if not '⚠' in warn else TR, False, 9)
            ws_q.row_dimensions[r].height = 45
            r += 1

    wb.save(out_path)
    print(f"Saved: {out_path}")

print("\nAll 5 Excel files complete.")

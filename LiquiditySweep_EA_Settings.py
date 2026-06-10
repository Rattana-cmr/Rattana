import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EA Input Settings"

# ── Colour palette ─────────────────────────────────────────────────────────
C_TITLE_BG   = "1A1A2E"   # dark navy
C_TITLE_FG   = "FFD700"   # gold
C_HEAD_BG    = "16213E"   # deep blue
C_HEAD_FG    = "FFFFFF"
C_GROUP_BG   = "0F3460"   # section header
C_GROUP_FG   = "E0E0E0"
C_FIXED_BG   = "2C2C54"   # same value on all accounts
C_FIXED_FG   = "A0A0FF"
C_NOTE_BG    = "3D0000"   # warning/note rows
C_NOTE_FG    = "FF8080"
C_ALT1       = "1E1E3A"
C_ALT2       = "16203A"
ACCT_COLORS  = ["8B0000","8B4513","2E5E1A","1A4D6E","1A3A6E","2E1A6E"]
ACCT_FG      = ["FFAAAA","FFCC88","AAFFAA","AADDFF","AABBFF","DDAAFF"]

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_fg, bold=False, size=10):
    return Font(color=hex_fg, bold=bold, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="333366")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Column widths ──────────────────────────────────────────────────────────
col_widths = [32, 22, 12, 12, 12, 12, 12, 12]   # param | description | $50...$10k
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 22

# ── Data ───────────────────────────────────────────────────────────────────
accounts = ["$50", "$100", "$300", "$500", "$1 000", "$10 000"]
balances = [50, 100, 300, 500, 1000, 10000]

def risk_usd(bal, pct):
    return f"${bal * pct / 100:.2f}"

def daily_loss(bal, pct, r):
    return f"${abs(bal * pct / 100 * r):.2f}"

# Each row: (group, parameter_name, description, v50, v100, v300, v500, v1k, v10k, row_type)
# row_type: "group" | "fixed" | "vary" | "note" | "calc"
rows = [
    # ── TITLE ROW (special) ──
    ("TITLE", "LiquiditySweep Reversal EA  v6.7  —  Input Settings by Account Size",
     "","","","","","","title"),

    # ── ACCOUNT INFO ──
    ("group","═══  ACCOUNT REFERENCE  ═══","","","","","","","group"),
    ("vary","Account Balance","Starting capital",
     "$50","$100","$300","$500","$1 000","$10 000","vary"),
    ("calc","Risk $ per Trade","At configured risk %",
     risk_usd(50,1.0), risk_usd(100,0.8), risk_usd(300,0.5),
     risk_usd(500,0.3), risk_usd(1000,0.3), risk_usd(10000,0.3),"calc"),
    ("calc","Max Daily Loss $","At configured loss R",
     daily_loss(50,1.0,2), daily_loss(100,0.8,2), daily_loss(300,0.5,2.5),
     daily_loss(500,0.3,3), daily_loss(1000,0.3,3), daily_loss(10000,0.3,3),"calc"),
    ("note","⚠ IMPORTANT","$50–$100 requires a MICRO or CENT account (0.01 lot min).",
     "◄ MICRO","◄ MICRO","","","","","note"),

    # ── MULTI-SYMBOL SCANNER ──
    ("group","═══  MULTI-SYMBOL SCANNER  ═══","","","","","","","group"),
    ("fixed","InpSymbolList","Main 7 forex pairs",
     "EURUSD,GBPUSD,USDJPY,AUDUSD,NZDUSD,USDCAD,USDCHF","","","","","","fixed"),
    ("fixed","InpAdditionalSymbols","Cross pairs (ULTRA+HIGH only)",
     "EURJPY,GBPJPY,AUDJPY,CADJPY","","","","","","fixed"),
    ("fixed","InpIncludeGold","Scan XAUUSD",
     "true","","","","","","fixed"),

    # ── TIMEFRAME ──
    ("group","═══  TIMEFRAME SETTINGS  ═══","","","","","","","group"),
    ("fixed","InpSweepTF","Sweep detection TF (overridden by ULTRA)",
     "M5","","","","","","fixed"),
    ("fixed","InpConfTF","BOS confirmation TF",
     "M1","","","","","","fixed"),
    ("fixed","InpUseM1Sweep","Force M1 sweep (auto in ULTRA)",
     "false","","","","","","fixed"),

    # ── LIQUIDITY POOL ──
    ("group","═══  LIQUIDITY POOL SETTINGS  ═══","","","","","","","group"),
    ("vary","InpLookbackN","Bars to look back for cluster",
     "20","25","30","30","30","40","vary"),
    ("vary","InpDeltaPips","Cluster radius (pips)",
     "2.0","2.0","2.0","2.0","2.0","2.0","vary"),
    ("vary","InpSweepMinPips","Minimum sweep distance (pips)",
     "1.5","2.0","2.5","3.0","3.0","3.0","vary"),

    # ── ENTRY & RISK ──
    ("group","═══  ENTRY & RISK SETTINGS  ═══","","","","","","","group"),
    ("vary","InpRiskPercent","Risk per trade (% of balance)",
     "1.0","0.8","0.5","0.3","0.3","0.3","vary"),
    ("vary","InpSlBufferPips","SL buffer beyond sweep extreme",
     "2.0","2.0","2.0","2.0","2.0","2.0","vary"),
    ("fixed","InpMaxSlippage","Max slippage (points)",
     "5","","","","","","fixed"),
    ("fixed","InpUseFVG","Use FVG entry",
     "false","","","","","","fixed"),
    ("fixed","InpDirectBOSEntry","Direct market entry",
     "true","","","","","","fixed"),
    ("fixed","InpTP_R","Take profit (× risk)",
     "1.5","","","","","","fixed"),
    ("fixed","InpMinRR","Minimum R:R to enter",
     "1.2","","","","","","fixed"),
    ("fixed","InpPartialTP_R","Partial close target (× risk)",
     "1.0","","","","","","fixed"),
    ("vary","InpMaxSlPips","Hard SL cap in pips (0 = off)",
     "10","12","15","15","15","20","vary"),

    # ── TRADE MANAGEMENT ──
    ("group","═══  TRADE MANAGEMENT  ═══","","","","","","","group"),
    ("vary","InpMaxDailyTrades","Max entries per day (global)",
     "10","15","20","25","30","50","vary"),
    ("vary","InpMaxTradesPerSymbol","Per-symbol daily cap (0 = auto)",
     "2","3","4","5","0","0","vary"),
    ("vary","InpMaxDailyNetR","Stop trading after +R today",
     "3.0","4.0","5.0","6.0","8.0","8.0","vary"),
    ("vary","InpMaxDailyLossR","Stop trading after −R today",
     "-2.0","-2.0","-2.5","-3.0","-3.0","-3.0","vary"),
    ("vary","InpMaxOpenPositions","Max simultaneous positions",
     "3","5","7","8","10","12","vary"),
    ("fixed","InpUseCorrelationFilter","USD exposure filter",
     "false","","","","","","fixed"),
    ("fixed","InpUseTrailingStop","Trailing stop",
     "true","","","","","","fixed"),
    ("fixed","InpTrailingStart","Trail activates at (× risk)",
     "0.5","","","","","","fixed"),
    ("fixed","InpTrailingStep","Trail step (× risk)",
     "0.2","","","","","","fixed"),
    ("fixed","InpUseBreakEven","Move SL to break-even",
     "true","","","","","","fixed"),
    ("fixed","InpBreakEvenR","Break-even activates at (× risk)",
     "0.5","","","","","","fixed"),

    # ── ADVANCED ──
    ("group","═══  ADVANCED FEATURES  ═══","","","","","","","group"),
    ("fixed","InpAggressiveMode","Aggression level",
     "AGGRESSIVE_ULTRA","","","","","","fixed"),
    ("fixed","InpBOSMode","BOS confirmation mode",
     "BOS_BYPASS","","","","","","fixed"),
    ("fixed","InpAllowReEntry","Re-enter same sweep",
     "true","","","","","","fixed"),
    ("vary","InpMaxEntriesPerSweep","Max entries per sweep signal",
     "2","2","3","3","3","3","vary"),
    ("fixed","InpTrendFilter","SMA trend filter",
     "FILTER_NONE","","","","","","fixed"),
    ("fixed","InpMinATR","Minimum ATR (pips)",
     "1.0","","","","","","fixed"),

    # ── SESSION ──
    ("group","═══  SESSION FILTER  ═══","","","","","","","group"),
    ("fixed","InpUseSessionFilter","Restrict to session window",
     "true","","","","","","fixed"),
    ("fixed","InpSessionStartHour (GMT)","Session open hour",
     "7","","","","","","fixed"),
    ("fixed","InpSessionEndHour (GMT)","Session close hour",
     "23","","","","","","fixed"),

    # ── SPREAD LIMITS ──
    ("group","═══  SPREAD LIMITS (pips)  ═══","","","","","","","group"),
    ("fixed","InpSpreadLimitEURUSD","EURUSD max spread",
     "1.5","","","","","","fixed"),
    ("fixed","InpSpreadLimitGBPUSD","GBPUSD max spread",
     "2.0","","","","","","fixed"),
    ("fixed","InpSpreadLimitUSDJPY","USDJPY max spread",
     "1.5","","","","","","fixed"),
    ("fixed","InpSpreadLimitXAUUSD","XAUUSD max spread",
     "3.5","","","","","","fixed"),
    ("fixed","InpSpreadLimitDefault","All other pairs",
     "2.0","","","","","","fixed"),

    # ── EXPERT ──
    ("group","═══  EXPERT SETTINGS  ═══","","","","","","","group"),
    ("fixed","InpMagicNumber","Unique EA identifier",
     "20250609","","","","","","fixed"),
    ("fixed","InpComment","Trade comment tag",
     "LSweep_HF","","","","","","fixed"),
    ("vary","InpDebugMode","Journal logging (turn OFF after testing)",
     "false","false","false","false","false","false","vary"),
]

# ── Write header ───────────────────────────────────────────────────────────
# Row 1: main title
ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = "⚡  LIQUIDITY SWEEP REVERSAL EA  v6.7  —  Input Settings Guide"
c.fill = fill(C_TITLE_BG)
c.font = Font(color=C_TITLE_FG, bold=True, size=14, name="Calibri")
c.alignment = center()
c.border = border

# Row 2: column headers
headers = ["Parameter", "Description", "$50", "$100", "$300", "$500", "$1 000", "$10 000"]
for col, hdr in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=hdr)
    c.fill = fill(C_HEAD_BG)
    c.font = font(C_HEAD_FG, bold=True, size=10)
    c.alignment = center()
    c.border = border
    if col >= 3:
        c.fill = PatternFill("solid", fgColor=ACCT_COLORS[col-3])
        c.font = Font(color=ACCT_FG[col-3], bold=True, size=10, name="Calibri")

# ── Write data rows ────────────────────────────────────────────────────────
current_row = 3
alt = 0

for row_data in rows:
    rtype = row_data[-1]

    if rtype == "title":
        continue   # already written

    if rtype == "group":
        ws.merge_cells(f"A{current_row}:H{current_row}")
        c = ws.cell(row=current_row, column=1, value=row_data[1])
        c.fill = fill(C_GROUP_BG)
        c.font = font(C_GROUP_FG, bold=True, size=10)
        c.alignment = left()
        c.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        alt = 0
        continue

    # data row
    vals = list(row_data[2:8])   # 6 account columns
    param  = row_data[1]
    desc   = row_data[2]
    vals   = list(row_data[3:9])

    # background
    if rtype == "note":
        row_bg = C_NOTE_BG
        row_fg = C_NOTE_FG
    elif rtype == "calc":
        row_bg = "1E3A1E"
        row_fg = "80FF80"
    elif rtype == "fixed":
        row_bg = C_FIXED_BG
        row_fg = C_FIXED_FG
    else:
        row_bg = C_ALT1 if alt % 2 == 0 else C_ALT2
        row_fg = "E8E8FF"
    alt += 1

    ws.row_dimensions[current_row].height = 18

    # col A – parameter name
    c = ws.cell(row=current_row, column=1, value=param)
    c.fill = fill(row_bg)
    c.font = font(row_fg, bold=(rtype == "vary"), size=9)
    c.alignment = left()
    c.border = border

    # col B – description
    c = ws.cell(row=current_row, column=2, value=desc)
    c.fill = fill(row_bg)
    c.font = font("A0A0C0", size=9)
    c.alignment = left()
    c.border = border

    # cols C-H – account values
    is_fixed = (rtype == "fixed")
    for col_i, val in enumerate(vals, 3):
        c = ws.cell(row=current_row, column=col_i)
        # For fixed rows, only first value column gets the value; rest get "  ← same"
        if is_fixed:
            if col_i == 3:
                c.value = val
            else:
                c.value = "← same"
                c.font = Font(color="505080", italic=True, size=8, name="Calibri")
                c.fill = fill("1A1A30")
                c.alignment = center()
                c.border = border
                continue
        else:
            c.value = val

        c.fill = PatternFill("solid", fgColor=ACCT_COLORS[col_i-3])
        c.font = Font(color=ACCT_FG[col_i-3],
                      bold=(rtype == "vary"),
                      size=9, name="Calibri")
        c.alignment = center()
        c.border = border

    current_row += 1

# ── Legend sheet ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legend & Notes")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 70

legend_rows = [
    ("LEGEND","","title"),
    ("Row colour","Meaning","head"),
    ("BRIGHT + BOLD","Variable by account size — change these","vary"),
    ("BLUE / DIM","Same value for all account sizes","fixed"),
    ("GREEN (italic)","Calculated reference (not an input)","calc"),
    ("RED / WARNING","Important notice — read before trading","note"),
    ("","",""),
    ("NOTES","","title"),
    ("note","","head"),
    ("$50 – $100",
     "Requires a MICRO or CENT account. Standard accounts have 0.01 lot minimum\n"
     "which already exceeds the calculated risk size for small balances.","note"),
    ("InpRiskPercent",
     "Higher % on small accounts to get a workable lot size. Reduce to 0.5% once\n"
     "balance grows above $200.","note"),
    ("InpDebugMode",
     "Set to false after your first test session. With 11 symbols on M1, debug=true\n"
     "generates thousands of journal lines per hour.","note"),
    ("InpMaxDailyTrades",
     "Global cap across all symbols. Raise to 50 if you want more frequency.","note"),
    ("InpMaxTradesPerSymbol",
     "0 = auto-distribute from global cap. Set to a number to give each symbol\n"
     "an explicit daily ceiling.","note"),
    ("InpBOSMode",
     "BOS_BYPASS = most trades, least filtering.\n"
     "BOS_RELAXED = medium quality.\n"
     "BOS_NORMAL = strictest, fewest trades.","note"),
    ("Session filter",
     "Hours are GMT (not broker time). Verify your broker's GMT offset before\n"
     "adjusting InpSessionStartHour / InpSessionEndHour.","note"),
    ("InpMaxSlPips",
     "Hard-caps the stop loss distance. Tighter = more trades stopped out early\n"
     "on volatile pairs. Set 0 to disable the cap entirely.","note"),
]

row = 1
for item in legend_rows:
    if len(item) < 3:
        row += 1
        continue
    lbl, txt, ltype = item
    if ltype == "title":
        ws2.merge_cells(f"A{row}:B{row}")
        c = ws2.cell(row=row, column=1, value=lbl)
        c.fill = fill(C_TITLE_BG)
        c.font = Font(color=C_TITLE_FG, bold=True, size=12, name="Calibri")
        c.alignment = center()
    elif ltype == "head":
        c = ws2.cell(row=row, column=1, value=lbl)
        c.fill = fill(C_HEAD_BG)
        c.font = font(C_HEAD_FG, bold=True)
        c = ws2.cell(row=row, column=2, value=txt)
        c.fill = fill(C_HEAD_BG)
        c.font = font(C_HEAD_FG, bold=True)
    else:
        c = ws2.cell(row=row, column=1, value=lbl)
        c.fill = fill(C_GROUP_BG)
        c.font = font(C_GROUP_FG, bold=True, size=9)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c = ws2.cell(row=row, column=2, value=txt)
        c.fill = fill(C_ALT1)
        c.font = font("D0D0FF", size=9)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.row_dimensions[row].height = max(18, txt.count("\n") * 14 + 18)
    row += 1

path = "/home/user/Rattana/LiquiditySweep_EA_Settings_v6.7.xlsx"
wb.save(path)
print("Saved:", path)

#!/usr/bin/env python3
"""ICT SMC EA — Complete Input Settings Guide  V1.0–V1.5"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/Rattana/ICT_SMC_EA_Settings_Guide.xlsx"

# ── colours ──────────────────────────────────────────────────────────
def f(c):   return PatternFill(fill_type='solid', fgColor=c)
def fn(c='FFFFFF', b=False, sz=10): return Font(color=c, bold=b, size=sz, name='Calibri')
def al(h='left', v='center', w=True): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def bd():
    s = Side(style='thin', color='3A3A5A')
    return Border(left=s, right=s, top=s, bottom=s)

HDR  = '0D2137'; GRP = '12243A'; ODD = '161625'; EVN = '1C1C2E'
NEW  = '0A2A0A'; WARN= '2A0A0A'; NEW15 = '1A0A2A'
TW='FFFFFF'; TG='FFD700'; TGR='00DD77'; TR='FF6060'; TB='64AAFF'; TSL='B0B0C0'; TOR='FFA040'; TPU='CC88FF'
C50H='252500'; C50='1C1C00'; C100H='002500'; C100='001C00'
VERS=['V1.0','V1.1','V1.2','V1.3','V1.4','V1.5']

# ── helper: write a cell ─────────────────────────────────────────────
def wc(ws, r, c, val, bg=ODD, fc=TW, bold=False, sz=10, h='left', wrap=True, merge=0):
    cell = ws.cell(row=r, column=c, value=str(val) if val is not None else '')
    cell.fill = f(bg); cell.font = fn(fc, bold, sz)
    cell.alignment = al(h, 'center', wrap); cell.border = bd()
    if merge > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+merge-1)

# ═══════════════════════════════════════════════════════════════════════
# PARAMETER DATA
# ═══════════════════════════════════════════════════════════════════════
# grp, name, type, v10,v11,v12,v13,v14,v15, what, f50,fn50, f100,fn100, warning
P = [
  # ── TRADING STYLE ──────────────────────────────────────────────────
  ('TRADING STYLE','TradingStyle','Enum','—','—','BALANCED','SMART_ACTIVE','SMART_ACTIVE','SMART_ACTIVE',
   'Master preset that overrides many individual settings.\n'
   '• Conservative: score≥80, all ICT strict, CD 30m  (1-5 trades/day)\n'
   '• Balanced: respects input toggles, score≥70  (3-8 trades/day)\n'
   '• Aggressive: MSS/Trend OFF, OTE 55-90%, score≥50  (8-15 trades/day)\n'
   '• Smart Active (V1.3+): all ICT ON, better detection, score≥55  (8-15 trades/day)\n'
   '• Smart Active+ (V1.4): all ICT ON, OTE 55-90%, score≥45  (15-25 trades/day)\n'
   '• Ultra Active: filters minimal, max frequency (V1.5: OTE 55-90%, score bypass)',
   'STYLE_CONSERVATIVE','Safest for $50 — strictest entries, fewer but higher-quality trades. Protects small capital from overtrading.',
   'STYLE_BALANCED','Good balance of trade frequency and quality for $100.',
   'Overrides many individual inputs. Settings below marked * may not take effect.'),

  # ── RISK MANAGEMENT ────────────────────────────────────────────────
  ('RISK MANAGEMENT','RiskMode','Enum','—','FIXED_PCT','FIXED_PCT','FIXED_PCT','FIXED_PCT','FIXED_PCT',
   'How position size is calculated:\n'
   '• RISK_FIXED_PCT: lot = (balance × risk%) ÷ (SL pips × pip value)\n'
   '• RISK_FIXED_LOT: always use FixedLot value (ignores RiskPercent)\n'
   '• RISK_DYNAMIC_EQ: same as FIXED_PCT but uses equity not balance',
   'RISK_FIXED_LOT','Use fixed 0.01 lot on $50. Percentage mode can give lot sizes below broker minimum.',
   'RISK_FIXED_PCT','Percentage mode scales correctly with $100 account.',
   'Not in V1.0 — V1.0 always uses RISK_FIXED_PCT equivalent'),

  ('RISK MANAGEMENT','RiskPercent','%','0.5','0.5','0.5','0.5','0.5','0.5',
   'Percentage of account balance to risk per single trade.\n'
   'Only used when RiskMode = RISK_FIXED_PCT or RISK_DYNAMIC_EQ.\n'
   'Example: $100 × 1.0% = $1.00 max risk per trade.',
   '1.0','Not used when RiskMode=FIXED_LOT. Keep ≥1.0 as fallback anyway.',
   '1.0','$100 × 1.0% = $1 risk/trade. Appropriate for early-stage live trading.',
   ''),

  ('RISK MANAGEMENT','FixedLot','lot','0.0','0.0','0.0','0.0','0.0','0.0',
   'Fixed lot size used when RiskMode = RISK_FIXED_LOT.\n'
   '0.0 = use RiskPercent calculation instead.\n'
   'Minimum lot on most brokers for XAUUSD: 0.01',
   '0.01','Set 0.01 and combine with RiskMode = RISK_FIXED_LOT for predictable micro-account risk.',
   '0.01','Use 0.01 as safety floor with FIXED_LOT, or 0.0 to let RiskPercent calculate.',
   ''),

  ('RISK MANAGEMENT','MaxDailyLossPercent','%','10.0','10.0','10.0','10.0','10.0','10.0',
   'EA stops trading for the day when total daily P&L loss reaches this % of balance.\n'
   'Example: $50 × 6% = $3.00 maximum loss per day.\n'
   'Resets at midnight broker time.',
   '6.0','$50 × 6% = $3 daily stop. Default 10% = $5 on $50 which is too aggressive for a micro start.',
   '6.0','$100 × 6% = $6 daily stop. Keeps drawdown manageable while allowing normal trade volume.',
   ''),

  ('RISK MANAGEMENT','MaxTradesPerDay','int','10','10','25','25','25','25',
   'Maximum new trades allowed per day. Counter resets at midnight.\n'
   'Should match your TradingStyle target frequency:\n'
   '  Conservative 1-5 | Balanced 3-8 | SmartActive 8-15 | SmartActive+ 15-25',
   '5','Limit to 5/day on $50 — fewer chances to overexpose small capital.',
   '10','10/day works for V1.0-V1.2. Raise to 15-20 for V1.3/V1.4/V1.5 Smart Active styles.',
   ''),

  ('RISK MANAGEMENT','RewardRiskRatio','ratio','2.0','2.0','2.0','2.0','2.0','2.0',
   'Target profit relative to stop loss distance.\n'
   '2.0 = TP is 2× the SL distance (2:1 R:R).\n'
   'At 2:1 R:R you break even with only 34% win rate.',
   '2.0','Never go below 2.0 on $50 — you need the mathematical edge.',
   '2.0','Keep 2.0. Do not reduce below 1.5 or the edge disappears.',
   ''),

  ('RISK MANAGEMENT','MaxLotLimit','lot','0.10','0.10','0.10','0.10','0.10','0.10',
   'Hard ceiling on position size regardless of risk calculation.\n'
   'Prevents oversized positions from glitches or very tight SL signals.\n'
   'Scale this with your account size.',
   '0.01','Lock at 0.01 — never allow more than minimum lot on $50.',
   '0.02','0.02 allows slightly larger positions as $100 account grows.',
   ''),

  ('RISK MANAGEMENT','MinRewardRiskRatio','ratio','2.0','2.0','2.0','2.0','2.0','2.0',
   'Trade is skipped if the actual calculated R:R is below this value.\n'
   'Rejects setups where SL is too wide vs the target TP.',
   '2.0','Always require 2:1 minimum R:R.',
   '1.8','Can relax slightly to 1.8 to allow more setups.',
   ''),

  # ── TAKE PROFIT MODE ───────────────────────────────────────────────
  ('TAKE PROFIT MODE','TPMode','Enum','—','FIXED_RR','FIXED_RR','FIXED_RR','FIXED_RR','FIXED_RR',
   'How take profit level is calculated:\n'
   '• TP_FIXED_RR: TP = entry ± (SL distance × RewardRiskRatio)\n'
   '• TP_ATR: TP = entry ± (ATR(14) × ATRMultiplierTP)\n'
   '• TP_HYBRID: TP = min(FixedRR, ATR) — most conservative of the two',
   'TP_FIXED_RR','Fixed R:R is predictable and simple. Best for small accounts.',
   'TP_FIXED_RR','Fixed R:R keeps trade math consistent.',
   'Not in V1.0'),

  ('TAKE PROFIT MODE','ATRMultiplierTP','x','—','3.0','3.0','3.0','3.0','3.0',
   'ATR multiplier used only when TPMode = TP_ATR or TP_HYBRID.\n'
   'TP = current ATR(14) × this value.\n'
   'ATR adapts to current market volatility automatically.',
   '3.0','Keep default — only relevant if you switch TPMode to TP_ATR.',
   '3.0','Keep default.',
   'Only used when TPMode ≠ TP_FIXED_RR'),

  # ── PARTIAL TAKE PROFIT ────────────────────────────────────────────
  ('PARTIAL TAKE PROFIT','UsePartialTP','bool','—','true','true','true','true','true',
   'When true: closes PartialClosePercent% of position at PartialCloseRR × R:R,\n'
   'then moves SL to breakeven for the remaining portion.\n'
   'V1.5: default changed to 35% at 1.5R — better expected value than 50% at 1R.',
   'true','Highly recommended for $50 — locks in partial profit, remainder runs free.',
   'true','Recommended for all account sizes.',
   'Not in V1.0'),

  ('PARTIAL TAKE PROFIT','PartialClosePercent','%','—','50.0','50.0','50.0','50.0','35.0',
   'Percentage of position to close at the partial TP level.\n'
   'V1.5 default changed from 50% to 35%.\n'
   'MATH: At 50%@1R you need >67% win rate to profit (if remainder hits SL after partial).\n'
   'At 35%@1.5R the math is more favorable — less exposed if remainder stops out.',
   '35.0','[V1.5] Use 35% — better EV than 50% at 1R for most win rates.',
   '35.0','[V1.5] Use 35% partial close.',
   '[V1.5 CHANGE] Default was 50.0 in V1.0-V1.4'),

  ('PARTIAL TAKE PROFIT','PartialCloseRR','R','—','1.0','1.0','1.0','1.0','1.5',
   'R:R level at which partial close triggers.\n'
   'V1.5 default changed from 1.0R to 1.5R.\n'
   '1.5 = trigger when profit = 1.5× the SL distance.\n'
   'Combining 35% partial at 1.5R gives a better payoff structure than 50% at 1R.',
   '1.5','[V1.5] Use 1.5R trigger — better than 1.0R for partial TP math.',
   '1.5','[V1.5] Use 1.5R trigger.',
   '[V1.5 CHANGE] Default was 1.0 in V1.0-V1.4'),

  # ── TRADE FILTERS ──────────────────────────────────────────────────
  ('TRADE FILTERS','UseTimeFilter','bool','true','true','true','true','true','true',
   'When true: EA only trades during active trading sessions (set below).\n'
   'When false: EA trades 24/5 regardless of time.',
   'true','Always keep ON — avoid low-liquidity hours.',
   'true','Keep ON.',
   ''),

  ('TRADE FILTERS','MaxSpreadPoints','pts','50','50','80','80','80','80',
   'Maximum spread (bid-ask gap) in broker points allowed before a trade.\n'
   'Protects against wide spreads during news or low liquidity.\n'
   'XAUUSD typical spread: 20-40 pts normal, 80-200 pts during news.\n'
   '0 = disable check.',
   '80','80 pts covers normal gold spreads. Wider = possible during news events.',
   '80','80 is appropriate for XAUUSD. Reduce to 50 if broker has tight spreads.',
   ''),

  ('TRADE FILTERS','MinStopDistance','pts','20','20','20','20','20','20',
   'Minimum SL distance from entry price in broker points.\n'
   'Brokers enforce a "freeze level" — orders too close to price are rejected.\n'
   'Must exceed your broker\'s stop level to avoid order rejection.',
   '20','Keep default 20 points.',
   '20','Keep default.',
   ''),

  ('TRADE FILTERS','MaxConsecutiveLosses','int','10','10','10','10','10','10',
   'EA halts trading after N consecutive losses.\n'
   'Lifetime counter unless ResetLossStreakDaily = true.\n'
   'Circuit breaker against runaway losing streaks.',
   '5','Lower to 5 on $50 — stops the EA before 5 losses damage the account too much.',
   '7','7 consecutive losses is a solid circuit breaker for $100.',
   ''),

  ('TRADE FILTERS','MaxDailyLossTrades','int','3','3','3','3','3','3',
   'Stop trading for the day after N losing trades (not the same as MaxDailyLossPercent).\n'
   'Resets at midnight.',
   '2','Stop after 2 losing trades on $50 — preserves capital for the next day.',
   '3','Standard 3 losing-trade daily stop.',
   ''),

  ('TRADE FILTERS','ResetLossStreakDaily','bool','true','true','true','true','true','true',
   'true = MaxConsecutiveLosses counter resets every midnight.\n'
   'false = counter accumulates for lifetime of the EA session.',
   'true','Reset daily — fresh start each morning.',
   'true','Keep true.',
   ''),

  # ── ICT STRUCTURE FILTERS ──────────────────────────────────────────
  ('ICT STRUCTURE FILTERS','UseMSSFilter','bool','—','true','true','true','true','true',
   'Require a Market Structure Shift on H1 before entry.\n'
   'MSS = H1 price breaks above a previous swing high (bullish MSS)\n'
   'or below a swing low (bearish MSS) — confirms direction change.\n'
   '* Overridden OFF by STYLE_AGGRESSIVE and STYLE_ULTRA_ACTIVE',
   'true','Keep ON — MSS confirms the move is real, not noise.',
   'true','Always ON for quality signals.',
   'Not in V1.0. Overridden by TradingStyle in V1.2+'),

  ('ICT STRUCTURE FILTERS','UseBOSFilter','bool','—','true','true','true','true','true',
   'Require Break of Structure on M15 for lower-timeframe confirmation.\n'
   'BOS = M15 price breaks above recent swing high (after bullish MSS).\n'
   'Adds a second confirmation layer on a lower timeframe.',
   'true','Keep ON.',
   'true','Keep ON.',
   'Not in V1.0'),

  ('ICT STRUCTURE FILTERS','RequireLiquiditySweep','bool','—','false','false','false','false','false',
   'Require a liquidity sweep (stop hunt) before entry.\n'
   'V1.5 FIX: sweep now detected using swing-based pools (was just range min/max).\n'
   'Smart money sweeps retail stops above swing highs / below swing lows before reversing.\n'
   'false = optional; true = must detect sweep (rare, reduces frequency to ~1-2/week)',
   'false','Keep false — liquidity sweeps are rare. Enabling reduces trades dramatically.',
   'false','Keep false unless you specifically want high-rarity setups.',
   '[V1.5] Detection algorithm fixed — swing-based pools instead of range extremes'),

  ('ICT STRUCTURE FILTERS','UseSMTFilter','bool','—','false','false','false','false','false',
   'SMT (Smart Money Divergence) filter.\n'
   'Compares price structure between this symbol and SMTSymbol.\n'
   'BUY if this symbol makes lower low but correlated symbol does not.',
   'false','Keep OFF for simplicity.',
   'false','Keep OFF unless you specifically trade SMT setups.',
   'Not in V1.0'),

  ('ICT STRUCTURE FILTERS','SMTSymbol','string','—','XAGUSD','XAGUSD','XAGUSD','XAGUSD','XAGUSD',
   'Correlated symbol for SMT divergence. Only active when UseSMTFilter = true.\n'
   'Gold → XAGUSD (silver). Indices → use correlated index.',
   'XAGUSD','Keep default.',
   'XAGUSD','Keep default.',
   'Symbol must exist on your broker'),

  ('ICT STRUCTURE FILTERS','UseDealingRange','bool','—','—','—','—','—','false',
   '[V1.5 NEW] ICT Dealing Range filter:\n'
   '• Buys only in Discount zone (price ≤ midpoint of H4 range)\n'
   '• Sells only in Premium zone (price ≥ midpoint of H4 range)\n'
   'Uses last 6 H4 bars (~24 hours) to define the institutional dealing range.\n'
   'Default false — enable to add institutional premium/discount filtering.\n'
   'When enabled, entries at wrong side of equilibrium are blocked.',
   'false','Keep OFF initially. Enable after verifying baseline performance.',
   'false','Optional quality filter. Test with false first.',
   '[V1.5 NEW] Not available in V1.0-V1.4'),

  # ── NEWS FILTER ────────────────────────────────────────────────────
  ('NEWS FILTER','UseNewsFilter','bool','—','false','false','false','false','false',
   'Block trading before/after high-impact economic news events.\n'
   'Uses MT5 Calendar API. Requires "Allow WebRequest" in MT5 settings.',
   'false','Disable in Strategy Tester (calendar limited). Enable for live trading.',
   'false','Optional but recommended for live, especially around NFP/FOMC.',
   'Not in V1.0. Requires Allow WebRequest enabled in MT5'),

  ('NEWS FILTER','NewsBlockBeforeMin','min','—','30','30','30','30','30',
   'Minutes BEFORE a high-impact news event to stop accepting new trades.',
   '30','Standard 30-minute pre-news block.',
   '30','Keep 30 minutes.',
   ''),

  ('NEWS FILTER','NewsBlockAfterMin','min','—','30','30','30','30','30',
   'Minutes AFTER a high-impact news event before resuming trades.\n'
   'Spreads and volatility remain elevated for 15-30 min post-news.',
   '30','30-minute post-news cooldown.',
   '30','Keep 30 minutes.',
   ''),

  # ── TRADE QUALITY SCORE ────────────────────────────────────────────
  ('TRADE QUALITY SCORE','UseTradeScore','bool','—','—','true','true','true','true',
   'Enable the minimum score gate.\n'
   'When false: score is shown on panel but never blocks trades.\n'
   'When true: trade skipped if score < MinimumTradeScore.\n'
   'V1.5 FIX: disabled filters no longer give free points — score is now meaningful.',
   'true','Keep ON — score filter is your quality gate on small accounts.',
   'true','Keep ON.',
   'Only in V1.2+. [V1.5] Score formula fixed — no free points for disabled filters'),

  ('TRADE QUALITY SCORE','MinimumTradeScore','0-100','—','70','70','70','70','70',
   'Minimum score (0-100) required to place a trade.\n'
   'Score = sum of: MSS confirmed (20pts) + BOS confirmed (20pts) +\n'
   '  Sweep confirmed (20pts) + FVG met (15pts) + base entry (15pts) + trend (10pts).\n'
   'V1.5 FIX: points only awarded when filter is ON and confirmed (not when disabled).\n'
   'Overridden by TradingStyle preset in V1.2+.',
   '70','70 blocks low-quality setups without being too restrictive.',
   '65','Slightly relax to 65 for more trades on $100 while keeping quality.',
   '[V1.5 FIX] Disabled filters no longer grant free score points'),

  # ── SESSIONS ───────────────────────────────────────────────────────
  ('SESSIONS (GMT)','AutoDetectGMT','bool','—','true','true','true','true','true',
   'Auto-detect broker GMT offset using: offset = (TimeCurrent - TimeGMT) / 3600.\n'
   'Works correctly on live accounts. In Strategy Tester always shows GMT+0 (normal).',
   'true','Keep ON for live trading.',
   'true','Keep ON.',
   'Always shows GMT+0 in Strategy Tester — this is expected, not a bug'),

  ('SESSIONS (GMT)','BrokerGMTOffset','hrs','0','0','0','0','0','0',
   'Manual GMT offset when AutoDetectGMT = false.\n'
   'EU brokers: GMT+2 (winter) / GMT+3 (summer DST).',
   '0','Leave 0 when AutoDetectGMT = true.',
   '0','Leave 0 when AutoDetectGMT = true.',
   ''),

  ('SESSIONS (GMT)','SessionSydney','bool','false','false','false','false','false','false',
   'Allow trading during Sydney session: 22:00-07:00 GMT.\n'
   'Low volume for XAUUSD. Not recommended for gold.',
   'false','Keep OFF — very low liquidity for gold.',
   'false','Keep OFF.',
   ''),

  ('SESSIONS (GMT)','SessionTokyo','bool','false','false','false','false','false','false',
   'Allow trading during Tokyo session: 00:00-09:00 GMT.\n'
   'Moderate XAUUSD activity. Enable for V1.4/V1.5 SmartActive+ expanded coverage.',
   'false','Keep OFF on $50.',
   'false','Can enable for SmartActive+ to expand session hours.',
   ''),

  ('SESSIONS (GMT)','SessionLondon','bool','true','true','true','true','true','true',
   'Allow trading during London session: 08:00-17:00 GMT.\n'
   'Highest liquidity period. Best session for XAUUSD and FX.',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','SessionNewYork','bool','true','true','true','true','true','true',
   'Allow trading during New York session: 13:00-22:00 GMT.\n'
   'High volatility with strong ICT move potential.',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','OverlapLondonNY','bool','true','true','true','true','true','true',
   'Allow trading during London/NY overlap: 13:00-17:00 GMT.\n'
   'Best 4 hours of the day — highest liquidity AND volatility combined.',
   'true','Definitely ON — highest quality ICT entries happen in this window.',
   'true','Always ON.',
   ''),

  ('SESSIONS (GMT)','OverlapTokyoLondon','bool','false','false','false','false','false','false',
   'Allow trading during Tokyo/London overlap: 08:00-09:00 GMT.\n'
   'Only 1 hour. Already covered by SessionLondon.',
   'false','Keep OFF.',
   'false','Keep OFF.',
   ''),

  # ── STOP LOSS ──────────────────────────────────────────────────────
  ('STOP LOSS','SLBufferPips','pips','15','15','15','15','15','15',
   'Extra pip buffer beyond the swing high/low for SL placement.\n'
   'Example: swing low at 2000.00, buffer=15 → SL placed ~15 pips below.\n'
   'Prevents SL from being hit by normal wick noise at the swing level.',
   '10','Reduce to 10 pips on $50 — tightens SL and reduces $ loss per trade.',
   '15','Default 15 pips provides good buffer against stop hunts.',
   ''),

  ('STOP LOSS','UseTrailingStop','bool','false','false','false','false','false','false',
   'Enable trailing stop loss.\n'
   'When true: SL moves up (for BUY) after TrailingStartPips profit.\n'
   'UsePartialTP is a more consistent alternative.',
   'false','Keep OFF — UsePartialTP is more effective for ICT style.',
   'false','Keep OFF unless backtest proves improvement.',
   ''),

  ('STOP LOSS','TrailingStartPips','pips','30','30','30','30','30','30',
   'Profit in pips before trailing stop activates.\n'
   'Only used when UseTrailingStop = true.',
   '30','Only relevant if UseTrailingStop = true.',
   '30','Only relevant if UseTrailingStop = true.',
   'Only used when UseTrailingStop = true'),

  ('STOP LOSS','TrailingStepPips','pips','10','10','10','10','10','10',
   'How many pips to trail behind current price.\n'
   'Smaller = tighter trail; Larger = more room for fluctuation.',
   '10','Only relevant if UseTrailingStop = true.',
   '10','Only relevant if UseTrailingStop = true.',
   'Only used when UseTrailingStop = true'),

  # ── POSITION MANAGEMENT ────────────────────────────────────────────
  ('POSITION MANAGEMENT','CloseOnFriday','bool','true','true','true','true','true','true',
   'Automatically close all open positions on Friday at FridayCloseHour GMT.\n'
   'Prevents weekend gap risk — price can jump 50-200 pips over weekends.',
   'true','Always ON — weekend gaps can easily wipe $50.',
   'true','Always ON.',
   ''),

  ('POSITION MANAGEMENT','FridayCloseHour','GMT hr','14','14','14','14','14','14',
   'GMT hour on Friday when positions are closed.\n'
   '14:00 GMT = before thin NY afternoon and approaching weekend gap risk.',
   '14','14:00 GMT is optimal.',
   '14','Keep 14:00 GMT.',
   ''),

  ('POSITION MANAGEMENT','UseBreakeven','bool','false','false','false','false','false','false',
   'Move SL to entry price after BreakevenTriggerPips profit.\n'
   'UsePartialTP is the superior alternative — use one or the other.',
   'false','Keep OFF — UsePartialTP at 1.5R is better.',
   'false','Keep OFF.',
   'Conflicts with UsePartialTP'),

  ('POSITION MANAGEMENT','BreakevenTriggerPips','pips','40','40','40','40','40','40',
   'Pips of profit required before SL moves to breakeven.\n'
   'Only used when UseBreakeven = true.',
   '30','Only relevant if UseBreakeven = true.',
   '40','Only relevant if UseBreakeven = true.',
   'Only used when UseBreakeven = true'),

  # ── SWING DETECTION ────────────────────────────────────────────────
  ('SWING DETECTION','SwingLookbackBarsH1','bars','50','50','50','50','50','50',
   'H1 bars to scan backward when finding swing highs/lows for OTE zone.\n'
   '50 bars ≈ 2 days of H1 history.',
   '50','Keep default.',
   '50','Keep default.',
   ''),

  ('SWING DETECTION','SwingConfirmBarsH1','bars','3','3','3','3','3','3',
   'Bars on EACH SIDE required to confirm an H1 swing point.\n'
   '3 = swing must be highest/lowest of 7 bars total.\n'
   'Overridden by effMSSConfirm in V1.3+ Smart Active styles.',
   '3','Default is fine — overridden by TradingStyle in V1.3+.',
   '3','Keep default.',
   'In V1.3+: overridden to 2 (SmartActive) or 1 (SmartActive+) by style'),

  ('SWING DETECTION','SwingLookbackBarsM15','bars','30','30','30','30','30','30',
   'M15 bars to scan for the SL anchor swing.\n'
   '30 bars ≈ 7.5 hours of M15.',
   '30','Keep default.',
   '30','Keep default.',
   ''),

  ('SWING DETECTION','SwingConfirmBarsM15','bars','5','5','5','5','5','5',
   'Bars each side to confirm an M15 swing (used for SL placement).\n'
   '5 = swing must be highest/lowest of 11 bars.',
   '5','Keep default.',
   '5','Keep default.',
   ''),

  ('SWING DETECTION','MaxSwingDistancePips','pips','500','500','500','500','500','500',
   'Maximum allowed H1 swing range in pips.\n'
   'Rejects abnormally large swings that suggest unusual market conditions.\n'
   '0 = no maximum.',
   '300','Tighter at 300 pips on $50 — extreme ranges often precede reversals.',
   '500','Keep default.',
   ''),

  ('SWING DETECTION','MaxSLPips','pips','30','30','30','30','30','30',
   'Maximum SL distance in pips. Trade skipped if SL > this value.\n'
   'Prevents large SL from consuming disproportionate % of small accounts.',
   '20','20 pips max SL on $50 keeps maximum loss tightly controlled.',
   '25','25-30 pips appropriate for $100.',
   ''),

  ('SWING DETECTION','MinSLPips','pips','10','10','10','10','10','10',
   'Minimum SL distance in pips.\n'
   'Skips trades with SL so tight it lives inside the spread.',
   '10','Keep default.',
   '10','Keep default.',
   ''),

  ('SWING DETECTION','ShowSwingLines','bool','true','true','true','true','true','true',
   'Draw M15 swing high/low lines on the chart.\n'
   'Shows where the EA anchored the SL.',
   'true','Keep ON for visual monitoring.',
   'true','Keep ON.',
   ''),

  # ── ICT TWINS MODEL ────────────────────────────────────────────────
  ('ICT TWINS MODEL','UseTwinsModel','bool','true','true','true','true','true','true',
   'Enable the full ICT Twins sequential entry model.\n'
   'Steps: HTF Level → MSS → BOS → Liquidity → OTE Zone → CISD → 1M Trigger.\n'
   'This IS the core strategy. Must always be ON.',
   'true','Always ON.',
   'true','Always ON.',
   ''),

  ('ICT TWINS MODEL','HTFLevelMinutes','min','15','15','15','15','15','15',
   'Timeframe for HTF reference level detection.\n'
   '15 = M15. EA checks proximity to M15 FVGs, prev-day H/L, H1/H4 levels.',
   '15','Keep default M15.',
   '15','Keep default.',
   ''),

  ('ICT TWINS MODEL','OTEMinPercent','0-1','0.65','0.65','0.65','0.65','0.65','0.65',
   'OTE zone minimum as % of the H1 swing range (measured from the base).\n'
   '0.65 = 65% retracement into the swing.\n'
   'Overridden by TradingStyle in V1.2+.',
   '0.65','Keep default. Overridden by TradingStyle in V1.2+ anyway.',
   '0.65','Keep default.',
   'Overridden by TradingStyle in V1.2+'),

  ('ICT TWINS MODEL','OTEMaxPercent','0-1','0.75','0.75','0.75','0.75','0.75','0.75',
   'OTE zone maximum as % of H1 swing range.\n'
   '0.75 = creates classic 65-75% OTE window (10% wide).\n'
   'V1.3 SmartActive widens to 60-85%; V1.4/V1.5 SmartActive+ widens to 55-90%.',
   '0.75','Keep default. Overridden by TradingStyle.',
   '0.75','Keep default.',
   'Overridden by TradingStyle in V1.2+'),

  ('ICT TWINS MODEL','OTESweetSpotPercent','0-1','0.705','0.705','0.705','0.705','0.705','0.705',
   'The golden ratio point within the OTE zone (70.5%).\n'
   'Entries very close to 70.5% receive bonus quality score points.',
   '0.705','Keep default.',
   '0.705','Keep default.',
   ''),

  ('ICT TWINS MODEL','MinFVGsRequired','int','0','0','0','0','0','0',
   'Minimum 1-minute Fair Value Gaps required near entry.\n'
   '0 = FVG check disabled (recommended — FVGs are plentiful when setup is valid).',
   '0','Keep 0 — FVG requirement with higher values reduces frequency too much.',
   '0','Keep 0.',
   ''),

  ('ICT TWINS MODEL','HTFToleranceATRMulti','x','2','2','2','2','2','2',
   'Price must be within N × ATR(14) of an HTF level to qualify as "at the level".\n'
   '2 = within 2 ATRs of the reference level.',
   '2','Default 2 ATRs is balanced.',
   '2','Keep default.',
   ''),

  ('ICT TWINS MODEL','HTFLevelRequired','bool','false','false','false','false','false','false',
   'Require price to be near a high-timeframe reference level before MSS step.\n'
   'Default false (auto-pass Step 1).\n'
   'true = strictest mode. In trending markets EA may not trade for days.',
   'false','Keep false. Setting true on trending gold = EA stops trading.',
   'false','Keep false.',
   'Setting true on strong trends = EA will not trade'),

  ('ICT TWINS MODEL','ShowOTEZone','bool','true','true','true','true','true','true',
   'Draw OTE zone rectangle on chart.\n'
   'Shows exactly where EA expects price to enter for a valid trade.',
   'true','Keep ON for visual confirmation.',
   'true','Keep ON.',
   ''),

  ('ICT TWINS MODEL','MinH1RangePips','pips','50','50','50','50','50','50',
   'Minimum H1 swing range (high to low) in pips.\n'
   'Filters out sideways/choppy markets where the swing is too small.\n'
   '0 = no minimum.',
   '30','Relax to 30 pips on $50 — 50 pip default was too strict, reducing frequency.',
   '40','40 pips balances quality vs frequency for $100.',
   ''),

  ('ICT TWINS MODEL','UseH1RangeFilter','bool','—','—','true','true','true','true',
   'Toggle the H1 range minimum pip check.\n'
   'false = skip MinH1RangePips check entirely.',
   'true','Keep ON.',
   'true','Keep ON.',
   'Not in V1.0 and V1.1'),

  ('ICT TWINS MODEL','MinH1RangeATRMulti','x','—','—','—','0.8','0.8','0.8',
   'H1 swing range must be ≥ (H1 ATR × this multiplier).\n'
   'Active in STYLE_SMART_ACTIVE and STYLE_SMART_ACTIVE_PLUS.\n'
   'Replaces the fixed pip check with a volatility-adaptive check.\n'
   '0.8 = range ≥ 80% of one H1 ATR period.',
   '0.8','Keep default. Filters extreme chop while allowing normal volatility.',
   '0.7','Relax to 0.7 for slightly more trades on $100.',
   'Only in V1.3, V1.4, V1.5'),

  # ── MSS / BOS / LIQUIDITY ──────────────────────────────────────────
  ('MSS / BOS / LIQUIDITY','MSSLookbackBars','bars','—','30','30','30','30','30',
   'H1 bars to scan backward when detecting Market Structure Shift.\n'
   '30 H1 bars ≈ 30 hours of history.',
   '30','Keep default.',
   '30','Keep default.',
   'Not in V1.0'),

  ('MSS / BOS / LIQUIDITY','MSSConfirmBars','bars','—','3','3','3','3','3',
   'Bars each side to confirm an H1 swing for MSS detection.\n'
   'In V1.3+ overridden to: 2 (SmartActive), 1 (SmartActive+) by style.',
   '3','Default — overridden by TradingStyle anyway.',
   '3','Keep default.',
   'Overridden by effMSSConfirm in V1.3+'),

  ('MSS / BOS / LIQUIDITY','BOSLookbackBars','bars','—','20','20','20','20','20',
   'M15 bars to scan when detecting Break of Structure.\n'
   '20 M15 bars ≈ 5 hours of M15 history.',
   '20','Keep default.',
   '20','Keep default.',
   'Not in V1.0'),

  ('MSS / BOS / LIQUIDITY','LiquidityLookbackBars','bars','—','50','50','50','50','50',
   'M15 bars to scan for liquidity levels (swing highs/lows).\n'
   'V1.5: these bars are now used to find actual swing highs/lows as pools\n'
   '(was incorrectly using range min/max in V1.0-V1.4).',
   '50','Keep default.',
   '50','Keep default.',
   '[V1.5 FIX] Now correctly scans for swing-based pools, not range extremes'),

  ('MSS / BOS / LIQUIDITY','LiquidityWickPips','pips','—','3','3','3','3','3',
   'Minimum wick size beyond a liquidity level to confirm a sweep.\n'
   'Wick ≥ 3 pips beyond swing level = liquidity sweep confirmed.',
   '3','Keep default 3 pips.',
   '3','Keep default.',
   'Not in V1.0'),

  # ── SYMBOL PRESET ──────────────────────────────────────────────────
  ('SYMBOL PRESET','SymbolPreset','Enum','—','PRESET_AUTO','PRESET_AUTO','PRESET_AUTO','PRESET_AUTO','PRESET_AUTO',
   'Pre-configured OTE and SL parameters for known symbols:\n'
   '• PRESET_AUTO: auto-detects from symbol name (recommended)\n'
   '• PRESET_XAUUSD: Gold — OTE 65-75%, MaxSL 30 pips\n'
   '• PRESET_BTCUSD: Bitcoin — OTE 62-78%, MaxSL 80 pips\n'
   '• PRESET_EURUSD: Euro — OTE 62-79%, MaxSL 25 pips\n'
   '• PRESET_GBPUSD: Pound — OTE 62-79%, MaxSL 30 pips',
   'PRESET_AUTO','Auto detects XAUUSD from symbol name.',
   'PRESET_AUTO','Keep auto.',
   'Not in V1.0'),

  # ── OPTIMIZATION MODE ──────────────────────────────────────────────
  ('OPTIMIZATION MODE','OptMode','Enum','—','OPT_BALANCED','OPT_BALANCED','OPT_BALANCED','OPT_BALANCED','OPT_BALANCED',
   'Secondary layer applied after TradingStyle (V1.2+) or standalone (V1.1):\n'
   '• OPT_CONSERVATIVE: score +10, risk 0.25%\n'
   '• OPT_BALANCED: no change from defaults\n'
   '• OPT_AGGRESSIVE: score -15, risk ×2',
   'OPT_CONSERVATIVE','Extra score filtering helps on $50.',
   'OPT_BALANCED','Balanced is fine for $100.',
   'Not in V1.0. In V1.2+ TradingStyle takes precedence'),

  # ── LOGGING ────────────────────────────────────────────────────────
  ('LOGGING','EnableScreenshot','bool','—','true','true','true','true','true',
   'Save a chart screenshot on every trade open and close.\n'
   'Files saved to MT5/MQL5/Files/ folder.',
   'true','Keep ON for trade journal.',
   'true','Keep ON.',
   'Not in V1.0'),

  ('LOGGING','EnableCSVLog','bool','—','true','true','true','true','true',
   'Export every trade to a CSV file.\n'
   'Columns: time, symbol, direction, entry, SL, TP, lot, score, result.',
   'true','Keep ON — CSV lets you analyze patterns.',
   'true','Keep ON.',
   'Not in V1.0'),

  # ── DEBUG / ADVANCED ───────────────────────────────────────────────
  ('DEBUG / ADVANCED','PostTradeCooldownMin','min','30','30','20','20','20','20',
   'Minutes to wait after closing a trade before next entry.\n'
   'Prevents back-to-back trades in the same move.\n'
   'Overridden by TradingStyle: Conservative=30, Balanced=20, SmartActive=10, SmartActive+=5.',
   '30','30 min on V1.0-V1.1; let TradingStyle control on V1.2+.',
   '20','20 min default on V1.2+.',
   'Overridden by TradingStyle in V1.2+'),

  ('DEBUG / ADVANCED','UseDailyTrendFilter','bool','true','true','true','true','true','true',
   'Only trade in the direction of the D1 trend.\n'
   'V1.5 FIX: H4 EMA handle now cached in OnInit (was created+released every tick).\n'
   'Uses D1 candle direction + H4 50-EMA alignment.\n'
   'Overridden OFF by STYLE_AGGRESSIVE.',
   'true','Keep ON — trading with D1 trend is core ICT principle.',
   'true','Keep ON.',
   '[V1.5 FIX] H4 EMA handle cached — was leaking memory every tick'),

  ('DEBUG / ADVANCED','BestHoursOnly','bool','true','true','true','true','true','true',
   'Restrict trading to 08:30-15:00 GMT.\n'
   'These are the highest-quality ICT entry hours.\n'
   'Overridden to false by STYLE_SMART_ACTIVE_PLUS for all-session coverage.',
   'true','Keep ON for $50 — focus on best-quality hours.',
   'true','Keep ON unless using SmartActive+.',
   'Overridden to false by STYLE_SMART_ACTIVE_PLUS'),

  ('DEBUG / ADVANCED','ForceTrades','bool','false','false','false','false','false','false',
   '⚠ TESTING ONLY: bypasses ALL filters and forces trade entries.\n'
   'Used in development to verify order execution.\n'
   'NEVER set to true on live accounts or real backtests.',
   'false','ALWAYS false. Never change this on live.',
   'false','ALWAYS false.',
   '⚠ DANGER: Setting true = random trades with no logic applied'),

  ('DEBUG / ADVANCED','DebugMode','bool','false','false','false','false','false','false',
   'Enable verbose logging to MT5 journal.\n'
   'Prints step-by-step logic for every tick.\n'
   'Useful for diagnosing why EA is not trading.',
   'false','Keep OFF in production. Enable temporarily to diagnose issues.',
   'false','Keep OFF.',
   'Generates massive journal output — can slow Strategy Tester significantly'),

  ('DEBUG / ADVANCED','RelaxedMode','bool','false','false','false','false','false','false',
   'Testing mode applying relaxed parameters.\n'
   'Only for verifying setup works in Strategy Tester.\n'
   'Not recommended for real backtests or live trading.',
   'false','Always false in production.',
   'false','Always false.',
   ''),
]

# ── STYLE TABLE (used in Style Guide sheet) ─────────────────────────
SGDATA = [
  ('V1.0','(no preset)','Manual','Manual','Manual','Manual','Manual','65-75%','70','30 min','~1-5/day'),
  ('V1.1','(no preset)','Manual','Manual','Manual','Manual','Manual','65-75%','70','30 min','~1-5/day'),
  ('V1.2','Conservative','ON','ON','ON','Manual','ON','65-75%','80','30 min','1-5/day'),
  ('V1.2','Balanced','Input','Input','Input','Input','Input','65-75%','70','20 min','3-8/day'),
  ('V1.2','Aggressive','OFF','Input','OFF','OFF','Input','55-90%','50','10 min','8-15/day'),
  ('V1.2','Ultra Active','OFF','OFF','OFF','OFF','OFF','40-95%','35','5 min','15-25/day'),
  ('V1.3','Conservative','ON','ON','ON','Manual','ON','65-75%','80','30 min','1-5/day'),
  ('V1.3','Balanced','Input','Input','Input','Input','Input','65-75%','70','20 min','3-8/day'),
  ('V1.3','Aggressive','OFF','Input','OFF','OFF','Input','55-90%','50','10 min','8-15/day'),
  ('V1.3','Smart Active ★','ON','ON','ON','Optional','ATR-adaptive','60-85%','55','10 min','8-15/day'),
  ('V1.3','Ultra Active','OFF','OFF','OFF','OFF','OFF','40-95%','35','5 min','15-25/day'),
  ('V1.4','Conservative','ON','ON','ON','Manual','ON','65-75%','80','30 min','1-5/day'),
  ('V1.4','Balanced','Input','Input','Input','Input','Input','65-75%','70','20 min','3-8/day'),
  ('V1.4','Aggressive','OFF','Input','OFF','OFF','Input','55-90%','50','10 min','8-15/day'),
  ('V1.4','Smart Active ★','ON','ON','ON','Optional','ATR-adaptive','60-85%','55','10 min','8-15/day'),
  ('V1.4','Smart Active+ ★★','ON','ON','ON','Optional','ATR-adaptive','55-90%','45','5 min','15-25/day'),
  ('V1.4','Ultra Active','OFF','OFF','OFF','OFF','OFF','40-95%','35','5 min','15-25/day'),
  # ── V1.5 rows ──────────────────────────────────────────────────────
  ('V1.5','Conservative','ON','ON','ON','Manual','ON','65-75%','80','30 min','1-5/day'),
  ('V1.5','Balanced','Input','Input','Input','Input','Input','65-75%','70','20 min','3-8/day'),
  ('V1.5','Aggressive','OFF','Input','OFF','OFF','Input','55-90%','50','10 min','8-15/day'),
  ('V1.5','Smart Active ★','ON','ON','ON','Optional','ATR-adaptive','60-85%','55','10 min','8-15/day'),
  ('V1.5','Smart Active+ ★★','ON','ON','ON','Optional','ATR-adaptive','55-90%','45','5 min','15-25/day'),
  ('V1.5','Ultra Active ✦','OFF','OFF','OFF','OFF','OFF','55-90%','OFF','5 min','15-25/day'),
]

# ═══════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1: README ─────────────────────────────────────────────────
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = "0D2137"
ws.column_dimensions['A'].width = 90

readme_lines = [
  ("ICT SMC EA — Complete Input Settings Guide", HDR, TG, True, 16),
  ("Versions: V1.0 · V1.1 · V1.2 · V1.3 · V1.4 · V1.5  |  Starting Capital: $50 and $100", HDR, TSL, False, 11),
  ("Created by: RATTANA CHHORM", HDR, TSL, False, 10),
  ("", HDR, TW, False, 10),
  ("HOW TO USE THIS WORKBOOK", GRP, TG, True, 12),
  ("", GRP, TW, False, 10),
  ("  Sheet: README           → This page. Overview and instructions.", ODD, TSL, False, 10),
  ("  Sheet: Style Guide      → TradingStyle preset comparison table.", ODD, TSL, False, 10),
  ("  Sheet: All Settings     → Full parameter reference — all versions side by side.", ODD, TSL, False, 10),
  ("  Sheet: $50 Quick Start  → Recommended settings for $50 starting capital.", ODD, TSL, False, 10),
  ("  Sheet: $100 Quick Start → Recommended settings for $100 starting capital.", ODD, TSL, False, 10),
  ("", ODD, TW, False, 10),
  ("ACCOUNT SIZE NOTES", GRP, TG, True, 12),
  ("", GRP, TW, False, 10),
  ("  $50 Account:", ODD, TB, True, 10),
  ("    • Use RiskMode = RISK_FIXED_LOT with FixedLot = 0.01 (minimum lot)", ODD, TSL, False, 10),
  ("    • MaxLotLimit = 0.01 — hard cap at minimum lot", ODD, TSL, False, 10),
  ("    • MaxDailyLossPercent = 6.0  ($3 max daily loss)", ODD, TSL, False, 10),
  ("    • MaxDailyLossTrades = 2  (stop after 2 losing trades per day)", ODD, TSL, False, 10),
  ("    • MaxTradesPerDay = 5  (conservative frequency)", ODD, TSL, False, 10),
  ("    • TradingStyle = STYLE_CONSERVATIVE  (V1.2+)", ODD, TSL, False, 10),
  ("    • MaxSLPips = 20, MinH1RangePips = 30", ODD, TSL, False, 10),
  ("    • XAUUSD: 0.01 lot × 20 pip SL ≈ $0.20 loss per trade (0.4% of $50)", ODD, TSL, False, 10),
  ("", EVN, TW, False, 10),
  ("  $100 Account:", EVN, TGR, True, 10),
  ("    • Use RiskMode = RISK_FIXED_PCT with RiskPercent = 1.0%  ($1 risk per trade)", EVN, TSL, False, 10),
  ("    • MaxLotLimit = 0.02", EVN, TSL, False, 10),
  ("    • MaxDailyLossPercent = 6.0  ($6 max daily loss)", EVN, TSL, False, 10),
  ("    • MaxDailyLossTrades = 3", EVN, TSL, False, 10),
  ("    • MaxTradesPerDay = 10  (balanced frequency)", EVN, TSL, False, 10),
  ("    • TradingStyle = STYLE_BALANCED  (V1.2+)", EVN, TSL, False, 10),
  ("    • MaxSLPips = 25, MinH1RangePips = 40", EVN, TSL, False, 10),
  ("", ODD, TW, False, 10),
  ("TRADING STYLE SUMMARY (V1.2+)", GRP, TG, True, 12),
  ("", GRP, TW, False, 10),
  ("  STYLE_CONSERVATIVE : score≥80 · OTE 65-75% · all ICT ON · CD 30min · 1-5 trades/day", ODD, TSL, False, 10),
  ("  STYLE_BALANCED     : score≥70 · OTE 65-75% · uses input toggles · CD 20min · 3-8 trades/day", ODD, TSL, False, 10),
  ("  STYLE_AGGRESSIVE   : score≥50 · OTE 55-90% · MSS/Trend OFF · CD 10min · 8-15 trades/day", EVN, TSL, False, 10),
  ("  STYLE_SMART_ACTIVE : score≥55 · OTE 60-85% · all ICT ON · CD 10min · 8-15 trades/day (V1.3+)", EVN, TGR, False, 10),
  ("  STYLE_SMART_ACTIVE+: score≥45 · OTE 55-90% · all ICT ON · CD 5min · 15-25 trades/day (V1.4+)", ODD, TB, False, 10),
  ("  STYLE_ULTRA_ACTIVE : score OFF · OTE 55-90% · filters minimal · CD 5min (V1.5: OTE fixed, score bypass)", ODD, TPU, False, 10),
  ("", ODD, TW, False, 10),
  ("V1.5 CHANGES SUMMARY", GRP, TG, True, 12),
  ("", GRP, TW, False, 10),
  ("  [FIX] H4 EMA handle cached in OnInit — was leaking a new handle every tick", ODD, TPU, False, 10),
  ("  [FIX] DetectLiquiditySweep now uses swing-based pools — was just finding range min/max", ODD, TPU, False, 10),
  ("  [FIX] CalculateTradeScore — disabled filters no longer give free points", ODD, TPU, False, 10),
  ("  [FIX] PartialClosePercent default 35% at 1.5R (was 50% at 1R — needs >67% WR to profit)", ODD, TPU, False, 10),
  ("  [FIX] UpdateDailyCounters preserves MSS/BOS/structural state at midnight", EVN, TPU, False, 10),
  ("  [FIX] UltraActive OTE tightened to 55-90% (was 40-95%) + score bypass", EVN, TPU, False, 10),
  ("  [NEW] UseDealingRange — optional filter: buys in discount, sells in premium only", EVN, TPU, False, 10),
  ("", ODD, TW, False, 10),
  ("RECOMMENDED VERSIONS", GRP, TG, True, 12),
  ("", GRP, TW, False, 10),
  ("  V1.0 → Base model. Only for reference — use V1.1+ for live.", ODD, TSL, False, 10),
  ("  V1.1 → Stable with MSS/BOS/Score/PartialTP. Good for conservative testing.", ODD, TSL, False, 10),
  ("  V1.2 → Adds TradingStyle presets and rejection statistics.", ODD, TSL, False, 10),
  ("  V1.3 → Improved detection algorithms. Use STYLE_SMART_ACTIVE for 8-15 trades/day.", ODD, TGR, False, 10),
  ("  V1.4 → Use STYLE_SMART_ACTIVE_PLUS for 15-25 trades/day with all ICT logic intact.", ODD, TB, False, 10),
  ("  V1.5 → Bug fixes + better math. Recommended for all new deployments.", NEW15, TPU, True, 10),
]

for i, (text, bg, fc, bold, sz) in enumerate(readme_lines, 1):
    cell = ws.cell(row=i, column=1, value=text)
    cell.fill = f(bg); cell.font = fn(fc, bold, sz)
    cell.alignment = al('left', 'center', False)
    ws.row_dimensions[i].height = 18

print("README sheet done.")

# ── Sheet 2: Style Guide ─────────────────────────────────────────────
ws2 = wb.create_sheet("Style Guide")
ws2.sheet_properties.tabColor = "1A3A1A"

SGCOLS = ['Version','Style','MSS','BOS','Trend','Liq Sweep','H1 Range','OTE Window','Min Score','Cooldown','Target Trades/Day']
SGWIDTHS = [8,22,7,7,7,12,10,14,10,12,18]
for i,(h,w) in enumerate(zip(SGCOLS,SGWIDTHS),1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    wc(ws2,1,i,h,HDR,TG,True,11,'center')
ws2.row_dimensions[1].height = 22

for r, row in enumerate(SGDATA, 2):
    ver = row[0]; style = row[1]
    if '★★' in style: bg = '0A1A2A'
    elif '★' in style: bg = '0A2A0A'
    elif 'Ultra' in style and '✦' in style: bg = '1A0A2A'   # V1.5 Ultra Active
    elif 'Ultra' in style: bg = '2A0A0A'
    elif 'Conservative' in style: bg = '1A1A2A'
    else: bg = ODD if r%2==0 else EVN
    for c, val in enumerate(row, 1):
        if '✦' in style:
            fc_ = TG if c==1 else TPU
        elif '★★' in style:
            fc_ = TG if c==1 else TB
        elif '★' in style:
            fc_ = TG if c==1 else TGR
        elif 'Ultra' in style:
            fc_ = TG if c==1 else TR
        else:
            fc_ = TG if c==1 else TSL
        if c in (3,4,5,6,7):
            fc_ = TGR if val=='ON' else (TR if val=='OFF' else TOR)
        wc(ws2, r, c, val, bg, fc_, bold=('★' in style or '✦' in style), h='center')
    ws2.row_dimensions[r].height = 18

# footer note
note_r = len(SGDATA)+3
wc(ws2, note_r, 1,
   '★ Smart Active = All ICT ON, better detection  |  ★★ Smart Active+ = All ICT ON + OTE 55-90%  |  ✦ V1.5 Ultra Active = OTE 55-90% fixed, score bypass',
   GRP, TB, False, 9, 'left', False, 10)
print("Style Guide sheet done.")

# ── Sheet 3: All Settings ────────────────────────────────────────────
ws3 = wb.create_sheet("All Settings")
ws3.sheet_properties.tabColor = "0D2137"

COL_WIDTHS3 = [20, 22, 8, 9, 9, 9, 9, 9, 9, 50, 22, 22, 35]
COL_HDRS3   = ['Group','Parameter','Type',
               'V1.0 Default','V1.1 Default','V1.2 Default','V1.3 Default','V1.4 Default','V1.5 Default',
               'What It Does','$50 Recommended','$100 Recommended','Notes / Warning']
for i,(h,w) in enumerate(zip(COL_HDRS3,COL_WIDTHS3),1):
    ws3.column_dimensions[get_column_letter(i)].width = w
    bg_ = NEW15 if h == 'V1.5 Default' else HDR
    fc_ = TPU if h == 'V1.5 Default' else TG
    wc(ws3,1,i,h,bg_,fc_,True,11,'center')
ws3.row_dimensions[1].height = 22
ws3.freeze_panes = 'A2'

cur_grp = None
r = 2
for row in P:
    grp,name,typ,v10,v11,v12,v13,v14,v15,what,f50_,fn50_,f100_,fn100_,warn = row
    if grp != cur_grp:
        for c in range(1,14):
            wc(ws3, r, c, grp if c==1 else '', GRP, TG, True, 10, 'left')
        ws3.row_dimensions[r].height = 18
        r += 1
        cur_grp = grp

    is_v15_new = (v14 == '—' and v15 != '—')
    is_v15_changed = (v14 != '—' and v15 != '—' and v14 != v15)
    row_bg = NEW15 if is_v15_new else (WARN if '⚠' in warn else (ODD if r%2==0 else EVN))

    wc(ws3,r,1,grp,row_bg,TSL,False,9)
    wc(ws3,r,2,name,row_bg,TW,True,10)
    wc(ws3,r,3,typ,row_bg,TSL,False,9,'center')
    for c_,val in zip(range(4,10),[v10,v11,v12,v13,v14,v15]):
        is_v15_col = (c_ == 9)
        if is_v15_col and is_v15_changed:
            fc_ = TPU; bg_c = NEW15
        elif is_v15_col and is_v15_new:
            fc_ = TPU; bg_c = NEW15
        elif val == '—':
            fc_ = TR; bg_c = row_bg
        elif val not in ('false','0','0.0'):
            fc_ = TGR; bg_c = row_bg
        else:
            fc_ = TSL; bg_c = row_bg
        cell = ws3.cell(row=r, column=c_, value=str(val))
        cell.fill = f(bg_c); cell.font = fn(fc_, False, 9)
        cell.alignment = al('center', 'center', True); cell.border = bd()
    wc(ws3,r,10,what,row_bg,TSL,False,9)
    wc(ws3,r,11,f50_,C50,TG,False,9)
    wc(ws3,r,12,f100_,C100,TGR,False,9)
    warn_text = warn if warn else '—'
    warn_bg = WARN if warn and '⚠' in warn else row_bg
    warn_fc = TR if warn and '⚠' in warn else (TPU if '[V1.5' in warn else TSL)
    wc(ws3,r,13,warn_text,warn_bg,warn_fc,False,9)
    ws3.row_dimensions[r].height = 60
    r += 1

print(f"All Settings sheet done. {r-2} parameter rows.")

# ── Sheet 4 & 5: Account Quick Start sheets ──────────────────────────
def build_account_sheet(wb, title, tab_color, account, amount):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    col_w = [20, 22, 8, 11, 11, 11, 11, 11, 11, 50, 26, 40]
    col_h = ['Group','Parameter','Type','V1.0','V1.1','V1.2','V1.3','V1.4','V1.5','What It Does',
             f'Recommended ({account})','Why This Setting']
    for i,(h,w) in enumerate(zip(col_h,col_w),1):
        ws.column_dimensions[get_column_letter(i)].width = w
        bg_ = NEW15 if h == 'V1.5' else HDR
        fc_ = TPU if h == 'V1.5' else TG
        wc(ws,1,i,h,bg_,fc_,True,11,'center')
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    ws.merge_cells('A2:L2')
    intro = (f"$50 Account Settings  |  Risk: Fixed 0.01 lot  |  Max daily loss: $3 (6%)  |  Style: CONSERVATIVE (V1.2+)  |  Max trades/day: 5" if amount==50
             else f"$100 Account Settings  |  Risk: 1% per trade (~$1)  |  Max daily loss: $6 (6%)  |  Style: BALANCED (V1.2+)  |  Max trades/day: 10")
    wc(ws,2,1,intro,C50H if amount==50 else C100H, TG,True,11,'center',False,11)
    ws.row_dimensions[2].height = 22

    cur_grp = None
    r = 3
    for row in P:
        grp,name,typ,v10,v11,v12,v13,v14,v15,what,f50_,fn50_,f100_,fn100_,warn = row
        rec = f50_ if amount==50 else f100_
        note = fn50_ if amount==50 else fn100_

        if grp != cur_grp:
            for c in range(1,13):
                wc(ws, r, c, grp if c==1 else '', GRP, TG, True, 10)
            ws.row_dimensions[r].height = 18
            r += 1
            cur_grp = grp

        row_bg = WARN if '⚠' in warn else (ODD if r%2==0 else EVN)
        wc(ws,r,1,grp,row_bg,TSL,False,9)
        wc(ws,r,2,name,row_bg,TW,True,10)
        wc(ws,r,3,typ,row_bg,TSL,False,9,'center')
        for c_,val in zip(range(4,10),[v10,v11,v12,v13,v14,v15]):
            is_v15_col = (c_ == 9)
            is_v15_changed = is_v15_col and v14 != '—' and v15 != '—' and v14 != v15
            is_v15_new = is_v15_col and v14 == '—' and v15 != '—'
            if is_v15_changed or is_v15_new:
                fc_ = TPU; bg_c = NEW15
            elif val == '—':
                fc_ = TR; bg_c = row_bg
            else:
                fc_ = TSL; bg_c = row_bg
            cell = ws.cell(row=r, column=c_, value=str(val))
            cell.fill = f(bg_c); cell.font = fn(fc_, False, 9)
            cell.alignment = al('center','center',True); cell.border = bd()
        wc(ws,r,10,what,row_bg,TSL,False,9)
        rc_bg = C50H if amount==50 else C100H
        rc_fc = TG if amount==50 else TGR
        wc(ws,r,11,rec,rc_bg,rc_fc,True,10,'center')
        wc(ws,r,12,note,row_bg,TSL,False,9)
        ws.row_dimensions[r].height = 55
        r += 1
    print(f"{title} sheet done.")

build_account_sheet(wb, "$50 Quick Start", "252500", "$50", 50)
build_account_sheet(wb, "$100 Quick Start", "002500", "$100", 100)

# ── Sheet 6: Version Comparison ──────────────────────────────────────
ws5 = wb.create_sheet("Version Comparison")
ws5.sheet_properties.tabColor = "16213E"

vc_params = [
  ('Feature','V1.0','V1.1','V1.2','V1.3','V1.4','V1.5'),
  ('EA Name','ICT SMC EA V1.0','ICT SMC EA V1.1','ICT SMC EA V1.2','ICT SMC EA V1.3','ICT SMC EA V1.4','ICT SMC EA V1.5'),
  ('TradingStyle Presets','None','None','4 styles','5 styles','6 styles','6 styles'),
  ('MSS Filter','No','Yes','Yes (style)','Yes (style)','Yes (style)','Yes (style)'),
  ('BOS Filter','No','Yes','Yes (style)','Yes (style)','Yes (style)','Yes (style)'),
  ('Liquidity Sweep','No','Optional','Optional','Optional','Optional','Optional (FIXED)'),
  ('Trade Quality Score','No','Yes (≥70)','Yes (toggle)','Yes (toggle)','Yes (toggle)','Yes (FIXED)'),
  ('Partial Take Profit','No','50%@1R','50%@1R','50%@1R','50%@1R','35%@1.5R (FIXED)'),
  ('TP Mode Selection','No','Yes','Yes','Yes','Yes','Yes'),
  ('Auto GMT Detection','No','Yes','Yes','Yes','Yes','Yes'),
  ('Symbol Presets','No','Yes','Yes','Yes','Yes','Yes'),
  ('CSV + Screenshot Log','No','Yes','Yes','Yes','Yes','Yes'),
  ('News Filter','No','Yes','Yes','Yes','Yes','Yes'),
  ('Filter Rejection Stats','No','No','Yes (daily)','Yes (daily+cumul)','Yes (daily+cumul)','Yes (daily+cumul)'),
  ('PrintFilterSummary','No','No','No','Yes (deinit)','Yes (deinit)','Yes (deinit)'),
  ('ATR-Adaptive H1 Range','No','No','No','Yes (SmartActive)','Yes (SA+)','Yes (SA+)'),
  ('H4 EMA Handle','Inline','Inline','Inline','Inline (LEAK)','Inline (LEAK)','Cached (FIXED)'),
  ('Daily State Reset','Full reset','Full reset','Full reset','Full reset','Full reset','Trade counters only (FIXED)'),
  ('UltraActive OTE','—','—','40-95%','40-95%','40-95%','55-90% (FIXED)'),
  ('Dealing Range Filter','No','No','No','No','No','Yes (optional)'),
  ('OTE Default Window','65-75%','65-75%','65-75%','65-75% (SA:60-85%)','65-75% (SA+:55-90%)','65-75% (SA+:55-90%)'),
  ('Default Style','-','-','BALANCED','SMART_ACTIVE','SMART_ACTIVE','SMART_ACTIVE'),
  ('Target Trades/Day (def)','1-5','1-5','3-8','8-15','8-15','8-15'),
  ('Magic Number','888777','888777','888777','888777','888777','888777'),
  ('Recommended Use','Reference only','Conservative test','Balanced live','Smart active live','Max frequency live','All new deployments'),
]

vc_widths = [28,20,20,20,20,20,22]
for i,w in enumerate(vc_widths,1):
    ws5.column_dimensions[get_column_letter(i)].width = w

for r, row in enumerate(vc_params,1):
    is_header = r == 1
    for c, val in enumerate(row,1):
        is_v15_col = (c == 7)
        if is_header:
            bg_ = NEW15 if is_v15_col else HDR
            fc_ = TPU if is_v15_col else TG
            bold_ = True; sz_ = 11
        elif is_v15_col:
            bg_ = NEW15; fc_ = TPU if 'FIXED' in str(val) or 'optional' in str(val).lower() else TW
            bold_ = False; sz_ = 10
        elif r % 2 == 0:
            bg_ = ODD; fc_ = TW if c==1 else TSL; bold_ = c==1; sz_ = 10
        else:
            bg_ = EVN; fc_ = TW if c==1 else TSL; bold_ = c==1; sz_ = 10
        # highlight V1.3/V1.4 new features in those columns
        if c in (4,5) and not is_header and 'Yes' in str(val):
            if vc_params[r-1][c-2] in ('No','—','-',''):
                bg_ = NEW; fc_ = TGR
        wc(ws5, r, c, val, bg_, fc_, bold_, sz_, 'center' if c>1 else 'left')
    ws5.row_dimensions[r].height = 22

print("Version Comparison sheet done.")

# ── Save ─────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\nSaved: {OUT}")
